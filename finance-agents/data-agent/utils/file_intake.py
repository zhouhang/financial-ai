"""Shared file-intake helpers for raw uploads and logical sheet files."""

from __future__ import annotations

import csv
import hashlib
import logging
import re
from pathlib import Path
from typing import Any

from config import UPLOAD_DIR

logger = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = {".xlsx", ".xls", ".csv"}


def _normalize_upload_ref(file_path: str, upload_root: Path) -> str:
    """Normalize uploaded file path to /uploads/... ref when possible."""
    path_str = str(file_path or "").strip()
    if not path_str:
        return ""
    if path_str.startswith("/uploads/"):
        return path_str
    if path_str.startswith("uploads/"):
        return f"/{path_str}"

    path_obj = Path(path_str)
    try:
        if path_obj.is_absolute():
            rel = path_obj.resolve().relative_to(upload_root)
            return f"/uploads/{rel.as_posix()}"
    except Exception:
        return path_str
    return path_str


def _resolve_upload_abs_path(file_path: str, upload_root: Path) -> Path:
    """Resolve upload ref or upload-root absolute path to a local absolute path."""
    path_str = str(file_path or "").strip()
    if not path_str:
        raise ValueError("文件路径不能为空")

    if path_str.startswith("/uploads/"):
        rel = path_str.lstrip("/")[len("uploads/"):]
        resolved = (upload_root / rel).resolve()
        resolved.relative_to(upload_root)
        return resolved

    path_obj = Path(path_str)
    if path_obj.is_absolute():
        resolved = path_obj.resolve()
        resolved.relative_to(upload_root)
        return resolved

    rel = path_str.lstrip("/")
    if not rel.startswith("uploads/"):
        raise ValueError(f"非法上传文件路径: {file_path}")
    resolved = (upload_root / rel[len("uploads/"):]).resolve()
    resolved.relative_to(upload_root)
    return resolved


def _coerce_cell_text(value: Any, *, strip: bool = True) -> str:
    text = "" if value is None else str(value)
    return text.strip() if strip else text


def _row_has_values(row: list[Any] | tuple[Any, ...] | None) -> bool:
    if not row:
        return False
    return any(_coerce_cell_text(value) != "" for value in row)


def _sanitize_filename_component(value: str, fallback: str, *, max_length: int = 48) -> str:
    text = str(value or "").strip()
    text = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", text)
    text = re.sub(r"\s+", "_", text)
    text = text.strip(" ._")
    if not text:
        text = fallback
    if len(text) > max_length:
        text = text[:max_length].rstrip(" ._") or fallback
    return text


def _build_file_fingerprint(upload_ref: str, abs_path: Path) -> str:
    seed = upload_ref or str(abs_path)
    return hashlib.sha1(seed.encode("utf-8")).hexdigest()[:8]


def _normalize_column_name(col_name: str, config: dict[str, Any]) -> str:
    normalized = str(col_name).strip()
    if config.get("ignore_whitespace", True):
        normalized = normalized.replace(" ", "").replace("\t", "")
    if not config.get("case_sensitive", False):
        normalized = normalized.lower()
    return normalized


def _normalize_columns_set(columns: list[str], config: dict[str, Any]) -> set[str]:
    return {_normalize_column_name(col, config) for col in columns}


def _build_alias_mapping(table_schema: dict[str, Any], config: dict[str, Any]) -> dict[str, str]:
    alias_map: dict[str, str] = {}
    for original_col, aliases in (table_schema.get("column_aliases") or {}).items():
        normalized_original = _normalize_column_name(str(original_col), config)
        for alias in aliases or []:
            alias_map[_normalize_column_name(str(alias), config)] = normalized_original
    return alias_map


def _normalize_file_columns(
    file_columns: list[str],
    table_schema: dict[str, Any],
    config: dict[str, Any],
) -> set[str]:
    alias_map = _build_alias_mapping(table_schema, config)
    normalized_set: set[str] = set()
    for col in file_columns:
        normalized_col = _normalize_column_name(str(col), config)
        normalized_set.add(alias_map.get(normalized_col, normalized_col))
    return normalized_set


def _schema_candidate_names(
    *,
    file_name: str,
    columns: list[str],
    validation_rules: dict[str, Any],
) -> list[str]:
    config = validation_rules.get("validation_config") or {}
    table_schemas = validation_rules.get("table_schemas") or []
    if not table_schemas:
        return []

    file_ext = Path(file_name).suffix.lower().lstrip(".")
    candidate_names: list[str] = []
    for table_schema in table_schemas:
        required_columns = table_schema.get("required_columns") or []
        required_set = _normalize_columns_set([str(col) for col in required_columns], config)
        file_set = _normalize_file_columns([str(col) for col in columns], table_schema, config)
        missing_required = required_set - file_set
        allowed_file_types = [str(item).lower().lstrip(".") for item in table_schema.get("file_type", [])]
        file_type_match = (not allowed_file_types) or (file_ext in allowed_file_types)
        if not missing_required and file_type_match:
            candidate_names.append(str(table_schema.get("table_name") or "").strip())
    return [name for name in candidate_names if name]


def _read_csv_header_and_rows(file_path: Path) -> tuple[list[str], bool]:
    with open(file_path, newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        header_row = next(reader, [])
        header = [_coerce_cell_text(cell) for cell in header_row]
        has_data_rows = any(_row_has_values(row) for row in reader)
    return header, has_data_rows


def _analyze_openpyxl_sheet(sheet: Any) -> tuple[list[str], bool]:
    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), tuple())
    header = [_coerce_cell_text(cell) for cell in first_row]
    has_data_rows = any(_row_has_values(row) for row in sheet.iter_rows(min_row=2, values_only=True))
    return header, has_data_rows


def _build_split_directory(source_abs_path: Path, fingerprint: str) -> Path:
    safe_stem = _sanitize_filename_component(source_abs_path.stem, "workbook")
    return source_abs_path.parent / f"{safe_stem}__split_{fingerprint}"


def _build_split_display_name(
    *,
    workbook_name: str,
    upload_ref: str,
    abs_path: Path,
    sheet_name: str,
    sheet_index: int,
    display_extension: str,
) -> str:
    fingerprint = _build_file_fingerprint(upload_ref, abs_path)
    workbook_stem = _sanitize_filename_component(Path(workbook_name).stem, "workbook")
    sheet_stem = _sanitize_filename_component(sheet_name, f"sheet_{sheet_index}")
    ext = display_extension if display_extension.startswith(".") else f".{display_extension}"
    return f"{workbook_stem}__{fingerprint}__s{sheet_index:02d}__{sheet_stem}{ext}"


def _build_split_storage_path(
    *,
    source_abs_path: Path,
    display_name: str,
    upload_ref: str,
) -> Path:
    fingerprint = _build_file_fingerprint(upload_ref, source_abs_path)
    split_dir = _build_split_directory(source_abs_path, fingerprint)
    split_dir.mkdir(parents=True, exist_ok=True)
    safe_name = _sanitize_filename_component(Path(display_name).stem, "sheet")
    return split_dir / f"{safe_name}.xlsx"


def _write_openpyxl_sheet_copy(source_sheet: Any, sheet_name: str, target_path: Path) -> None:
    import openpyxl

    target_wb = openpyxl.Workbook(write_only=True)
    target_ws = target_wb.create_sheet(title=sheet_name[:31] or "Sheet1")
    for row in source_sheet.iter_rows(values_only=True):
        target_ws.append(list(row))
    target_wb.save(target_path)


def _write_pandas_sheet_copy(frame: Any, sheet_name: str, target_path: Path) -> None:
    import pandas as pd

    df = pd.DataFrame(frame)
    with pd.ExcelWriter(target_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, header=False, sheet_name=(sheet_name[:31] or "Sheet1"))


def _build_logical_file_entry(
    *,
    file_path: str,
    display_name: str,
    workbook_original_filename: str,
    workbook_display_name: str,
    workbook_file_path: str,
    sheet_name: str | None,
    sheet_index: int | None,
    is_logical_split: bool,
) -> dict[str, Any]:
    return {
        "file_path": file_path,
        # Keep original_filename/display_name aligned for downstream compatibility.
        "original_filename": display_name,
        "display_name": display_name,
        "name": display_name,
        "workbook_original_filename": workbook_original_filename,
        "workbook_display_name": workbook_display_name,
        "workbook_file_path": workbook_file_path,
        "sheet_name": sheet_name,
        "sheet_index": sheet_index,
        "is_logical_split": is_logical_split,
    }


def _build_prefilter_decision(
    *,
    logical_file: dict[str, Any],
    columns: list[str],
    has_data_rows: bool,
    validation_rules: dict[str, Any],
) -> dict[str, Any]:
    candidate_table_names = _schema_candidate_names(
        file_name=str(logical_file.get("display_name") or logical_file.get("original_filename") or ""),
        columns=columns,
        validation_rules=validation_rules,
    ) if validation_rules else []

    has_header = any(str(col).strip() for col in columns)
    reason_code = ""
    reason = ""
    status = "kept"

    if not has_header:
        status = "dropped"
        reason_code = "empty_header"
        reason = "首行为空，无法识别表头"
    elif not has_data_rows:
        status = "dropped"
        reason_code = "no_data_rows"
        reason = "只有表头没有数据行"
    elif validation_rules and not candidate_table_names:
        status = "dropped"
        reason_code = "no_schema_candidate"
        reason = "按当前规则无法命中任何 schema"

    summary = {
        "display_name": logical_file.get("display_name") or logical_file.get("original_filename") or "",
        "file_path": logical_file.get("file_path") or "",
        "workbook_display_name": logical_file.get("workbook_display_name") or "",
        "workbook_original_filename": logical_file.get("workbook_original_filename") or "",
        "sheet_name": logical_file.get("sheet_name"),
        "sheet_index": logical_file.get("sheet_index"),
        "status": status,
        "reason_code": reason_code,
        "reason": reason,
        "candidate_table_names": candidate_table_names,
        "columns": list(columns),
        "is_logical_split": bool(logical_file.get("is_logical_split")),
    }
    return summary


def _prepare_csv_logical_entry(entry: dict[str, Any], validation_rules: dict[str, Any]) -> tuple[dict[str, Any], dict[str, Any]]:
    abs_path = entry["abs_path"]
    display_name = entry["display_name"]
    header, has_data_rows = _read_csv_header_and_rows(abs_path)
    logical_file = _build_logical_file_entry(
        file_path=entry["upload_ref"],
        display_name=display_name,
        workbook_original_filename=entry["original_filename"],
        workbook_display_name=display_name,
        workbook_file_path=entry["upload_ref"],
        sheet_name=None,
        sheet_index=None,
        is_logical_split=False,
    )
    summary = _build_prefilter_decision(
        logical_file=logical_file,
        columns=header,
        has_data_rows=has_data_rows,
        validation_rules=validation_rules,
    )
    return logical_file, summary


def _prepare_excel_logical_entries(entry: dict[str, Any], validation_rules: dict[str, Any]) -> list[tuple[dict[str, Any], dict[str, Any]]]:
    abs_path = entry["abs_path"]
    display_name = entry["display_name"]
    upload_ref = entry["upload_ref"]
    extension = entry["extension"]

    if extension == ".xlsx":
        import openpyxl

        wb = openpyxl.load_workbook(abs_path, read_only=True, data_only=True)
        try:
            sheet_names = list(wb.sheetnames)
            split_required = len(sheet_names) > 1
            results: list[tuple[dict[str, Any], dict[str, Any]]] = []
            for sheet_index, sheet_name in enumerate(sheet_names, start=1):
                source_sheet = wb[sheet_name]
                header, has_data_rows = _analyze_openpyxl_sheet(source_sheet)
                if split_required:
                    split_display_name = _build_split_display_name(
                        workbook_name=display_name,
                        upload_ref=upload_ref,
                        abs_path=abs_path,
                        sheet_name=sheet_name,
                        sheet_index=sheet_index,
                        display_extension=extension,
                    )
                    split_path = _build_split_storage_path(
                        source_abs_path=abs_path,
                        display_name=split_display_name,
                        upload_ref=upload_ref,
                    )
                    _write_openpyxl_sheet_copy(source_sheet, sheet_name, split_path)
                    file_path = _normalize_upload_ref(str(split_path), Path(UPLOAD_DIR).resolve())
                    logical_display_name = split_display_name
                else:
                    file_path = upload_ref
                    logical_display_name = display_name

                logical_file = _build_logical_file_entry(
                    file_path=file_path,
                    display_name=logical_display_name,
                    workbook_original_filename=entry["original_filename"],
                    workbook_display_name=display_name,
                    workbook_file_path=upload_ref,
                    sheet_name=sheet_name,
                    sheet_index=sheet_index,
                    is_logical_split=split_required,
                )
                summary = _build_prefilter_decision(
                    logical_file=logical_file,
                    columns=header,
                    has_data_rows=has_data_rows,
                    validation_rules=validation_rules,
                )
                results.append((logical_file, summary))
            return results
        finally:
            wb.close()

    import pandas as pd

    workbook = pd.read_excel(abs_path, sheet_name=None, header=None, dtype=object)
    split_required = len(workbook) > 1
    results = []
    for sheet_index, (sheet_name, frame) in enumerate(workbook.items(), start=1):
        rows = frame.values.tolist() if hasattr(frame, "values") else []
        header_row = rows[0] if rows else []
        header = [_coerce_cell_text(cell) for cell in header_row]
        has_data_rows = any(_row_has_values(row) for row in rows[1:])
        if split_required:
            split_display_name = _build_split_display_name(
                workbook_name=display_name,
                upload_ref=upload_ref,
                abs_path=abs_path,
                sheet_name=sheet_name,
                sheet_index=sheet_index,
                display_extension=extension,
            )
            split_path = _build_split_storage_path(
                source_abs_path=abs_path,
                display_name=split_display_name,
                upload_ref=upload_ref,
            )
            _write_pandas_sheet_copy(frame, sheet_name, split_path)
            file_path = _normalize_upload_ref(str(split_path), Path(UPLOAD_DIR).resolve())
            logical_display_name = split_display_name
        else:
            file_path = upload_ref
            logical_display_name = display_name

        logical_file = _build_logical_file_entry(
            file_path=file_path,
            display_name=logical_display_name,
            workbook_original_filename=entry["original_filename"],
            workbook_display_name=display_name,
            workbook_file_path=upload_ref,
            sheet_name=sheet_name,
            sheet_index=sheet_index,
            is_logical_split=split_required,
        )
        summary = _build_prefilter_decision(
            logical_file=logical_file,
            columns=header,
            has_data_rows=has_data_rows,
            validation_rules=validation_rules,
        )
        results.append((logical_file, summary))
    return results


def _normalize_uploaded_file_entry(
    item: Any,
    *,
    upload_root: Path,
) -> dict[str, Any] | None:
    if isinstance(item, dict):
        raw_path = str(item.get("file_path") or item.get("path") or "").strip()
        original_filename = str(item.get("original_filename") or item.get("name") or "").strip()
    else:
        raw_path = str(item or "").strip()
        original_filename = ""

    if not raw_path:
        return None

    upload_ref = _normalize_upload_ref(raw_path, upload_root)
    abs_path = _resolve_upload_abs_path(upload_ref or raw_path, upload_root)
    stored_name = abs_path.name
    display_name = original_filename or stored_name
    extension = abs_path.suffix.lower()

    return {
        "upload_ref": upload_ref or raw_path,
        "abs_path": abs_path,
        "stored_name": stored_name,
        "display_name": display_name,
        "original_filename": original_filename or display_name,
        "extension": extension,
    }


def build_upload_name_maps(raw_files: list[Any]) -> tuple[dict[str, str], dict[str, str]]:
    """Build filename/ref maps from either raw uploads or logical upload files."""
    upload_root = Path(UPLOAD_DIR).resolve()
    display_name_to_ref: dict[str, str] = {}
    ref_to_display_name: dict[str, str] = {}

    for item in raw_files:
        if isinstance(item, dict):
            file_path = str(item.get("file_path") or item.get("path") or "").strip()
            display_name = str(
                item.get("display_name")
                or item.get("name")
                or item.get("original_filename")
                or ""
            ).strip()
            original_filename = str(item.get("original_filename") or "").strip()
            is_logical_split = bool(item.get("is_logical_split"))
        else:
            file_path = str(item or "").strip()
            display_name = ""
            original_filename = ""
            is_logical_split = False

        if not file_path:
            continue

        upload_ref = _normalize_upload_ref(file_path, upload_root)
        try:
            abs_path = _resolve_upload_abs_path(upload_ref or file_path, upload_root)
        except ValueError:
            continue

        stored_name = abs_path.name
        final_display_name = display_name or stored_name

        if final_display_name:
            display_name_to_ref[final_display_name] = upload_ref or file_path
        if stored_name:
            display_name_to_ref[stored_name] = upload_ref or file_path
        if original_filename and not is_logical_split:
            display_name_to_ref[original_filename] = upload_ref or file_path

        if upload_ref:
            ref_to_display_name[upload_ref] = final_display_name
        ref_to_display_name[str(abs_path)] = final_display_name
        if stored_name:
            ref_to_display_name[stored_name] = final_display_name

    return display_name_to_ref, ref_to_display_name


def prepare_logical_upload_files(
    raw_files: list[Any],
    *,
    file_rule: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Expand raw uploads into logical files and apply sheet-level prefiltering."""
    upload_root = Path(UPLOAD_DIR).resolve()
    validation_rules = ((file_rule or {}).get("file_validation_rules") or {}) if file_rule else {}

    normalized_entries: list[dict[str, Any]] = []
    for item in raw_files:
        normalized = _normalize_uploaded_file_entry(item, upload_root=upload_root)
        if normalized is None:
            continue
        normalized_entries.append(normalized)

    kept_logical_files: list[dict[str, Any]] = []
    files_with_columns: list[dict[str, Any]] = []
    prefilter_summary: list[dict[str, Any]] = []

    for entry in normalized_entries:
        extension = entry["extension"]
        if extension not in SUPPORTED_EXTENSIONS:
            raise ValueError(f"不支持的文件类型：{extension}")

        if extension == ".csv":
            prepared_items = [_prepare_csv_logical_entry(entry, validation_rules)]
        else:
            prepared_items = _prepare_excel_logical_entries(entry, validation_rules)

        for logical_file, summary in prepared_items:
            prefilter_summary.append(summary)
            if summary["status"] != "kept":
                logger.info(
                    "[file_intake] 过滤逻辑文件 %s: %s (%s)",
                    summary.get("display_name"),
                    summary.get("reason"),
                    summary.get("reason_code"),
                )
                continue
            kept_logical_files.append(logical_file)
            files_with_columns.append(
                {
                    "file_name": str(logical_file.get("display_name") or logical_file.get("original_filename") or ""),
                    "columns": list(summary.get("columns") or []),
                }
            )

    return {
        "logical_uploaded_files": kept_logical_files,
        "files_with_columns": files_with_columns,
        "prefilter_summary": prefilter_summary,
        "kept_count": len(kept_logical_files),
        "dropped_count": sum(1 for item in prefilter_summary if item.get("status") == "dropped"),
        "source_file_count": len(normalized_entries),
    }
