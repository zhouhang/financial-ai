"""文件格式验证模块

用于对账流程的文件上传验证，判断是否为标准格式（2个文件，单sheet，header+data）。
标准格式直接走现有流程，非标准格式路由到智能分析。
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any
from dataclasses import dataclass

try:
    from openpyxl import load_workbook
except ImportError:
    import pandas as pd  # 降级使用 pandas

logger = logging.getLogger(__name__)


@dataclass
class FileValidationResult:
    """文件验证结果"""
    is_standard: bool  # 是否为标准格式
    reason: str | None  # 不标准的原因（如果 is_standard=False）
    file_count: int  # 文件数量
    total_sheet_count: int  # 总 sheet 数量
    details: list[dict[str, Any]]  # 每个文件的详细信息


def is_standard_format(file_paths: list[str | Path]) -> FileValidationResult:
    """检查上传的文件是否为标准格式

    标准格式定义：
    1. 恰好 2 个 Excel 文件
    2. 每个文件只有 1 个 sheet
    3. 每个 sheet 有表头行（第一行）
    4. 每个 sheet 至少有 1 行数据（除表头外）

    Args:
        file_paths: 上传的文件路径列表

    Returns:
        FileValidationResult 对象
    """
    # 检查文件数量
    file_count = len(file_paths)
    if file_count != 2:
        return FileValidationResult(
            is_standard=False,
            reason=f"文件数量不符：需要2个文件，实际上传了{file_count}个",
            file_count=file_count,
            total_sheet_count=0,
            details=[]
        )

    details = []
    total_sheet_count = 0

    for file_path in file_paths:
        file_path_str = str(file_path)

        # 转换为绝对路径（如果是相对路径 /uploads/... 则转换为 MCP 服务器的上传目录）
        if file_path_str.startswith('/uploads/'):
            # MCP 服务器的上传目录
            import os
            mcp_root = Path(__file__).parent.parent.parent.parent.parent / 'finance-mcp'
            file_path = mcp_root / file_path_str.lstrip('/')
        else:
            file_path = Path(file_path_str)

        # 检查文件扩展名
        file_ext = file_path.suffix.lower()
        if file_ext not in ['.xlsx', '.xls', '.csv']:
            return FileValidationResult(
                is_standard=False,
                reason=f"文件 '{file_path.name}' 格式不支持（仅支持 Excel 或 CSV）",
                file_count=file_count,
                total_sheet_count=0,
                details=details
            )

        # 检查文件是否存在
        if not file_path.exists():
            return FileValidationResult(
                is_standard=False,
                reason=f"文件不存在: {file_path}",
                file_count=file_count,
                total_sheet_count=0,
                details=details
            )

        try:
            # 分析文件（Excel 或 CSV）
            if file_ext == '.csv':
                file_info = _analyze_csv_file(file_path)
            else:
                file_info = _analyze_excel_file(file_path)

            details.append(file_info)
            total_sheet_count += file_info['sheet_count']

            # 检查 sheet 数量（CSV 文件视为只有 1 个 sheet）
            if file_info['sheet_count'] != 1:
                return FileValidationResult(
                    is_standard=False,
                    reason=f"文件 '{file_path.name}' 包含 {file_info['sheet_count']} 个 sheet（需要恰好1个）",
                    file_count=file_count,
                    total_sheet_count=total_sheet_count,
                    details=details
                )

            # 检查表头和数据
            sheet_info = file_info['sheets'][0]

            if not sheet_info['has_header']:
                return FileValidationResult(
                    is_standard=False,
                    reason=f"文件 '{file_path.name}' 缺少表头行",
                    file_count=file_count,
                    total_sheet_count=total_sheet_count,
                    details=details
                )

            if sheet_info['data_row_count'] == 0:
                return FileValidationResult(
                    is_standard=False,
                    reason=f"文件 '{file_path.name}' 没有数据行（只有表头）",
                    file_count=file_count,
                    total_sheet_count=total_sheet_count,
                    details=details
                )

        except Exception as e:
            logger.error(f"文件验证失败 {file_path}: {e}")
            return FileValidationResult(
                is_standard=False,
                reason=f"文件 '{file_path.name}' 解析失败: {str(e)}",
                file_count=file_count,
                total_sheet_count=total_sheet_count,
                details=details
            )

    # 所有检查通过
    return FileValidationResult(
        is_standard=True,
        reason=None,
        file_count=file_count,
        total_sheet_count=total_sheet_count,
        details=details
    )


def _analyze_excel_file(file_path: Path) -> dict[str, Any]:
    """分析单个 Excel 文件的结构

    Args:
        file_path: Excel 文件路径

    Returns:
        文件信息字典
    """
    try:
        # 优先使用 openpyxl（更快，不加载全部数据）
        wb = load_workbook(file_path, read_only=True, data_only=True)

        sheets_info = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_info = _analyze_sheet_openpyxl(sheet, sheet_name)
            sheets_info.append(sheet_info)

        wb.close()

        return {
            'filename': file_path.name,
            'filepath': str(file_path),
            'sheet_count': len(sheets_info),
            'sheets': sheets_info
        }

    except Exception as e:
        logger.warning(f"openpyxl 解析失败，尝试使用 pandas: {e}")
        # 降级到 pandas
        return _analyze_excel_file_pandas(file_path)


def _analyze_sheet_openpyxl(sheet, sheet_name: str) -> dict[str, Any]:
    """使用 openpyxl 分析 sheet

    Args:
        sheet: openpyxl worksheet 对象
        sheet_name: sheet 名称

    Returns:
        sheet 信息字典
    """
    # 获取所有行
    rows = list(sheet.iter_rows(values_only=True, max_row=10))  # 只读前10行用于快速验证

    if not rows:
        return {
            'sheet_name': sheet_name,
            'has_header': False,
            'data_row_count': 0,
            'headers': []
        }

    # 第一行作为表头
    header_row = rows[0]
    has_header = any(cell is not None and str(cell).strip() != '' for cell in header_row)

    headers = []
    if has_header:
        headers = [str(cell) if cell is not None else f"Column_{i}"
                  for i, cell in enumerate(header_row)]

    # 检查是否有数据行（第二行开始）
    data_rows = rows[1:]
    # 判断非空行（至少有一个非空单元格）
    data_row_count = sum(1 for row in data_rows
                        if any(cell is not None and str(cell).strip() != '' for cell in row))

    # 实际数据行数需要检查整个sheet，不能只看前10行
    # 但为了性能，我们使用 sheet.max_row 作为估计
    estimated_data_rows = max(0, sheet.max_row - 1) if sheet.max_row else 0

    return {
        'sheet_name': sheet_name,
        'has_header': has_header,
        'data_row_count': estimated_data_rows,  # 使用估计值
        'headers': headers
    }


def _analyze_csv_file(file_path: Path) -> dict[str, Any]:
    """分析 CSV 文件的结构

    Args:
        file_path: CSV 文件路径

    Returns:
        文件信息字典（格式与 _analyze_excel_file 一致）
    """
    import pandas as pd
    import chardet

    try:
        # 自动检测编码
        with open(file_path, 'rb') as f:
            raw_data = f.read()
            detected = chardet.detect(raw_data)
            encoding = detected.get('encoding', 'utf-8')

        # 读取 CSV 文件
        df = pd.read_csv(file_path, encoding=encoding, nrows=5)

        has_header = len(df.columns) > 0
        headers = list(df.columns) if has_header else []

        # 获取实际行数
        df_full = pd.read_csv(file_path, encoding=encoding)
        data_row_count = len(df_full)

        # CSV 文件视为只有 1 个 sheet
        sheet_info = {
            'sheet_name': 'CSV',
            'has_header': has_header,
            'data_row_count': data_row_count,
            'headers': headers
        }

        return {
            'filename': file_path.name,
            'filepath': str(file_path),
            'sheet_count': 1,  # CSV 文件视为只有 1 个 sheet
            'sheets': [sheet_info]
        }

    except Exception as e:
        logger.error(f"CSV 文件分析失败 {file_path}: {e}")
        raise


def _analyze_excel_file_pandas(file_path: Path) -> dict[str, Any]:
    """使用 pandas 分析 Excel 文件（降级方案）

    Args:
        file_path: Excel 文件路径

    Returns:
        文件信息字典
    """
    import pandas as pd

    # 读取所有 sheet 名称
    excel_file = pd.ExcelFile(file_path)
    sheet_names = excel_file.sheet_names

    sheets_info = []
    for sheet_name in sheet_names:
        try:
            # 只读取前几行用于快速验证
            df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=5)

            has_header = len(df.columns) > 0
            headers = list(df.columns) if has_header else []

            # 获取实际行数（重新读取，不限制行数，但只获取shape）
            df_full = pd.read_excel(file_path, sheet_name=sheet_name)
            data_row_count = len(df_full)

            sheets_info.append({
                'sheet_name': sheet_name,
                'has_header': has_header,
                'data_row_count': data_row_count,
                'headers': headers
            })

        except Exception as e:
            logger.error(f"Sheet '{sheet_name}' 分析失败: {e}")
            sheets_info.append({
                'sheet_name': sheet_name,
                'has_header': False,
                'data_row_count': 0,
                'headers': [],
                'error': str(e)
            })

    return {
        'filename': file_path.name,
        'filepath': str(file_path),
        'sheet_count': len(sheets_info),
        'sheets': sheets_info
    }


def detect_empty_sheets(file_paths: list[str | Path]) -> list[dict[str, Any]]:
    """检测空 sheet（没有数据行）

    Args:
        file_paths: 文件路径列表

    Returns:
        空 sheet 列表，每个元素包含 filename 和 sheet_name
    """
    empty_sheets = []

    for file_path in file_paths:
        file_path = Path(file_path)
        try:
            file_info = _analyze_excel_file(file_path)
            for sheet_info in file_info['sheets']:
                if sheet_info['data_row_count'] == 0:
                    empty_sheets.append({
                        'filename': file_info['filename'],
                        'sheet_name': sheet_info['sheet_name']
                    })
        except Exception as e:
            logger.error(f"检测空 sheet 失败 {file_path}: {e}")

    return empty_sheets


def detect_header_row(rows: list[tuple]) -> int:
    """检测表头行位置（区分表头和数据行）

    策略：
    1. 跳过全空行
    2. 找到第一个包含文本的行作为表头
    3. 检查下一行是否为数据（数字或日期）

    Args:
        rows: 行数据列表

    Returns:
        表头行的索引（0-based），如果找不到返回 -1
    """
    if not rows:
        return -1

    for i, row in enumerate(rows):
        # 跳过全空行
        if not any(cell is not None and str(cell).strip() != '' for cell in row):
            continue

        # 检查这一行是否像表头（主要包含文本）
        text_count = sum(1 for cell in row
                        if cell is not None and isinstance(cell, str) and cell.strip() != '')

        # 如果这一行大部分是文本，认为是表头
        if text_count > len(row) * 0.5:  # 50%以上是文本
            return i

    # 没找到明显的表头，返回第一行
    return 0


# 测试代码
if __name__ == "__main__":
    import sys

    logging.basicConfig(level=logging.INFO)

    if len(sys.argv) < 2:
        print("用法: python file_validation.py <file1_path> [file2_path]")
        sys.exit(1)

    file_paths = sys.argv[1:]

    result = is_standard_format(file_paths)

    print(f"\n{'='*60}")
    print(f"文件格式验证结果")
    print(f"{'='*60}")
    print(f"是否为标准格式: {'✅ 是' if result.is_standard else '❌ 否'}")
    if not result.is_standard:
        print(f"原因: {result.reason}")
    print(f"文件数量: {result.file_count}")
    print(f"总 Sheet 数量: {result.total_sheet_count}")
    print(f"\n文件详情:")
    for detail in result.details:
        print(f"  - {detail['filename']}: {detail['sheet_count']} 个 sheet")
        for sheet in detail['sheets']:
            print(f"    • {sheet['sheet_name']}: {sheet['data_row_count']} 行数据")
