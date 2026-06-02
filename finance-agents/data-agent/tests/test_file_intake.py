from __future__ import annotations

from contextlib import contextmanager
import sys
from pathlib import Path

DATA_AGENT_ROOT = Path(__file__).resolve().parents[1]
if str(DATA_AGENT_ROOT) not in sys.path:
    sys.path.insert(0, str(DATA_AGENT_ROOT))

from utils import file_intake
from utils.file_intake import build_upload_name_maps, prepare_logical_upload_files


def test_xls_engine_is_declared_for_runtime_installs() -> None:
    repo_root = Path(__file__).resolve().parents[3]
    data_agent_requirements = (
        repo_root / "finance-agents" / "data-agent" / "requirements.txt"
    ).read_text(encoding="utf-8")
    data_agent_pyproject = (
        repo_root / "finance-agents" / "data-agent" / "pyproject.toml"
    ).read_text(encoding="utf-8")
    finance_mcp_requirements = (
        repo_root / "finance-mcp" / "requirements.txt"
    ).read_text(encoding="utf-8")

    assert "xlrd>=2.0.1" in data_agent_requirements
    assert '"xlrd>=2.0.1"' in data_agent_pyproject
    assert "xlrd>=2.0.1" in finance_mcp_requirements


def test_oss_logical_upload_ref_is_mapped_without_local_file() -> None:
    logical_file = {
        "file_path": "/uploads/oss/company-1/a.csv",
        "display_name": "orders.csv",
        "original_filename": "orders.csv",
        "is_logical_split": False,
    }

    display_name_to_ref, ref_to_display_name = build_upload_name_maps([logical_file])

    assert display_name_to_ref["orders.csv"] == "/uploads/oss/company-1/a.csv"
    assert ref_to_display_name["/uploads/oss/company-1/a.csv"] == "orders.csv"


def test_oss_logical_upload_ref_with_columns_is_accepted_without_local_file() -> None:
    logical_file = {
        "file_path": "/uploads/oss/company-1/a.csv",
        "display_name": "orders.csv",
        "original_filename": "orders.csv",
        "columns": ["订单号", "金额"],
        "has_data_rows": True,
    }
    file_rule = {
        "file_validation_rules": {
            "validation_config": {},
            "table_schemas": [
                {
                    "table_name": "订单表",
                    "file_type": ["csv"],
                    "required_columns": ["订单号", "金额"],
                }
            ],
        }
    }

    result = prepare_logical_upload_files([logical_file], file_rule=file_rule)

    assert result["kept_count"] == 1
    assert result["logical_uploaded_files"][0]["file_path"] == "/uploads/oss/company-1/a.csv"
    assert result["files_with_columns"] == [
        {"file_name": "orders.csv", "columns": ["订单号", "金额"]}
    ]


def test_oss_logical_upload_ref_without_columns_reads_headers_from_storage(
    monkeypatch,
    tmp_path: Path,
) -> None:
    materialized = tmp_path / "orders.csv"
    materialized.write_text("订单号,金额\nA001,12.3\n", encoding="utf-8-sig")
    calls: list[str] = []

    def fake_materialize_oss_logical_file(file_ref: str):
        calls.append(file_ref)

        @contextmanager
        def _materialized():
            yield materialized

        return _materialized()

    monkeypatch.setattr(
        file_intake,
        "_materialize_oss_logical_file",
        fake_materialize_oss_logical_file,
    )

    result = prepare_logical_upload_files(
        [
            {
                "file_path": "/uploads/oss/company-1/a.csv",
                "original_filename": "orders.csv",
                "size_bytes": 123,
            }
        ],
        file_rule={
            "file_validation_rules": {
                "validation_config": {},
                "table_schemas": [
                    {
                        "table_name": "订单表",
                        "file_type": ["csv"],
                        "required_columns": ["订单号", "金额"],
                    }
                ],
            }
        },
    )

    assert calls == ["/uploads/oss/company-1/a.csv"]
    assert result["kept_count"] == 1
    assert result["logical_uploaded_files"][0]["file_path"] == "/uploads/oss/company-1/a.csv"
    assert result["files_with_columns"] == [
        {"file_name": "orders.csv", "columns": ["订单号", "金额"]}
    ]


def test_oss_logical_upload_multi_sheet_workbook_preserves_logical_refs(
    monkeypatch,
    tmp_path: Path,
) -> None:
    import openpyxl

    materialized = tmp_path / "workbook.xlsx"
    workbook = openpyxl.Workbook()
    first_sheet = workbook.active
    first_sheet.title = "订单"
    first_sheet.append(["订单号", "金额"])
    first_sheet.append(["A001", 12.3])
    second_sheet = workbook.create_sheet("退款")
    second_sheet.append(["退款单号", "退款金额"])
    second_sheet.append(["R001", 4.5])
    workbook.save(materialized)

    def fake_materialize_oss_logical_file(file_ref: str):
        assert file_ref == "/uploads/oss/company-1/workbook.xlsx"

        @contextmanager
        def _materialized():
            yield materialized

        return _materialized()

    monkeypatch.setattr(
        file_intake,
        "_materialize_oss_logical_file",
        fake_materialize_oss_logical_file,
    )

    result = prepare_logical_upload_files(
        [
            {
                "file_path": "/uploads/oss/company-1/workbook.xlsx",
                "original_filename": "workbook.xlsx",
            }
        ],
        file_rule={
            "file_validation_rules": {
                "validation_config": {},
                "table_schemas": [
                    {
                        "table_name": "订单表",
                        "file_type": ["xlsx"],
                        "required_columns": ["订单号", "金额"],
                    },
                    {
                        "table_name": "退款表",
                        "file_type": ["xlsx"],
                        "required_columns": ["退款单号", "退款金额"],
                    },
                ],
            }
        },
    )

    logical_files = result["logical_uploaded_files"]
    summaries = result["prefilter_summary"]

    assert result["kept_count"] == 2
    assert {item["sheet_name"] for item in logical_files} == {"订单", "退款"}
    assert {item["sheet_index"] for item in logical_files} == {1, 2}
    assert all(item["is_logical_split"] is True for item in logical_files)
    assert all(item["file_path"].startswith("/uploads/oss/company-1/workbook.xlsx#sheet=") for item in logical_files)
    assert len({item["file_path"] for item in logical_files}) == 2
    assert all(item["workbook_file_path"] == "/uploads/oss/company-1/workbook.xlsx" for item in logical_files)
    assert all(summary["file_path"].startswith("/uploads/oss/company-1/workbook.xlsx#sheet=") for summary in summaries)
    assert all(str(tmp_path) not in summary["file_path"] for summary in summaries)

    display_name_to_ref, _ = build_upload_name_maps(logical_files)
    for item in logical_files:
        assert display_name_to_ref[item["display_name"]] == item["file_path"]


def test_oss_logical_upload_xls_without_columns_reads_headers_with_pandas(
    monkeypatch,
    tmp_path: Path,
) -> None:
    import openpyxl
    import pandas as pd

    materialized = tmp_path / "legacy.xls"
    materialized.write_bytes(b"xls")
    header_frame = pd.DataFrame([["订单号", "金额"], ["A001", 12.3]])
    materialize_calls: list[str] = []
    read_excel_calls: list[dict[str, object]] = []

    def fake_materialize_oss_logical_file(file_ref: str):
        materialize_calls.append(file_ref)

        @contextmanager
        def _materialized():
            yield materialized

        return _materialized()

    def fake_read_excel(path: Path, **kwargs):
        read_excel_calls.append({"path": path, **kwargs})
        return {"Sheet1": header_frame}

    monkeypatch.setattr(
        file_intake,
        "_materialize_oss_logical_file",
        fake_materialize_oss_logical_file,
    )
    monkeypatch.setattr(pd, "read_excel", fake_read_excel)
    monkeypatch.setattr(
        openpyxl,
        "load_workbook",
        lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("openpyxl called")),
    )

    result = prepare_logical_upload_files(
        [
            {
                "file_path": "/uploads/oss/company-1/legacy.xls",
                "original_filename": "legacy.xls",
            }
        ],
        file_rule={
            "file_validation_rules": {
                "validation_config": {},
                "table_schemas": [
                    {
                        "table_name": "订单表",
                        "file_type": ["xls"],
                        "required_columns": ["订单号", "金额"],
                    }
                ],
            }
        },
    )

    assert materialize_calls == ["/uploads/oss/company-1/legacy.xls"]
    assert read_excel_calls == [
        {
            "path": materialized,
            "sheet_name": None,
            "header": None,
            "dtype": object,
        }
    ]
    assert result["kept_count"] == 1
    logical_file = result["logical_uploaded_files"][0]
    assert logical_file["file_path"] == "/uploads/oss/company-1/legacy.xls"
    assert logical_file["workbook_file_path"] == "/uploads/oss/company-1/legacy.xls"
    assert logical_file["sheet_name"] == "Sheet1"
    assert logical_file["is_logical_split"] is False
    assert result["files_with_columns"] == [
        {"file_name": "legacy.xls", "columns": ["订单号", "金额"]}
    ]


def test_oss_logical_upload_xls_multi_sheet_preserves_logical_refs(
    monkeypatch,
    tmp_path: Path,
) -> None:
    import openpyxl
    import pandas as pd

    materialized = tmp_path / "legacy.xls"
    materialized.write_bytes(b"xls")

    def fake_materialize_oss_logical_file(file_ref: str):
        assert file_ref == "/uploads/oss/company-1/legacy.xls"

        @contextmanager
        def _materialized():
            yield materialized

        return _materialized()

    def fake_read_excel(path: Path, **kwargs):
        assert path == materialized
        return {
            "订单": pd.DataFrame([["订单号", "金额"], ["A001", 12.3]]),
            "退款": pd.DataFrame([["退款单号", "退款金额"], ["R001", 4.5]]),
        }

    monkeypatch.setattr(
        file_intake,
        "_materialize_oss_logical_file",
        fake_materialize_oss_logical_file,
    )
    monkeypatch.setattr(pd, "read_excel", fake_read_excel)
    monkeypatch.setattr(
        openpyxl,
        "load_workbook",
        lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("openpyxl called")),
    )

    result = prepare_logical_upload_files(
        [
            {
                "file_path": "/uploads/oss/company-1/legacy.xls",
                "original_filename": "legacy.xls",
            }
        ],
        file_rule={
            "file_validation_rules": {
                "validation_config": {},
                "table_schemas": [
                    {
                        "table_name": "订单表",
                        "file_type": ["xls"],
                        "required_columns": ["订单号", "金额"],
                    },
                    {
                        "table_name": "退款表",
                        "file_type": ["xls"],
                        "required_columns": ["退款单号", "退款金额"],
                    },
                ],
            }
        },
    )

    logical_files = result["logical_uploaded_files"]
    summaries = result["prefilter_summary"]

    assert result["kept_count"] == 2
    assert {item["sheet_name"] for item in logical_files} == {"订单", "退款"}
    assert {item["sheet_index"] for item in logical_files} == {1, 2}
    assert all(item["is_logical_split"] is True for item in logical_files)
    assert all(item["file_path"].startswith("/uploads/oss/company-1/legacy.xls#sheet=") for item in logical_files)
    assert len({item["file_path"] for item in logical_files}) == 2
    assert all(item["workbook_file_path"] == "/uploads/oss/company-1/legacy.xls" for item in logical_files)
    assert all(summary["file_path"].startswith("/uploads/oss/company-1/legacy.xls#sheet=") for summary in summaries)
    assert all(str(tmp_path) not in summary["file_path"] for summary in summaries)


def test_oss_logical_upload_ref_with_explicit_empty_columns_prefilters_empty_header(
    monkeypatch,
) -> None:
    calls: list[str] = []
    monkeypatch.setattr(
        file_intake,
        "_materialize_oss_logical_file",
        lambda file_ref: calls.append(file_ref),
    )

    result = prepare_logical_upload_files(
        [
            {
                "file_path": "/uploads/oss/company-1/a.csv",
                "original_filename": "orders.csv",
                "columns": [],
            }
        ],
        file_rule=None,
    )

    assert calls == []
    assert result["kept_count"] == 0
    assert result["prefilter_summary"][0]["reason_code"] == "empty_header"
