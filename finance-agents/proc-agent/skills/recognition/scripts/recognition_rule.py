"""核算报表填充脚本

将手工凭证 Excel 中提取的数据填充到已有的 BI 报表中：
1. BI 费用明细表 (BI费用明细表.xlsx)  — 将手工凭证费用科目数据追加填充到已有表
2. BI 损益毛利明细表 (BI损益毛利明细表.xlsx) — 将手工凭证收入成本数据追加填充到已有表

业务逻辑依据: AI自动化逻辑20260103.xlsx（章节一、四、五）

输入文件（均由用户通过前端上传）：
- 手工凭证 Excel（必填）：提取数据源
- BI费用明细表 Excel（必填）：待填充的目标表
- BI损益毛利明细表 Excel（必填）：待填充的目标表
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import sys
import traceback
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd

logger = logging.getLogger(__name__)

# ── 路径配置 ──────────────────────────────────────────────────────────────────

SKILL_DIR = Path(__file__).parent.parent          # skills/recognition/
DATA_DIR = SKILL_DIR / "data"                     # 参考模板目录
REFERENCES_DIR = SKILL_DIR / "references"         # references 目录
PROC_AGENT_DIR = SKILL_DIR.parent.parent          # proc-agent 根目录
DEFAULT_RESULT_DIR = PROC_AGENT_DIR / "result"    # 默认输出目录

# ── 收入类型映射（从 references/revenue_type_reference.md 加载）──────────────────


def load_revenue_type_mapping() -> Dict[str, Dict[str, str]]:
    """从 references/revenue_type_reference.md 加载收入类型映射。
    
    Returns:
        dict，键为科目名称，值为{"收入类型": ..., "收入明细": ...}
    """
    mapping: Dict[str, Dict[str, str]] = {}
    ref_file = REFERENCES_DIR / "revenue_type_reference.md"
    
    if not ref_file.exists():
        logger.warning(f"收入类型参考文件不存在: {ref_file}")
        return mapping
    
    try:
        with open(ref_file, "r", encoding="utf-8") as f:
            content = f.read()
        
        # 解析markdown表格
        lines = content.split("\n")
        for line in lines:
            line = line.strip()
            # 过滤表头行和分隔符行（分隔符行通常是 |---|---| 格式）
            if line.startswith("|") and "科目名称" not in line and not line.startswith("|-"):
                parts = line.split("|")
                if len(parts) >= 4:
                    subject = parts[1].strip()
                    revenue_type = parts[2].strip()
                    revenue_detail = parts[3].strip()
                    if subject and subject != "科目名称":
                        mapping[subject] = {
                            "收入类型": revenue_type,
                            "收入明细": revenue_detail,
                        }
        logger.info(f"收入类型映射加载完成，共 {len(mapping)} 条")
    except Exception as e:
        logger.error(f"加载收入类型映射失败: {e}")
    
    return mapping


# 全局收入类型映射缓存
REVENUE_TYPE_MAPPING: Dict[str, Dict[str, str]] = {}


def get_revenue_type_mapping() -> Dict[str, Dict[str, str]]:
    """获取收入类型映射（带缓存）。"""
    global REVENUE_TYPE_MAPPING
    if not REVENUE_TYPE_MAPPING:
        REVENUE_TYPE_MAPPING = load_revenue_type_mapping()
    return REVENUE_TYPE_MAPPING


def get_revenue_info(subject_name: str) -> Dict[str, str]:
    """根据科目名称获取收入类型和收入明细。
    
    Args:
        subject_name: 科目名称
        
    Returns:
        {"收入类型": ..., "收入明细": ...}，如果未找到则返回 {"收入类型": "", "收入明细": ""}
    """
    mapping = get_revenue_type_mapping()
    return mapping.get(subject_name, {"收入类型": "", "收入明细": ""})

# 服务地址（可通过环境变量覆盖）
DATA_AGENT_HOST = os.getenv("DATA_AGENT_HOST", "localhost")
DATA_AGENT_PORT = os.getenv("DATA_AGENT_PORT", "8100")
BASE_URL = f"http://{DATA_AGENT_HOST}:{DATA_AGENT_PORT}"

# ── 业务常量：公司主体明细（关联判断依据）────────────────────────────────────────

COMPANY_ENTITIES = {
    "福建福禄网络科技有限公司",
    "福禄（武汉）科技集团有限公司",
    "福禄（香港）有限公司",
    "福禄福擎网络科技有限公司",
    "福禄控股有限公司",
    "海南福禄网络科技有限公司",
    "杭州福之禄至信息技术有限公司",
    "杭州福之禄至信息技术有限公司上海分公司",
    "湖北福之禄至网络科技有限公司",
    "湖北氪金网络科技有限公司",
    "江苏福之禄至信息技术有限公司",
    "江西福禄网络科技有限公司",
    "喀什一起玩网络科技有限公司",
    "南京宏昇网络科技有限公司",
    "上海禄至网络科技有限公司",
    "沭阳数宇科技有限公司",
    "天津如意咨询有限公司",
    "武汉福福文化传媒有限公司",
    "武汉福禄传媒有限公司",
    "武汉福禄传媒有限公司第二分公司",
    "武汉福禄传媒有限公司第一分公司",
    "武汉福禄创新股权投资合伙企业(有限合伙)",
    "武汉福禄供应链管理有限公司",
    "武汉福禄数云网络科技有限公司",
    "武汉福禄私募基金管理有限公司",
    "武汉福禄网络科技有限公司",
    "武汉福禄网络科技有限公司上海分公司",
    "武汉福禄相随网络科技有限公司",
    "武汉福禄信息科技有限公司",
    "武汉福穗网络科技有限公司",
    "武汉福游网络科技有限公司",
    "武汉福悦网络科技有限公司",
    "武汉福韵网络科技有限公司",
    "武汉立硕科技有限公司",
    "武汉市福禄共想网络科技有限公司",
    "武汉搜卡科技有限公司",
    "武汉天识科技有限公司",
    "武汉一起游网络科技有限公司",
    "武汉亿禄网络科技有限公司",
    "西藏福禄网络科技有限公司",
    "西藏葫芦娃网络科技有限公司",
    "咸宁福蕊网络科技有限公司",
    "新疆福禄网络科技有限公司",
    "新疆福佑网络科技有限公司",
    "新疆葫芦娃网络科技有限公司",
    "新沂福云网络科技有限公司",
    "新沂福之禄至信息技术有限公司",
    "周口福蕊网络科技有限公司",
}

# ── 业务常量：科目分类 ──────────────────────────────────────────────────────────

# 章节一：收入类科目关键词
INCOME_SUBJECT_KEYWORDS = ["主营业务收入"]
# 章节一：销售成本科目关键词
SALES_COST_SUBJECT_KEYWORDS = ["主营业务成本_销售成本", "主营业务成本-销售成本"]
# 章节一：技术服务成本科目关键词
TECH_COST_SUBJECT_KEYWORDS = ["主营业务成本_技术服务成本", "主营业务成本-技术服务成本"]

# 章节一：寄售收入科目
CONSIGNMENT_SUBJECT = "主营业务收入_寄售收入"
# 章节一：技术服务收入科目关键词（代运营维度）
TECH_SERVICE_INCOME_KEYWORD = "主营业务收入_技术服务收入"

# 章节四：费用类科目关键词
EXPENSE_SUBJECT_KEYWORDS = [
    "研发支出_费用化支出",
    "主营业务成本_直接服务费",
    "主营业务成本_工资性支出",
    "其他业务收入",
    "其他业务成本",
    "管理费用",
    "销售费用",
    "财务费用",
    "营业外收入",
    "营业外支出",
    "投资收益",
    "资产减值损失",
    "所得税",
]
# 章节四：管理费用中需剔除的子科目
MGMT_EXPENSE_EXCLUDE_KEYWORDS = ["管理费用_研发费", "管理费用-研发费"]

# ── 工具函数 ───────────────────────────────────────────────────────────────────


def clean_amount(value: Any) -> float:
    """清洗金额字段，返回浮点数（空/NaN 返回 0.0）。"""
    if value is None:
        return 0.0
    s = str(value).strip()
    if s in ("", "-", "nan", "NaN", "None"):
        return 0.0
    s = s.replace(",", "").replace("，", "")
    match = re.search(r"-?[\d.]+", s)
    return float(match.group()) if match else 0.0


def parse_tax_rate_str(rate_str: str) -> float:
    """将税率字符串（如 '6%'、'13%'、'0.06'）转换为小数。"""
    s = str(rate_str).strip()
    if not s:
        return 0.0
    match = re.search(r"([\d.]+)\s*%", s)
    if match:
        return float(match.group(1)) / 100.0
    match = re.search(r"[\d.]+", s)
    if match:
        v = float(match.group())
        return v / 100.0 if v > 1 else v
    return 0.0


def get_tax_rate_from_subject(subject_name: str) -> float:
    """从科目名称后缀推断税率（章节一规则6）。"""
    s = str(subject_name)
    for suffix, rate in [
        ("-13%", 0.13), ("_13%", 0.13),
        ("-9%", 0.09), ("_9%", 0.09),
        ("-3%", 0.03), ("_3%", 0.03),
        ("-1%", 0.01), ("_1%", 0.01),
    ]:
        if s.endswith(suffix):
            return rate
    return 0.06  # 默认 6%


def _is_company_like(name: str) -> bool:
    """判断字符串是否看起来像公司/客户/供应商名称。

    公司名称具备以下特征之一：
    - 含有限公司/集团/公司/工厂/商行等公司尾缀
    - 其不是税率字符串、商品大类名、摘要描述词
    """
    if not name:
        return False
    s = name.strip()
    if not s:
        return False
    # 过滤税率字符串
    if re.match(r'^\d+(\.\d+)?%?$', s):
        return False
    # 过滤过短的描述性词（小于5个字符且不含公司关键词）
    company_keywords = ["公司", "集团", "企业", "工厂", "商行", "巴士物幼", "体育", "店", "机构", "学校",
                        "科技", "网络", "信息", "数字", "资讯", "证券", "银行",
                        "Limited", "Ltd", "Inc", "Corp", "LLC"]
    has_company_kw = any(kw in s for kw in company_keywords)
    # 长度小于5且无公司关键词，认为是描述性词而非公司名
    if len(s) < 5 and not has_company_kw:
        return False
    return True


def parse_summary(summary: str) -> Dict[str, str]:
    """解析摘要字段，按 & 和 + 分列结构化字段。

    摘要格式（手工凭证原表）：
        <描述文字>&<调整类型细分>+<客户>+<商户号(客户)>+<税率>
        <描述文字>&<调整类型细分>+<客户>+<供应商>+<商户号(客户)>+<商户号(供应商)>+<税率>

    说明：
    - & 前：整段描述，作为 adj_type_raw（原始调整类型描述）
    - & 后 plus_parts[0]：调整类型细分，固定为"调整收入"/"调整成本"/"调整费用"等
    - & 后 plus_parts[1]：客户（只接受公司名）
    - & 后 plus_parts[2]：供应商（只接受公司名），或商户号（纯数字），或税率
    - & 后 plus_parts[3]：商户号(客户) 或税率
    - & 后 plus_parts[4]：商户号(供应商) 或税率
    - & 后 plus_parts[5]：税率

    注意：位置3/4/5 的内容需按类型判断（公司名/纯数字/税率%）
    """
    s = str(summary).strip()
    adj_type_raw = ""   # & 前的完整描述
    adj_type_sub = ""   # & 后 plus_parts[0]，调整类型细分
    rest = s

    if "&" in s:
        idx = s.index("&")
        adj_type_raw = s[:idx].strip()
        rest = s[idx + 1:]

    plus_parts = [p.strip() for p in rest.split("+")]

    # plus_parts[0] 是调整类型细分（调整收入/调整成本/调整费用 等），不是客户
    adj_type_sub = plus_parts[0] if len(plus_parts) > 0 else ""

    # plus_parts[1] 起才是客户/供应商/商户号/税率
    remaining = plus_parts[1:] if len(plus_parts) > 1 else []

    customer = ""
    vendor = ""
    merchant_cust = ""
    merchant_vend = ""
    tax_rate_str = ""

    # 按类型逐一识别 remaining 中各段
    company_slots = []   # 收集公司名（前两个公司名依次为客户、供应商）
    for part in remaining:
        p = part.strip()
        if not p:
            continue
        if re.match(r'^\d+(\.\d+)?%$', p):
            # 税率（如 6%、13%）
            if not tax_rate_str:
                tax_rate_str = p
        elif re.match(r'^\d+$', p):
            # 纯数字：商户号
            if not merchant_cust:
                merchant_cust = p
            elif not merchant_vend:
                merchant_vend = p
        elif _is_company_like(p):
            # 公司名
            company_slots.append(p)
        # 其余（如描述性文字）忽略

    customer = company_slots[0] if len(company_slots) > 0 else ""
    vendor = company_slots[1] if len(company_slots) > 1 else ""

    # adj_type 对外统一使用 & 后的调整类型细分（如"调整收入"/"调整成本"），
    # 便于后续 if adj_type == "调整收入" 的判断；同时保留 adj_type_raw 供调试
    adj_type = adj_type_sub if adj_type_sub else adj_type_raw

    return {
        "调整类型": adj_type,
        "调整类型原始": adj_type_raw,
        "客户": customer,
        "供应商": vendor,
        "商户号_客户": merchant_cust,
        "商户号_供应商": merchant_vend,
        "税率字符串": tax_rate_str,
    }


def is_related_company(name: str) -> str:
    """判断是否关联公司（章节一规则3/4）。"""
    return "关联" if str(name).strip() in COMPANY_ENTITIES else "非关联"


def get_subject_parts(subject_name: str) -> Tuple[str, str, str]:
    """按 _ 分列科目名称为三级（章节四规则3），每级去除首尾空格。"""
    parts = str(subject_name).split("_")
    level1 = parts[0].strip() if len(parts) > 0 else ""
    level2 = parts[1].strip() if len(parts) > 1 else ""
    level3 = parts[2].strip() if len(parts) > 2 else ""
    return level1, level2, level3


def is_income_subject(subject_name: str) -> bool:
    """判断是否为章节一收入类科目。"""
    return any(kw in subject_name for kw in INCOME_SUBJECT_KEYWORDS)


def is_sales_cost_subject(subject_name: str) -> bool:
    """判断是否为章节一销售成本科目。"""
    return any(kw in subject_name for kw in SALES_COST_SUBJECT_KEYWORDS)


def is_tech_cost_subject(subject_name: str) -> bool:
    """判断是否为章节一技术服务成本科目。"""
    return any(kw in subject_name for kw in TECH_COST_SUBJECT_KEYWORDS)


def is_expense_subject(subject_name: str) -> bool:
    """判断是否为章节四费用类科目（管理费用剔除研发费子科目）。"""
    s = str(subject_name)
    for exclude in MGMT_EXPENSE_EXCLUDE_KEYWORDS:
        if exclude in s:
            return False
    return any(kw in s for kw in EXPENSE_SUBJECT_KEYWORDS)


# ── 读取手工凭证 Excel ─────────────────────────────────────────────────────────


def _read_voucher_sheet(xl: pd.ExcelFile, sheet_name: str) -> pd.DataFrame:
    """读取手工凭证 Excel 中的单个 Sheet，自动识别表头行并统一列名。

    Returns:
        DataFrame，含标准列名；若无法识别表头则返回空 DataFrame
    """
    df_raw = xl.parse(sheet_name, header=None, dtype=str)

    # 查找表头行（包含"科目名称"或"凭证号"的行）
    header_row = None
    for i in range(min(10, len(df_raw))):
        row_vals = [str(v).strip() for v in df_raw.iloc[i].values]
        if "科目名称" in row_vals or "凭证号" in row_vals:
            header_row = i
            break
    if header_row is None:
        logger.warning(f"Sheet '{sheet_name}' 未找到表头行，跳过")
        return pd.DataFrame()

    df = xl.parse(sheet_name, header=header_row, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(how="all").reset_index(drop=True)

    # 列名别名统一
    col_aliases = {
        "借方发生额": ["借方发生额", "借方金额", "借方"],
        "贷方发生额": ["贷方发生额", "贷方金额", "贷方"],
        "科目编码": ["科目编码", "科目编号", "科目代码"],
        "业务时间": ["业务时间", "业务日期", "记账日期", "日期"],
    }
    rename_map: Dict[str, str] = {}
    for std_name, aliases in col_aliases.items():
        if std_name not in df.columns:
            for alias in aliases:
                if alias in df.columns:
                    rename_map[alias] = std_name
                    break
    if rename_map:
        df = df.rename(columns=rename_map)

    # 整理表（含"收入成本毛利整理"关键词）列名映射到标准原表列名
    # 整理表已有：客户、供应商、是否关联（客户）、是否关联（供应商）、业务模式 等字段
    # 统一到原表规范列名，方便后续统一处理
    summary_sheet_rename: Dict[str, str] = {
        "是否关联（客户）": "_预解析_是否关联客户",
        "是否关联（供应商）": "_预解析_是否关联供应商",
        "业务模式": "_预解析_核算类型",
    }
    if "客户" in df.columns:
        # 整理表有客户列，标记为预解析来源（不需要从摘要二次解析）
        df["_来源_客户已解析"] = True
        for src, dst in summary_sheet_rename.items():
            if src in df.columns:
                df = df.rename(columns={src: dst})
    else:
        df["_来源_客户已解析"] = False

    logger.info(f"读取 Sheet '{sheet_name}' 完成，共 {len(df)} 行")
    return df


def read_voucher_file(file_path: str) -> pd.DataFrame:
    """读取手工凭证 Excel，自动识别所有相关 Sheet 并合并。

    同时读取：
    - 手工凭证原表（来源类型包含手工凭证和其它系统引入）
    - 手工凭证-收入成本毛利整理表（已整理好客户/供应商字段）

    Returns:
        合并后的 DataFrame，含标准列名
    """
    xl = pd.ExcelFile(file_path)
    logger.info(f"手工凭证文件 Sheets: {xl.sheet_names}")

    # 分类识别 Sheet
    original_sheet = None    # 手工凭证原表
    summary_sheet = None     # 收入成本毛利整理表
    for sheet in xl.sheet_names:
        if "收入成本毛利整理" in sheet or "收入成本整理" in sheet:
            summary_sheet = sheet
        elif "手工凭证原表" in sheet:
            original_sheet = sheet
        elif "手工凭证" in sheet and original_sheet is None:
            original_sheet = sheet

    if original_sheet is None and summary_sheet is None:
        # 降级：读取第一个 Sheet
        original_sheet = xl.sheet_names[0]

    dfs: List[pd.DataFrame] = []

    if original_sheet:
        df_orig = _read_voucher_sheet(xl, original_sheet)
        if not df_orig.empty:
            df_orig["_sheet_来源"] = "原表"
            # 如果整理表存在，则原表只保留"手工凭证"来源类型的行，
            # 避免与整理表（已包含"其它系统引入"数据）重复处理
            if summary_sheet:
                src_col = "来源类型"
                if src_col in df_orig.columns:
                    orig_count_before = len(df_orig)
                    df_orig = df_orig[df_orig[src_col].str.contains("手工凭证", na=False)].copy()
                    logger.info(f"原表过滤来源类型=手工凭证: {orig_count_before} → {len(df_orig)} 行（已排除其它系统引入行以避免与整理表重复）")
            dfs.append(df_orig)

    if summary_sheet:
        df_summ = _read_voucher_sheet(xl, summary_sheet)
        if not df_summ.empty:
            df_summ["_sheet_来源"] = "整理表"
            dfs.append(df_summ)

    if not dfs:
        logger.error("未能读取任何有效 Sheet")
        return pd.DataFrame()

    # 合并所有 Sheet，缺失列填充空字符串
    combined = pd.concat(dfs, ignore_index=True, sort=False)
    combined = combined.fillna("")
    logger.info(f"手工凭证合并后共 {len(combined)} 行，来自 Sheets: orig={original_sheet}, summary={summary_sheet}")
    return combined


def identify_file_type(file_path: str) -> str:
    """识别文件类型：
    - 'voucher'      手工凭证文件
    - 'bi_expense'   BI费用明细表文件
    - 'bi_profit'    BI损益毛利明细表文件
    - 'unknown'      未知文件
    """
    name = Path(file_path).name

    # 优先通过文件名识别
    if any(kw in name for kw in ["手工凭证", "凭证原表", "凭证"]):
        return "voucher"
    if "BI费用明细" in name or "费用明细表" in name:
        return "bi_expense"
    if any(kw in name for kw in ["BI损益", "损益毛利", "供应商", "代运营毛利"]):
        return "bi_profit"

    # 文件名无法判断时，通过内容特征识别
    try:
        df_raw = pd.read_excel(file_path, header=None, dtype=str, nrows=5)
        content = df_raw.to_string()
        # 手工凭证特征：借方发生额/贷方发生额/科目名称 同时存在
        voucher_score = sum(
            1 for kw in ["凭证号", "借方发生额", "贷方发生额", "科目名称", "摘要"] if kw in content
        )
        if voucher_score >= 3:
            return "voucher"
        # BI费用明细表特征
        expense_score = sum(
            1 for kw in ["eas一级科目", "eas含税金额", "eas不含税金额", "费用二级项目", "数据来源"] if kw in content
        )
        if expense_score >= 3:
            return "bi_expense"
        # BI损益毛利明细表特征
        profit_score = sum(
            1 for kw in ["供应商名称", "调整后gmv", "调整后成本", "代运营收入", "合并毛利"] if kw in content
        )
        if profit_score >= 2:
            return "bi_profit"
    except Exception:
        pass
    return "unknown"


def read_existing_bi_file(file_path: str) -> Tuple[pd.DataFrame, List[str]]:
    """读取用户上传的已有 BI 报表文件，返回 (DataFrame, 原始列名列表)。

    自动识别表头行（前10行内），保留原始列名以便追加后保持格式一致。
    """
    xl = pd.ExcelFile(file_path)
    # 读取第一个 Sheet
    target_sheet = xl.sheet_names[0]

    df_raw = xl.parse(target_sheet, header=None, dtype=str)

    # 查找表头行（包含关键列名的行）
    expense_keywords = ["eas一级科目", "eas含税金额", "数据来源", "rp单据", "费用二级项目"]
    profit_keywords = ["供应商名称", "调整后gmv", "合并毛利", "代运营收入"]

    header_row = 0
    for i in range(min(10, len(df_raw))):
        row_vals = [str(v).strip() for v in df_raw.iloc[i].values]
        row_str = " ".join(row_vals)
        if any(kw in row_str for kw in expense_keywords + profit_keywords):
            header_row = i
            break

    df = xl.parse(target_sheet, header=header_row, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    # 过滤全空行
    df = df.dropna(how="all").reset_index(drop=True)

    original_columns = list(df.columns)
    logger.info(f"读取已有 BI 文件: {Path(file_path).name}，共 {len(df)} 行，列: {original_columns}")
    return df, original_columns


# ── 章节四+五：从手工凭证提取费用数据 → 追加到已有 BI 费用明细表 ──────────────────


def _get_tax_rate_from_subject_str(subject_name: str, summary_tax_rate: float) -> Tuple[float, str]:
    """按同步规则返回税率小数和税率显示字符串。
    
    费用表同步规则：优先摘要中的税率，其次科目名称后缀，无法判断则默认0%。
    损益毛利表同步规则：按科目名称后缀确定税率（13%/9%/3%/1%/其他6%）。
    
    Returns:
        (税率小数, 税率显示字符串如 '6%')
    """
    if summary_tax_rate > 0:
        rate = summary_tax_rate
    else:
        rate = get_tax_rate_from_subject(subject_name)
    pct = int(round(rate * 100))
    rate_str = f"{pct}%" if pct > 0 else "0%"
    return rate, rate_str


def extract_expense_from_voucher(df: pd.DataFrame) -> pd.DataFrame:
    """从手工凭证中提取费用明细数据（章节四 + 章节五字段映射）。

    章节四：过滤费用科目 → 解析摘要 → 计算含税金额
    章节五：字段映射，字段名与 BI费用明细表 列名对齐

    来源优先级（税率）：
    1. 整理表已解析的税率列（_来源_客户已解析 == True 且税率列有值）
    2. 原表摘要中 & 后的税率字符串
    3. 科目名称后缀推断税率

    Returns:
        DataFrame，列名与 BI费用明细表 完全一致，可直接追加
    """
    rows = []

    for _, row in df.iterrows():
        subject_name = str(row.get("科目名称", "")).strip()
        if not is_expense_subject(subject_name):
            continue

        debit = clean_amount(row.get("借方发生额", ""))
        credit = clean_amount(row.get("贷方发生额", ""))

        # 章节四：摘要解析
        summary = str(row.get("摘要", "")).strip()

        # ── 税率取值逻辑（三层优先级）──────────────────────────────────────────
        already_parsed = str(row.get("_来源_客户已解析", "")).strip().lower() in ("true", "1")
        summary_tax_rate = 0.0

        if already_parsed:
            # 整理表：优先使用已解析的税率列
            tax_rate_from_row = clean_amount(row.get("税率", ""))
            if tax_rate_from_row > 0:
                summary_tax_rate = tax_rate_from_row if tax_rate_from_row <= 1 else tax_rate_from_row / 100.0
        
        if summary_tax_rate == 0.0 and "&" in summary:
            # 原表（或整理表无税率）：从摘要 & 后解析
            # 费用摘要格式：调整费用&[客户+]税率
            after_amp = summary[summary.index("&") + 1:].strip()
            if "+" in after_amp:
                rate_str = after_amp.split("+")[-1].strip()
                summary_tax_rate = parse_tax_rate_str(rate_str)
            else:
                summary_tax_rate = parse_tax_rate_str(after_amp)

        # 税率：摘要/整理表优先，其次科目后缀
        tax_rate, tax_rate_str = _get_tax_rate_from_subject_str(subject_name, summary_tax_rate)

        # 章节四：科目三级分列
        level1, level2, level3 = get_subject_parts(subject_name)

        # 章节四：金额计算
        eas_no_tax = round(debit + credit, 2)
        eas_with_tax = round(eas_no_tax * (1 + tax_rate), 2)
        eas_tax = round(eas_with_tax - eas_no_tax, 2)

        # 调整类型
        adj_type = ""
        if already_parsed:
            adj_type = str(row.get("调整类型", "")).strip()
        elif "&" in summary:
            adj_type = summary[:summary.index("&")].strip()

        # 月分区：直接取手工凭证月份列
        period = str(row.get("月份", "")).strip()

        # 行政组织：整理表可能有"行政组织"或"所属组织"列，兜底取公司名称
        admin_org = str(row.get("行政组织", "")).strip()
        if not admin_org:
            admin_org = str(row.get("所属组织", "")).strip()

        # 章节五：字段映射 → BI费用明细表
        rows.append({
            "日期": str(row.get("业务时间", "")).strip(),
            "所属公司": str(row.get("公司名称", "")).strip(),
            "客户": "",
            "店铺_平台": str(row.get("平台", "")).strip(),
            "数据来源": "手工凭证",
            "rp单据": str(row.get("凭证号", "")).strip(),
            "行政中心": admin_org,
            "利润中心": admin_org,
            "eas一级科目": level1,
            "费用二级项目": level2,
            "费用三级项目": level3,
            "账单上提取规则": "",
            "是否归属店铺费用": "",
            "是否参与综合店铺分摊": "",
            "eas含税金额": eas_with_tax,
            "eas不含税金额": eas_no_tax,
            "eas税额": eas_tax,
            "费用项目": adj_type,
            "报表取数": "",
            "报表利润中心id": "",
            "报表利润中心": "",
            "报表费用类型": "",
            "是否公摊": "",
            "中心群": "",
            "摘要备注": summary,
            "月分区": period,
            "备注": "",
            "税率": tax_rate_str,
        })

    return pd.DataFrame(rows)


# 保留旧名兼容调用
generate_expense_detail = extract_expense_from_voucher


# ── 章节一：从手工凭证提取收入成本明细 ──────────────────────────────────────────


def generate_income_cost_detail(df: pd.DataFrame) -> pd.DataFrame:
    """从手工凭证生成手工收入成本明细表（章节一）。

    与 extract_profit_detail_from_voucher 使用相同的客户/供应商解析逻辑：
    - 整理表数据：直接使用已解析的客户/供应商字段
    - 原表数据：摘要含 & 时从摘要解析，无 & 时留空

    Returns:
        含字段：月份, 公司名称, 商品大类, 供应商, 科目名称,
               含税销售额, 含税采购成本, 含税差额收入, 是否技术服务,
               收入类型, 收入明细
    """
    rows = []

    for _, row in df.iterrows():
        subject_name = str(row.get("科目名称", "")).strip()

        is_income = is_income_subject(subject_name)
        is_sales_cost = is_sales_cost_subject(subject_name)
        is_tech_cost = is_tech_cost_subject(subject_name)

        if not (is_income or is_sales_cost or is_tech_cost):
            continue

        debit = clean_amount(row.get("借方发生额", ""))
        credit = clean_amount(row.get("贷方发生额", ""))
        summary = str(row.get("摘要", "")).strip()

        # ── 客户/供应商取值逻辑（与 extract_profit_detail_from_voucher 保持一致）──
        already_parsed = str(row.get("_来源_客户已解析", "")).strip().lower() in ("true", "1")

        if already_parsed:
            # 整理表数据：直接读取已解析字段
            customer = str(row.get("客户", "")).strip()
            vendor = str(row.get("供应商", "")).strip()
            adj_type = str(row.get("调整类型", "")).strip()
            pre_acct_type = str(row.get("_预解析_核算类型", "")).strip()
            accounting_type = pre_acct_type if pre_acct_type else ("净额" if CONSIGNMENT_SUBJECT in subject_name else "总额")
            # 税率：整理表税率列优先
            tax_rate_from_row = clean_amount(row.get("税率", ""))
            if tax_rate_from_row > 0:
                tax_rate = tax_rate_from_row if tax_rate_from_row <= 1 else tax_rate_from_row / 100.0
            else:
                tax_rate = get_tax_rate_from_subject(subject_name)
        else:
            # 原表数据：仅摘要含 & 时解析客户/供应商
            if "&" in summary:
                parsed = parse_summary(summary)
                adj_type = parsed["调整类型"]
                customer = parsed["客户"]
                vendor = parsed["供应商"]
                tax_rate_str_p = parsed["税率字符串"]
                tax_rate = parse_tax_rate_str(tax_rate_str_p) if tax_rate_str_p else get_tax_rate_from_subject(subject_name)
            else:
                # 无 & 分隔符：客户/供应商留空
                adj_type = ""
                customer = ""
                vendor = ""
                tax_rate = get_tax_rate_from_subject(subject_name)
                logger.debug(f"generate_income_cost_detail 摘要无&，客户/供应商留空: 摘要='{summary}'")
            accounting_type = "净额" if CONSIGNMENT_SUBJECT in subject_name else "总额"

        # eas收入/成本不含税（章节一规则7）
        eas_income_no_tax = 0.0
        eas_cost_no_tax = 0.0

        if is_income:
            if CONSIGNMENT_SUBJECT in subject_name:
                if "调整收入" in adj_type:
                    eas_income_no_tax = credit
                elif "调整成本" in adj_type:
                    eas_cost_no_tax = -credit
                else:
                    eas_income_no_tax = credit
            else:
                eas_income_no_tax = credit

        elif is_sales_cost or is_tech_cost:
            eas_cost_no_tax = debit

        # 含税金额计算（章节一规则8）
        taxed_income = round(eas_income_no_tax * (1 + tax_rate), 2)
        taxed_cost = round(eas_cost_no_tax * (1 + tax_rate), 2)
        taxed_diff = round(taxed_income - taxed_cost, 2)

        is_tech_service = (TECH_SERVICE_INCOME_KEYWORD in subject_name) or is_tech_cost

        # 获取收入类型和收入明细（从映射表）
        revenue_info = get_revenue_info(subject_name)

        rows.append({
            "月份": str(row.get("月份", "")).strip(),
            "公司名称": str(row.get("公司名称", "")).strip(),
            "商品大类": str(row.get("商品大类", "")).strip(),
            "供应商": vendor,
            "科目名称": subject_name,
            "含税销售额": taxed_income,
            "含税采购成本": taxed_cost,
            "含税差额收入": taxed_diff,
            "是否技术服务": is_tech_service,
            "收入类型": revenue_info["收入类型"],
            "收入明细": revenue_info["收入明细"],
        })

    return pd.DataFrame(rows)


# ── 章节三：从手工凭证按行提取 BI 损益毛利明细表数据 ─────────────────────────────


# 损益毛利明细表输出列（按行写入模式，对应 BI损益毛利明细表 同步规则）
BI_PROFIT_DETAIL_COLS = [
    "日期", "数据源", "订单类型",
    "公司", "店铺名称加平台", "客户", "是否关联客户",
    "商品大类", "供应商", "是否关联供应商",
    "核算类型或返点类型", "eas科目", "科目名称",
    "含税销售额", "含税采购成本", "含税差额收入",
    "eas收入不含税", "eas成本不含税", "eas税额", "eas差额收入不含税",
    "收入类型", "收入明细", "税率",
]


def extract_profit_detail_from_voucher(df: pd.DataFrame) -> pd.DataFrame:
    """从手工凭证按行提取收入成本明细，对齐 BI损益毛利明细表列格式（按行写入模式）。

    同步规则：
    - 过滤收入类、销售成本类、技术服务成本类科目
    - 客户/供应商：优先使用整理表中已解析的字段，其次从摘要中解析
      - 摘要格式：调整类型&客户+供应商+商户号（客户）+商户号（供应商）+税率
      - 当摘要无 & 分隔符时，无法解析客户/供应商，保持为空
    - 关联判断：查 COMPANY_ENTITIES
    - eas收入/成本不含税：按科目类型分配
    - 含税金额：不含税 × (1 + 税率)
    - 收入类型/收入明细：查 revenue_type_reference.md
    - 税率：优先摘要，其次科目名称后缀

    Returns:
        DataFrame，列名与 BI损益毛利明细表 同步规则完全一致
    """
    rows = []

    for _, row in df.iterrows():
        subject_name = str(row.get("科目名称", "")).strip()
        is_income = is_income_subject(subject_name)
        is_sales_cost = is_sales_cost_subject(subject_name)
        is_tech_cost = is_tech_cost_subject(subject_name)

        if not (is_income or is_sales_cost or is_tech_cost):
            continue

        debit = clean_amount(row.get("借方发生额", ""))
        credit = clean_amount(row.get("贷方发生额", ""))
        summary = str(row.get("摘要", "")).strip()

        # ── 客户/供应商取值逻辑 ───────────────────────────────────────────────────
        # 优先：整理表已解析字段（_来源_客户已解析 == True）
        # 其次：原表摘要含 & 时从摘要解析
        # 兜底：无 & 时客户/供应商留空（摘要内容不能作为客户名）
        already_parsed = str(row.get("_来源_客户已解析", "")).strip().lower() in ("true", "1")

        if already_parsed:
            # 整理表数据：直接读取已解析字段
            customer = str(row.get("客户", "")).strip()
            vendor = str(row.get("供应商", "")).strip()
            adj_type = str(row.get("调整类型", "")).strip()
            # 从整理表的预解析列获取关联状态（若存在则直接用，否则重新计算）
            pre_related_cust = str(row.get("_预解析_是否关联客户", "")).strip()
            pre_related_vend = str(row.get("_预解析_是否关联供应商", "")).strip()
            is_related_cust = pre_related_cust if pre_related_cust else (is_related_company(customer) if customer else "")
            is_related_vend = pre_related_vend if pre_related_vend else (is_related_company(vendor) if vendor else "")
            # 核算类型：整理表已有预解析列
            pre_acct_type = str(row.get("_预解析_核算类型", "")).strip()
            accounting_type = pre_acct_type if pre_acct_type else ("净额" if CONSIGNMENT_SUBJECT in subject_name else "总额")
            # 税率从整理表的税率列（如有），否则从摘要/科目推断
            tax_rate_from_row = clean_amount(row.get("税率", ""))
            if tax_rate_from_row > 0:
                tax_rate = tax_rate_from_row if tax_rate_from_row <= 1 else tax_rate_from_row / 100.0
            else:
                tax_rate = get_tax_rate_from_subject(subject_name)
        else:
            # 原表数据：仅在摘要含 & 时解析客户/供应商
            if "&" in summary:
                parsed = parse_summary(summary)
                adj_type = parsed["调整类型"]
                customer = parsed["客户"]
                vendor = parsed["供应商"]
                tax_rate_str_parsed = parsed["税率字符串"]
                if tax_rate_str_parsed:
                    tax_rate = parse_tax_rate_str(tax_rate_str_parsed)
                else:
                    tax_rate = get_tax_rate_from_subject(subject_name)
            else:
                # 无 & 分隔符：无法从摘要解析客户/供应商，保持为空
                adj_type = ""
                customer = ""
                vendor = ""
                tax_rate = get_tax_rate_from_subject(subject_name)
                logger.debug(f"摘要无&分隔符，客户/供应商留空: 摘要='{summary}'")

            is_related_cust = is_related_company(customer) if customer else ""
            is_related_vend = is_related_company(vendor) if vendor else ""
            accounting_type = "净额" if CONSIGNMENT_SUBJECT in subject_name else "总额"

        pct = int(round(tax_rate * 100))
        tax_rate_display = f"{pct}%" if pct > 0 else "0%"

        # ── eas收入/成本不含税（章节一规则7）────────────────────────────────────────
        eas_income_no_tax = 0.0
        eas_cost_no_tax = 0.0

        if is_income:
            if CONSIGNMENT_SUBJECT in subject_name:
                # 寄售收入：通过调整类型细分判断（用 in 匹配，兼容"调整收入"前后有描述的情况）
                if "调整收入" in adj_type:
                    eas_income_no_tax = credit
                elif "调整成本" in adj_type:
                    # 调整成本：eas成本不含税 = 负的贷方发生额
                    eas_cost_no_tax = -credit
                else:
                    eas_income_no_tax = credit
            else:
                eas_income_no_tax = credit
        elif is_sales_cost or is_tech_cost:
            eas_cost_no_tax = debit

        # ── 含税金额计算（章节一规则8）──────────────────────────────────────────────
        taxed_income = round(eas_income_no_tax * (1 + tax_rate), 2)
        taxed_cost = round(eas_cost_no_tax * (1 + tax_rate), 2)
        taxed_diff = round(taxed_income - taxed_cost, 2)

        # eas税额 = 含税销售额 - eas收入不含税（仅针对收入方，成本行税额=0）
        eas_tax_amount = round(taxed_income - eas_income_no_tax, 2)
        eas_diff_no_tax = round(eas_income_no_tax - eas_cost_no_tax, 2)

        # ── 收入类型和收入明细 ────────────────────────────────────────────────────
        revenue_info = get_revenue_info(subject_name)

        # ── 日期 ──────────────────────────────────────────────────────────────────
        date_val = str(row.get("业务时间", "")).strip()

        rows.append({
            "日期": date_val,
            "数据源": "手工凭证",
            "订单类型": str(row.get("凭证号", "")).strip(),
            "公司": str(row.get("公司名称", "")).strip(),
            "店铺名称加平台": str(row.get("平台", "")).strip(),
            "客户": customer,
            "是否关联客户": is_related_cust,
            "商品大类": str(row.get("商品大类", "")).strip(),
            "供应商": vendor,
            "是否关联供应商": is_related_vend,
            "核算类型或返点类型": accounting_type,
            "eas科目": str(row.get("科目编码", "")).strip(),
            "科目名称": subject_name,
            "含税销售额": taxed_income if taxed_income != 0 else "",
            "含税采购成本": taxed_cost if taxed_cost != 0 else "",
            "含税差额收入": taxed_diff if taxed_diff != 0 else "",
            "eas收入不含税": eas_income_no_tax if eas_income_no_tax != 0 else "",
            "eas成本不含税": eas_cost_no_tax if eas_cost_no_tax != 0 else "",
            "eas税额": eas_tax_amount if eas_tax_amount != 0 else "",
            "eas差额收入不含税": eas_diff_no_tax if eas_diff_no_tax != 0 else "",
            "收入类型": revenue_info["收入类型"],
            "收入明细": revenue_info["收入明细"],
            "税率": tax_rate_display,
        })

    result_df = pd.DataFrame(rows)
    return result_df


# 保留旧名兼容调用
generate_profit_detail = extract_profit_detail_from_voucher


# ── Excel 格式化写入：将新数据追加到已有 BI 表并保存 ──────────────────────────────


BI_EXPENSE_COLS = [
    "日期", "所属公司", "客户", "店铺_平台", "数据来源", "rp单据",
    "行政中心", "利润中心", "eas一级科目", "费用二级项目", "费用三级项目",
    "账单上提取规则", "是否归属店铺费用", "是否参与综合店铺分摊",
    "eas含税金额", "eas不含税金额", "eas税额",
    "费用项目", "报表利润中心", "报表费用类型", "是否公摊", "摘要备注", "期间",
]

PROFIT_COLS = BI_PROFIT_DETAIL_COLS  # 对齐新的按行写入列名


def append_to_existing_expense_excel(
    existing_file: str,
    new_rows_df: pd.DataFrame,
    output_path: str,
) -> None:
    """将新提取的费用数据追加到已有 BI 费用明细表，并保存到 output_path。

    参数:
        existing_file: 用户上传的已有 BI 费用明细表路径
        new_rows_df:   从手工凭证提取的新费用明细 DataFrame
        output_path:   输出文件路径
    """
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    # 读取已有 Excel，保留原始格式（使用 openpyxl 直接操作）
    wb = openpyxl.load_workbook(existing_file)
    ws = wb.active

    # 找到表头行（数据起始行）
    header_row_idx = 1
    for i in range(1, min(12, ws.max_row + 1)):
        row_vals = [str(ws.cell(row=i, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
        row_str = " ".join(row_vals)
        if any(kw in row_str for kw in ["eas一级科目", "eas含税金额", "数据来源", "rp单据"]):
            header_row_idx = i
            break

    # 获取列名映射：列名 → 列号
    col_name_to_idx: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        val = str(ws.cell(row=header_row_idx, column=c).value or "").strip()
        if val:
            col_name_to_idx[val] = c

    logger.info(f"已有费用明细表列映射: {col_name_to_idx}")

    # 动态新增缺失列时使用全新样式对象（避免 StyleProxy unhashable 错误）
    _hdr_font = Font(name="微软雅黑", bold=True, size=9)
    _hdr_thin = Side(style="thin", color="C0C0C0")
    _hdr_border = Border(left=_hdr_thin, right=_hdr_thin, top=_hdr_thin, bottom=_hdr_thin)
    _hdr_align = Alignment(horizontal="center", vertical="center")

    def _add_expense_col(col_name: str) -> None:
        new_col_idx = ws.max_column + 1
        cell = ws.cell(row=header_row_idx, column=new_col_idx, value=col_name)
        cell.font = _hdr_font
        cell.border = _hdr_border
        cell.alignment = _hdr_align
        col_name_to_idx[col_name] = new_col_idx
        logger.info(f"已动态新增\"{col_name}\"列（第 {new_col_idx} 列）到费用明细表")

    # 如果已有表中缺少"月分区"列，在末尾动态新增
    if "月分区" not in col_name_to_idx:
        _add_expense_col("月分区")

    # 如果已有表中缺少"税率"列，在末尾动态新增
    if "税率" not in col_name_to_idx:
        _add_expense_col("税率")

    # 样式：复制倒数第一个数据行的样式作为新行样式基准
    data_font = Font(name="微软雅黑", size=9)
    thin = Side(style="thin", color="C0C0C0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    amount_cols_set = {"eas含税金额", "eas不含税金额", "eas税额"}
    # 费用明细表计算列：eas税额 = eas含税金额 - eas不含税金额
    calc_cols_expense = {"eas税额"}
    from openpyxl.utils import get_column_letter as _get_expense_col_letter

    # 将 new_rows_df 的列名适配到已有表的列名
    # 支持列名别名兼容
    col_aliases = {
        "rp单据": ["rp单据", "rp单据汇总", "凭证号"],
        "店铺_平台": ["店铺_平台", "店铺名称加平台", "平台"],
    }

    def find_col_idx(col_name: str) -> Optional[int]:
        """在已有表中找到列号，支持别名。"""
        if col_name in col_name_to_idx:
            return col_name_to_idx[col_name]
        for std, aliases in col_aliases.items():
            if col_name in aliases:
                for alias in aliases:
                    if alias in col_name_to_idx:
                        return col_name_to_idx[alias]
        return None

    # 追加新行到已有表末尾
    next_row = ws.max_row + 1
    appended_count = 0
    for _, new_row in new_rows_df.iterrows():
        # 获取计算基础列的列号（用于构建公式）
        with_tax_col_idx = find_col_idx("eas含税金额")
        no_tax_col_idx = find_col_idx("eas不含税金额")
        with_tax_letter = _get_expense_col_letter(with_tax_col_idx) if with_tax_col_idx else ""
        no_tax_letter = _get_expense_col_letter(no_tax_col_idx) if no_tax_col_idx else ""

        for col_name, value in new_row.items():
            col_idx = find_col_idx(str(col_name))
            if col_idx is None:
                continue

            cell = ws.cell(row=next_row, column=col_idx)
            cell.font = data_font
            cell.border = border

            if col_name in calc_cols_expense:
                # 计算列：写入 Excel 公式
                if col_name == "eas税额" and with_tax_letter and no_tax_letter:
                    cell.value = f'=IF(OR({with_tax_letter}{next_row}<>"",{no_tax_letter}{next_row}<>""),IFERROR({with_tax_letter}{next_row}-{no_tax_letter}{next_row},""),"")'
                else:
                    if value != "":
                        value = clean_amount(value)
                    cell.value = value if value != "" else None
                cell.number_format = "#,##0.00"
                cell.alignment = right_align
            else:
                if col_name in amount_cols_set:
                    value = clean_amount(value)
                cell.value = value if value != "" else None
                if col_name in amount_cols_set and value != "":
                    cell.number_format = "#,##0.00"
                    cell.alignment = right_align
                else:
                    cell.alignment = left_align
        next_row += 1
        appended_count += 1

    wb.save(output_path)
    logger.info(f"BI 费用明细表已追加 {appended_count} 行并保存: {output_path}")


def write_expense_excel(df: pd.DataFrame, output_path: str) -> None:
    """将 BI 费用明细表写入格式化 Excel（格式参考 data/BI费用明细表.xlsx）。
    此函数用于兼容旧逻辑（无已有文件时从零创建）。"""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BI费用"

    header_font = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
    data_font = Font(name="微软雅黑", size=9)
    header_fill = PatternFill(fill_type="solid", fgColor="4472C4")
    even_fill = PatternFill(fill_type="solid", fgColor="EBF3FB")
    thin = Side(style="thin", color="C0C0C0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    for col in BI_EXPENSE_COLS:
        if col not in df.columns:
            df[col] = ""
    cols = BI_EXPENSE_COLS
    amount_cols = {"eas含税金额", "eas不含税金额", "eas税额"}
    amount_col_indices = {i + 1 for i, c in enumerate(cols) if c in amount_cols}

    ws.row_dimensions[1].height = 22
    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        fill = even_fill if row_idx % 2 == 0 else None
        for col_idx, col_name in enumerate(cols, start=1):
            value = row.get(col_name, "")
            if col_name in amount_cols:
                value = clean_amount(value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = border
            if fill:
                cell.fill = fill
            if col_idx in amount_col_indices:
                cell.number_format = "#,##0.00"
                cell.alignment = right_align
            else:
                cell.alignment = left_align

    for col_idx, col_name in enumerate(cols, start=1):
        sample_lens = [len(str(col_name))]
        for r in range(min(50, len(df))):
            sample_lens.append(len(str(df.iloc[r].get(col_name, ""))))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(sample_lens) * 1.8 + 2, 40)

    wb.save(output_path)
    logger.info(f"BI 费用明细表已保存: {output_path}")


def append_to_existing_profit_excel(
    existing_file: str,
    new_rows_df: pd.DataFrame,
    output_path: str,
) -> None:
    """将新提取的损益毛利明细数据追加到已有 BI 损益毛利明细表，并保存到 output_path。

    支持新型 22 列和旧型汇总格式，如已有表中缺少“税率”列则在末尾动态新增。

    参数:
        existing_file: 用户上传的已有 BI 损益毛利明细表路径
        new_rows_df:   从手工凭证提取的新毛利明细 DataFrame
        output_path:   输出文件路径
    """
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(existing_file)
    # 优先选取名称含"BI"或"明细"的 Sheet，避免误操作"同步规则"等辅助 Sheet
    ws = wb.active
    for sh in wb.worksheets:
        name = sh.title.upper()
        if "BI" in name or "明细" in sh.title or "毛利" in sh.title:
            ws = sh
            break
    logger.info(f"损益毛利明细表操作 Sheet: {ws.title}")
    header_keywords = [
        # 新格式关键列
        "日期", "数据源", "含税销售额", "eas收入不含税", "客户", "供应商",
        # 旧格式关键列
        "供应商名称", "调整后gmv", "合并毛利", "代运营收入",
    ]
    header_row_idx = 1
    for i in range(1, min(12, ws.max_row + 1)):
        row_vals = [str(ws.cell(row=i, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
        row_str = " ".join(row_vals)
        if any(kw in row_str for kw in header_keywords):
            header_row_idx = i
            break

    # 获取列名映射：列名 → 列号
    col_name_to_idx: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        val = str(ws.cell(row=header_row_idx, column=c).value or "").strip()
        if val:
            col_name_to_idx[val] = c

    logger.info(f"已有损益毛利明细表列映射: {col_name_to_idx}")

    # 列名别名映射（已有表列名可能与新格式列名略有差异）
    col_aliases_map: Dict[str, List[str]] = {
        "供应商": ["供应商", "供应商名称"],
        "公司": ["公司", "公司名称"],
        "店铺名称加平台": ["店铺名称加平台", "店铺_平台", "店铺平台"],
        "核算类型或返点类型": ["核算类型或返点类型", "核算类型", "返点类型"],
    }

    def find_col_idx(col_name: str) -> Optional[int]:
        if col_name in col_name_to_idx:
            return col_name_to_idx[col_name]
        for std, aliases in col_aliases_map.items():
            if col_name in aliases:
                for alias in aliases:
                    if alias in col_name_to_idx:
                        return col_name_to_idx[alias]
        return None

    # 如果已有表中缺少“税率”列，在最后一列后面动态新增（使用全新样式对象避免 StyleProxy unhashable）
    if "税率" not in col_name_to_idx:
        _p_hdr_font = Font(name="微软雅黑", bold=True, size=9)
        _p_hdr_thin = Side(style="thin", color="C0C0C0")
        _p_hdr_border = Border(left=_p_hdr_thin, right=_p_hdr_thin, top=_p_hdr_thin, bottom=_p_hdr_thin)
        _p_hdr_align = Alignment(horizontal="center", vertical="center")
        new_col_idx = ws.max_column + 1
        tax_header_cell = ws.cell(row=header_row_idx, column=new_col_idx, value="税率")
        tax_header_cell.font = _p_hdr_font
        tax_header_cell.border = _p_hdr_border
        tax_header_cell.alignment = _p_hdr_align
        col_name_to_idx["税率"] = new_col_idx
        logger.info(f"已动态新增“税率”列（第 {new_col_idx} 列）到损益毛利明细表")

    data_font = Font(name="微软雅黑", size=9)
    thin = Side(style="thin", color="C0C0C0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    amount_cols_set = {
        "含税销售额", "含税采购成本", "含税差额收入",
        "eas收入不含税", "eas成本不含税", "eas税额", "eas差额收入不含税",
        # 兼容旧格式列
        "调整后gmv", "调整后成本", "调整后毛利", "代运营收入", "代运营成本", "代运营毛利", "合并毛利",
    }
    # 损益毛利明细表的计算列：写入 Excel 公式
    calc_cols_profit = {"含税差额收入", "eas税额", "eas差额收入不含税"}
    from openpyxl.utils import get_column_letter as _get_col_letter

    # 找到已有数据的最后一行（跳过合计行）
    total_row_idx: Optional[int] = None
    for r in range(ws.max_row, header_row_idx, -1):
        first_cell_val = str(ws.cell(row=r, column=1).value or "").strip()
        if first_cell_val == "合计":
            total_row_idx = r
            break

    if total_row_idx is not None:
        next_row = total_row_idx
        ws.insert_rows(next_row, amount=len(new_rows_df))
    else:
        next_row = ws.max_row + 1

    appended_count = 0
    for _, new_row in new_rows_df.iterrows():
        # 首先收集本行各计算基础列的列号（用于构建公式）
        income_no_tax_col_idx = find_col_idx("eas收入不含税")
        cost_no_tax_col_idx = find_col_idx("eas成本不含税")
        taxed_income_col_idx = find_col_idx("含税销售额")
        taxed_cost_col_idx = find_col_idx("含税采购成本")
        income_no_tax_letter = _get_col_letter(income_no_tax_col_idx) if income_no_tax_col_idx else ""
        cost_no_tax_letter = _get_col_letter(cost_no_tax_col_idx) if cost_no_tax_col_idx else ""
        taxed_income_letter = _get_col_letter(taxed_income_col_idx) if taxed_income_col_idx else ""
        taxed_cost_letter = _get_col_letter(taxed_cost_col_idx) if taxed_cost_col_idx else ""

        for col_name, value in new_row.items():
            col_idx = find_col_idx(str(col_name))
            if col_idx is None:
                continue

            cell = ws.cell(row=next_row, column=col_idx)
            cell.font = data_font
            cell.border = border

            if col_name in calc_cols_profit:
                # 计算列：写入 Excel 公式
                if col_name == "含税差额收入" and taxed_income_letter and taxed_cost_letter:
                    cell.value = f'=IF(OR({taxed_income_letter}{next_row}<>"",{taxed_cost_letter}{next_row}<>""),IFERROR({taxed_income_letter}{next_row}-{taxed_cost_letter}{next_row},""),"")'
                elif col_name == "eas税额" and taxed_income_letter and income_no_tax_letter:
                    cell.value = f'=IF(OR({taxed_income_letter}{next_row}<>"",{income_no_tax_letter}{next_row}<>""),IFERROR({taxed_income_letter}{next_row}-{income_no_tax_letter}{next_row},""),"")'
                elif col_name == "eas差额收入不含税" and income_no_tax_letter and cost_no_tax_letter:
                    cell.value = f'=IF(OR({income_no_tax_letter}{next_row}<>"",{cost_no_tax_letter}{next_row}<>""),IFERROR({income_no_tax_letter}{next_row}-{cost_no_tax_letter}{next_row},""),"")'
                else:
                    if value != "":
                        value = clean_amount(value)
                    cell.value = value if value != "" else None
                cell.number_format = "#,##0.00"
                cell.alignment = right_align
            else:
                if col_name in amount_cols_set and value != "":
                    value = clean_amount(value)
                cell.value = value if value != "" else None
                if col_name in amount_cols_set and value != "":
                    cell.number_format = "#,##0.00"
                    cell.alignment = right_align
                else:
                    cell.alignment = left_align
        next_row += 1
        appended_count += 1

    wb.save(output_path)
    logger.info(f"BI 损益毛利明细表已追加 {appended_count} 行并保存: {output_path}")


def write_profit_excel(df: pd.DataFrame, output_path: str) -> None:
    """将 BI 损益毛利明细表写入格式化 Excel（格式参考供应商&代运营毛利表原表）。
    此函数用于兼容旧逻辑（无已有文件时从零创建）。"""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "供应商&代运营毛利表"

    header_font = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
    data_font = Font(name="微软雅黑", size=9)
    sum_font = Font(name="微软雅黑", bold=True, size=10)
    red_font = Font(name="微软雅黑", size=9, color="FF0000")
    red_bold_font = Font(name="微软雅黑", bold=True, size=10, color="FF0000")

    header_fill = PatternFill(fill_type="solid", fgColor="70AD47")
    even_fill = PatternFill(fill_type="solid", fgColor="E2EFDA")
    sum_fill = PatternFill(fill_type="solid", fgColor="A9D18E")

    thin = Side(style="thin", color="C0C0C0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    for col in PROFIT_COLS:
        if col not in df.columns:
            df[col] = ""
    cols = PROFIT_COLS
    amount_cols = {"调整后gmv", "调整后成本", "调整后毛利", "代运营收入", "代运营成本", "代运营毛利", "合并毛利"}
    loss_cols = {"调整后毛利", "代运营毛利", "合并毛利"}
    amount_col_indices = {i + 1: c for i, c in enumerate(cols) if c in amount_cols}

    ws.row_dimensions[1].height = 22
    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        fill = even_fill if row_idx % 2 == 0 else None
        combined = clean_amount(row.get("合并毛利", 0))
        is_loss = combined < 0

        for col_idx, col_name in enumerate(cols, start=1):
            value = row.get(col_name, "")
            if col_name in amount_cols and value != "":
                value = clean_amount(value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if fill:
                cell.fill = fill
            if col_idx in amount_col_indices and value != "":
                cell.number_format = "#,##0.00"
                cell.alignment = right_align
                cell.font = red_font if (is_loss and col_name in loss_cols) else data_font
            else:
                cell.alignment = left_align
                cell.font = data_font

    # 合计行
    if not df.empty:
        sum_row = len(df) + 2
        ws.cell(row=sum_row, column=1, value="合计").font = sum_font
        ws.cell(row=sum_row, column=1).fill = sum_fill
        ws.cell(row=sum_row, column=1).alignment = center_align
        ws.cell(row=sum_row, column=1).border = border

        totals: Dict[str, float] = {}
        for col_name in amount_cols:
            if col_name in df.columns:
                totals[col_name] = round(
                    df[col_name].apply(lambda x: clean_amount(x) if x != "" else 0).sum(), 2
                )

        total_combined = totals.get("合并毛利", 0)
        is_neg = total_combined < 0

        for col_idx, col_name in enumerate(cols, start=1):
            cell = ws.cell(row=sum_row, column=col_idx)
            cell.fill = sum_fill
            cell.border = border
            if col_idx > 1 and col_name in totals:
                cell.value = totals[col_name]
                cell.number_format = "#,##0.00"
                cell.font = red_bold_font if (is_neg and col_name in loss_cols) else sum_font
                cell.alignment = right_align
            elif col_idx > 1:
                cell.alignment = center_align

    for col_idx, col_name in enumerate(cols, start=1):
        sample_lens = [len(str(col_name))]
        for r in range(min(50, len(df))):
            sample_lens.append(len(str(df.iloc[r].get(col_name, ""))))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(sample_lens) * 1.8 + 2, 35)

    wb.save(output_path)
    logger.info(f"BI 损益毛利明细表已保存: {output_path}")


# ── 手工底稿 Excel 写入 ────────────────────────────────────────────────────────

# 费用底稿列：与 extract_expense_from_voucher 输出保持完全一致
BI_EXPENSE_DRAFT_COLS = [
    "日期", "所属公司", "客户", "店铺_平台", "数据来源", "rp单据",
    "行政中心", "利润中心", "eas一级科目", "费用二级项目", "费用三级项目",
    "账单上提取规则", "是否归属店铺费用", "是否参与综合店铺分摊",
    "eas含税金额", "eas不含税金额", "eas税额",
    "费用项目", "报表利润中心id", "报表利润中心", "报表费用类型",
    "是否公摊", "中心群", "摘要备注", "月分区", "备注", "税率",
]


def write_expense_draft_excel(df: pd.DataFrame, output_path: str) -> None:
    """将手工凭证费用底稿写入格式化 Excel（蓝色表头，仅含本次凭证提取的数据）。

    计算列使用 Excel 公式写入：
    - eas税额 = eas含税金额 - eas不含税金额
    """
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "费用手工底稿"

    header_font = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
    data_font = Font(name="微软雅黑", size=9)
    header_fill = PatternFill(fill_type="solid", fgColor="4472C4")
    even_fill = PatternFill(fill_type="solid", fgColor="EBF3FB")
    thin = Side(style="thin", color="C0C0C0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    # 以实际 DataFrame 列为准，BI_EXPENSE_DRAFT_COLS 中未出现的列跳过
    cols = [c for c in BI_EXPENSE_DRAFT_COLS if c in df.columns]
    # 保留 df 中 BI_EXPENSE_DRAFT_COLS 未覆盖的额外列（追加在末尾）
    for c in df.columns:
        if c not in cols and not c.startswith("_"):
            cols.append(c)
    amount_cols = {"eas含税金额", "eas不含税金额", "eas税额"}
    # 计算列：eas税额 = eas含税金额 - eas不含税金额
    calc_cols_expense = {"eas税额"}
    amount_col_indices = {i + 1 for i, c in enumerate(cols) if c in amount_cols}

    # 建立列名到列字母的映射，用于构建公式引用
    col_name_to_letter = {col_name: get_column_letter(i + 1) for i, col_name in enumerate(cols)}

    ws.row_dimensions[1].height = 22
    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        fill = even_fill if row_idx % 2 == 0 else None
        for col_idx, col_name in enumerate(cols, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = border
            if fill:
                cell.fill = fill

            if col_name in calc_cols_expense:
                # 计算列：写入 Excel 公式
                with_tax_letter = col_name_to_letter.get("eas含税金额", "")
                no_tax_letter = col_name_to_letter.get("eas不含税金额", "")
                if col_name == "eas税额" and with_tax_letter and no_tax_letter:
                    cell.value = f'=IF(OR({with_tax_letter}{row_idx}<>"",{no_tax_letter}{row_idx}<>""),IFERROR({with_tax_letter}{row_idx}-{no_tax_letter}{row_idx},""),"")'
                else:
                    value = row.get(col_name, "")
                    if value != "":
                        value = clean_amount(value)
                    cell.value = value if value != "" else None
                cell.number_format = "#,##0.00"
                cell.alignment = right_align
            else:
                value = row.get(col_name, "")
                if col_name in amount_cols:
                    value = clean_amount(value)
                cell.value = value if value != "" else None
                if col_idx in amount_col_indices:
                    cell.number_format = "#,##0.00"
                    cell.alignment = right_align
                else:
                    cell.alignment = left_align

    for col_idx, col_name in enumerate(cols, start=1):
        sample_lens = [len(str(col_name))]
        for r in range(min(50, len(df))):
            sample_lens.append(len(str(df.iloc[r].get(col_name, ""))))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(sample_lens) * 1.8 + 2, 40)

    wb.save(output_path)
    logger.info(f"BI 费用明细手工底稿已保存: {output_path}，共 {len(df)} 行")


def write_profit_draft_excel(df: pd.DataFrame, output_path: str) -> None:
    """将手工凭证损益毛利底稿写入格式化 Excel（绳色表头，仅含本次凭证提取的数据）。

    计算列使用 Excel 公式写入：
    - 含税差额收入 = 含税销售额 - 含税采购成本
    - eas税额 = 含税销售额 - eas收入不含税
    - eas差额收入不含税 = eas收入不含税 - eas成本不含税
    """
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "损益毛利手工底稿"

    header_font = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
    data_font = Font(name="微软雅黑", size=9)
    header_fill = PatternFill(fill_type="solid", fgColor="70AD47")
    even_fill = PatternFill(fill_type="solid", fgColor="E2EFDA")
    thin = Side(style="thin", color="C0C0C0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    cols = [c for c in BI_PROFIT_DETAIL_COLS if c in df.columns]
    for c in df.columns:
        if c not in cols and not c.startswith("_"):
            cols.append(c)
    amount_cols = {
        "含税销售额", "含税采购成本", "含税差额收入",
        "eas收入不含税", "eas成本不含税", "eas税额", "eas差额收入不含税",
    }
    # 计算列：使用 Excel 公式表达
    calc_cols = {"含税差额收入", "eas税额", "eas差额收入不含税"}
    amount_col_indices = {i + 1 for i, c in enumerate(cols) if c in amount_cols}

    # 建立列名到列号的映射，用于构建公式引用
    col_name_to_letter = {col_name: get_column_letter(i + 1) for i, col_name in enumerate(cols)}

    ws.row_dimensions[1].height = 22
    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        fill = even_fill if row_idx % 2 == 0 else None
        for col_idx, col_name in enumerate(cols, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = border
            if fill:
                cell.fill = fill

            if col_name in calc_cols:
                # 计算列：写入 Excel 公式
                taxed_income_col = col_name_to_letter.get("含税销售额", "")
                taxed_cost_col = col_name_to_letter.get("含税采购成本", "")
                income_no_tax_col = col_name_to_letter.get("eas收入不含税", "")
                cost_no_tax_col = col_name_to_letter.get("eas成本不含税", "")

                if col_name == "含税差额收入" and taxed_income_col and taxed_cost_col:
                    formula = f'=IF(OR({taxed_income_col}{row_idx}<>"",{taxed_cost_col}{row_idx}<>""),IFERROR({taxed_income_col}{row_idx}-{taxed_cost_col}{row_idx},""),"")'
                    cell.value = formula
                elif col_name == "eas税额" and taxed_income_col and income_no_tax_col:
                    formula = f'=IF(OR({taxed_income_col}{row_idx}<>"",{income_no_tax_col}{row_idx}<>""),IFERROR({taxed_income_col}{row_idx}-{income_no_tax_col}{row_idx},""),"")'
                    cell.value = formula
                elif col_name == "eas差额收入不含税" and income_no_tax_col and cost_no_tax_col:
                    formula = f'=IF(OR({income_no_tax_col}{row_idx}<>"",{cost_no_tax_col}{row_idx}<>""),IFERROR({income_no_tax_col}{row_idx}-{cost_no_tax_col}{row_idx},""),"")'
                    cell.value = formula
                else:
                    value = row.get(col_name, "")
                    if value != "":
                        value = clean_amount(value)
                    cell.value = value if value != "" else None
                cell.number_format = "#,##0.00"
                cell.alignment = right_align
            else:
                value = row.get(col_name, "")
                if col_name in amount_cols and value != "":
                    value = clean_amount(value)
                cell.value = value if value != "" else None
                if col_idx in amount_col_indices and value != "":
                    cell.number_format = "#,##0.00"
                    cell.alignment = right_align
                else:
                    cell.alignment = left_align

    for col_idx, col_name in enumerate(cols, start=1):
        sample_lens = [len(str(col_name))]
        for r in range(min(50, len(df))):
            sample_lens.append(len(str(df.iloc[r].get(col_name, ""))))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(sample_lens) * 1.8 + 2, 35)

    wb.save(output_path)
    logger.info(f"BI 损益毛利明细手工底稿已保存: {output_path}，共 {len(df)} 行")


# ── 主处理函数 ─────────────────────────────────────────────────────────────────


def process(
    input_files: List[str],
    output_dir: str,
    chat_id: str = "default",
) -> Dict[str, Any]:
    """主处理函数——将手工凭证提取的数据填充到已有 BI 表中，或仅生成手工底稿。

    参数:
        input_files: 用户上传的文件路径列表：
                     - 手工凭证 Excel（必填）
                     - BI费用明细表 Excel（可选）：存在时同步追加 + 生成底稿；缺少时仅生成底稿
                     - BI损益毛利明细表 Excel（可选）：存在时同步追加 + 生成底稿；缺少时仅生成底稿
        output_dir: 输出目录（含 chat_id）
        chat_id: 用户会话 ID

    返回:
        包含状态、文件路径、下载 URL 的字典
    """
    result: Dict[str, Any] = {
        "status": "pending",
        "result_files": [],
        "download_urls": [],
        "metadata": {},
    }

    # ── 1. 识别三类文件
    voucher_file: Optional[str] = None
    bi_expense_file: Optional[str] = None
    bi_profit_file: Optional[str] = None

    for fp in input_files:
        file_type = identify_file_type(fp)
        if file_type == "voucher" and voucher_file is None:
            voucher_file = fp
        elif file_type == "bi_expense" and bi_expense_file is None:
            bi_expense_file = fp
        elif file_type == "bi_profit" and bi_profit_file is None:
            bi_profit_file = fp
        else:
            logger.warning(f"文件识别类型为 '{file_type}'，文件: {Path(fp).name}")

    logger.info(
        f"文件识别结果——手工凭证: {Path(voucher_file).name if voucher_file else '未找到'}，"
        f"BI费用明细表: {Path(bi_expense_file).name if bi_expense_file else '未找到'}，"
        f"BI损益毛利明细表: {Path(bi_profit_file).name if bi_profit_file else '未找到'}"
    )

    # 校验必填文件（仅手工凭证必须存在）
    if voucher_file is None:
        result["status"] = "error"
        result["error"] = {
            "code": "REQUIRED_FILE_MISSING",
            "message": "缺少必要文件：手工凭证文件",
            "suggestion": (
                "请上传手工凭证 Excel（文件名应含「手工凭证」或「凭证」）。\n"
                "如需同步到 BI 报表，请同时上传对应的 BI 报表文件；\n"
                "若只上传手工凭证，则仅生成手工底稿供下载。"
            ),
        }
        return result

    # 记录 BI 报表的缺失情况（仅警告，不阻断流程）
    if bi_expense_file is None:
        logger.info("未上传 BI费用明细表，将跳过同步操作，仅生成费用手工底稿")
    if bi_profit_file is None:
        logger.info("未上传 BI损益毛利明细表，将跳过同步操作，仅生成损益毛利手工底稿")

    # ── 2. 读取手工凭证数据
    logger.info(f"读取手工凭证: {voucher_file}")
    try:
        df = read_voucher_file(voucher_file)
    except Exception as e:
        result["status"] = "error"
        result["error"] = {
            "code": "VOUCHER_FORMAT_ERROR",
            "message": f"手工凭证文件读取失败: {str(e)}",
            "suggestion": "请检查文件格式，需包含：科目名称、借方发生额、贷方发生额、摘要、凭证号等字段",
        }
        return result

    logger.info(f"凭证读取完成，共 {len(df)} 行，列: {list(df.columns)}")
    result["metadata"]["record_count"] = len(df)
    result["metadata"]["source_file"] = Path(voucher_file).name
    result["metadata"]["bi_expense_file"] = Path(bi_expense_file).name if bi_expense_file else "（未上传）"
    result["metadata"]["bi_profit_file"] = Path(bi_profit_file).name if bi_profit_file else "（未上传）"

    # ── 3. 准备输出目录
    os.makedirs(output_dir, exist_ok=True)
    expense_output = str(Path(output_dir) / "BI费用明细表.xlsx")
    profit_output = str(Path(output_dir) / "BI损益毛利明细表.xlsx")
    expense_draft_output = str(Path(output_dir) / "BI费用明细表(手工底稿).xlsx")
    profit_draft_output = str(Path(output_dir) / "BI损益毛利明细表(手工底稿).xlsx")

    # ── 4. 提取费用数据 → 同步到 BI 费用明细表（可选）+ 生成手工底稿
    logger.info("开始提取费用数据（章节四+五）...")
    try:
        expense_df = extract_expense_from_voucher(df)
        logger.info(f"提取费用明细: {len(expense_df)} 条")
        if expense_df.empty:
            logger.warning("费用明细数据为空，凭证中可能无费用类科目")

        if bi_expense_file is not None:
            # 有 BI 费用明细表：同步追加
            append_to_existing_expense_excel(
                existing_file=bi_expense_file,
                new_rows_df=expense_df,
                output_path=expense_output,
            )
            result["result_files"].append(expense_output)
            result["download_urls"].append(f"{BASE_URL}/result/{chat_id}/BI费用明细表.xlsx")
            result["metadata"]["expense_appended_count"] = len(expense_df)
            logger.info(f"费用数据已同步到 BI 费用明细表: {expense_output}")
        else:
            # 无 BI 费用明细表：跳过同步
            result["metadata"]["expense_appended_count"] = 0
            logger.info("未上传 BI费用明细表，跳过费用同步")

        # 手工底稿：无论是否同步，只要有数据就生成
        if not expense_df.empty:
            write_expense_draft_excel(expense_df, expense_draft_output)
            result["result_files"].append(expense_draft_output)
            result["download_urls"].append(f"{BASE_URL}/result/{chat_id}/BI费用明细表(手工底稿).xlsx")
            logger.info(f"费用手工底稿已生成: {expense_draft_output}")
    except Exception as e:
        logger.error(f"费用数据处理失败: {e}\n{traceback.format_exc()}")
        result.setdefault("errors", []).append(f"费用数据处理失败: {str(e)}")

    # ── 5. 提取收入成本数据 → 同步到 BI 损益毛利明细表（可选）+ 生成手工底稿
    logger.info("开始提取收入成本数据（章节一）...")
    try:
        profit_rows_df = extract_profit_detail_from_voucher(df)
        logger.info(f"损益毛利明细（按行）: {len(profit_rows_df)} 条")
        if profit_rows_df.empty:
            logger.warning("毛利数据为空，凭证中可能无收入/成本类科目")

        if bi_profit_file is not None:
            # 有 BI 损益毛利明细表：同步追加
            append_to_existing_profit_excel(
                existing_file=bi_profit_file,
                new_rows_df=profit_rows_df,
                output_path=profit_output,
            )
            result["result_files"].append(profit_output)
            result["download_urls"].append(f"{BASE_URL}/result/{chat_id}/BI损益毛利明细表.xlsx")
            result["metadata"]["profit_appended_count"] = len(profit_rows_df)
            logger.info(f"损益毛利数据已同步到 BI 损益毛利明细表: {profit_output}")
        else:
            # 无 BI 损益毛利明细表：跳过同步
            result["metadata"]["profit_appended_count"] = 0
            logger.info("未上传 BI损益毛利明细表，跳过损益毛利同步")

        # 手工底稿：无论是否同步，只要有数据就生成
        if not profit_rows_df.empty:
            write_profit_draft_excel(profit_rows_df, profit_draft_output)
            result["result_files"].append(profit_draft_output)
            result["download_urls"].append(f"{BASE_URL}/result/{chat_id}/BI损益毛利明细表(手工底稿).xlsx")
            logger.info(f"损益毛利手工底稿已生成: {profit_draft_output}")
    except Exception as e:
        logger.error(f"损益毛利数据处理失败: {e}\n{traceback.format_exc()}")
        result.setdefault("errors", []).append(f"损益毛利数据处理失败: {str(e)}")

    # ── 6. 最终状态
    if result["result_files"]:
        result["status"] = "success" if "errors" not in result else "partial_success"
    else:
        result["status"] = "error"
        result.setdefault("error", {
            "code": "OUTPUT_GEN_FAILED",
            "message": "报表填充失败，未产生任何输出文件",
            "suggestion": "请检查凭证数据是否包含费用/收入/成本类科目（如销售费用、管理费用、主营业务收入等）",
        })

    return result

# ── 命令行入口 ─────────────────────────────────────────────────────────────────


def main():
    """命令行入口，供 skill_handler 调用。"""
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
    )

    parser = argparse.ArgumentParser(description="核算报表生成：手工凭证 → BI费用明细表 + BI损益毛利明细表")
    parser.add_argument("--input", action="append", default=[], dest="inputs",
                        help="输入文件路径（可多次指定）")
    parser.add_argument("--output-dir", default=str(DEFAULT_RESULT_DIR / "default"),
                        help="输出目录")
    parser.add_argument("--chat-id", default="default", help="用户会话 ID")
    args = parser.parse_args()

    if not args.inputs:
        env_inputs = os.getenv("INPUT_FILES", "")
        if env_inputs:
            args.inputs = [f.strip() for f in env_inputs.split(",") if f.strip()]

    if not args.inputs:
        print(json.dumps({
            "status": "error",
            "error": {"code": "NO_INPUT", "message": "未指定输入文件，请使用 --input 参数"},
        }, ensure_ascii=False))
        sys.exit(1)

    chat_id = args.chat_id
    if chat_id == "default" and "/" in args.output_dir:
        chat_id = Path(args.output_dir).name

    result = process(
        input_files=args.inputs,
        output_dir=args.output_dir,
        chat_id=chat_id,
    )

    print(json.dumps(result, ensure_ascii=False, indent=2))

    if result["status"] in ("success", "partial_success"):
        for url in result.get("download_urls", []):
            logger.info(f"下载地址: {url}")
        sys.exit(0)
    else:
        sys.exit(1)


if __name__ == "__main__":
    main()
