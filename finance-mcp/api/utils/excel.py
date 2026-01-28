"""
Excel file processing utilities
"""
from openpyxl import load_workbook
from typing import List, Dict, Any
from pathlib import Path


def parse_excel_file(file_path: str, max_rows: int = 100) -> Dict[str, Any]:
    """
    Parse Excel file and extract preview data

    Args:
        file_path: Path to Excel file
        max_rows: Maximum number of rows to preview (default: 100)

    Returns:
        Dictionary containing file metadata and sheet data
    """
    file_path = Path(file_path)

    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)

    sheets_data = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Get all rows (limited to max_rows)
        rows = []
        headers = []

        for idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if idx == 0:
                # First row as headers
                headers = [str(cell) if cell is not None else f"Column_{i+1}"
                          for i, cell in enumerate(row)]
            elif idx <= max_rows:
                # Convert row to list of strings
                row_data = [str(cell) if cell is not None else "" for cell in row]
                rows.append(row_data)
            else:
                break

        # Get total row count
        total_rows = sheet.max_row - 1  # Exclude header row

        sheets_data.append({
            "name": sheet_name,
            "headers": headers,
            "rows": rows,
            "total_rows": total_rows,
            "total_columns": len(headers)
        })

    workbook.close()

    return {
        "filename": file_path.name,
        "sheets": sheets_data
    }


def get_sheet_names(file_path: str) -> List[str]:
    """
    Get list of sheet names from Excel file

    Args:
        file_path: Path to Excel file

    Returns:
        List of sheet names
    """
    workbook = load_workbook(filename=file_path, read_only=True)
    sheet_names = workbook.sheetnames
    workbook.close()
    return sheet_names
