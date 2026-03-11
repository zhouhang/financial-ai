"""
数据整理规则同步执行工具

根据 bus_proc_rules 表中存储的 JSON 规则，对上传文件执行字段映射、数据转换，
生成新的输出文件（xlsx）。

支持的 rule_type：
  - direct_mapping      : 直接从源字段取值
  - constant            : 常量值（可为 null）
  - extract             : 按分隔符提取第 N 级子串
  - formula             : 数学公式计算（依赖其他目标字段）
  - parse_from_field    : 多步骤解析字段值
  - conditional_value   : 条件匹配取固定值
  - conditional_formula : 条件匹配取公式结果
  - lookup              : 从 lookup_table 查找对应值（运行时查表）
"""
from __future__ import annotations

import json
import logging
import os
import re
from pathlib import Path
from typing import Any, Optional

import pandas as pd
from mcp import Tool

logger = logging.getLogger(__name__)


# ════════════════════════════════════════════════════════════════════════════
# Tool 定义
# ════════════════════════════════════════════════════════════════════════════

def create_sync_rule_tools() -> list[Tool]:
    """创建数据同步规则工具列表"""
    return [
        Tool(
            name="sync_rule_execute",
            description=(
                "根据规则编码（rule_code）从数据库获取数据整理规则，"
                "对上传文件执行字段映射和数据转换，生成目标 Excel 文件。\n"
                "uploaded_files 格式为文件校验工具（validate_uploaded_files）返回的 matched_results。"
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "uploaded_files": {
                        "type": "array",
                        "description": (
                            "文件校验结果列表，每个元素格式：\n"
                            "{ file_name: 文件名, file_path: 绝对路径, table_id: 表ID, table_name: 表名 }"
                        ),
                        "items": {
                            "type": "object",
                            "properties": {
                                "file_name": {"type": "string"},
                                "file_path": {"type": "string"},
                                "table_id": {"type": "string"},
                                "table_name": {"type": "string"},
                            },
                            "required": ["file_name", "table_name"],
                        },
                    },
                    "rule_code": {
                        "type": "string",
                        "description": "整理规则编码，用于从 bus_proc_rules 表中获取规则 JSON",
                    },
                    "output_dir": {
                        "type": "string",
                        "description": "生成文件的存放目录（绝对路径）。若目录不存在会自动创建。",
                    },
                },
                "required": ["uploaded_files", "rule_code", "output_dir"],
            },
        )
    ]


# ════════════════════════════════════════════════════════════════════════════
# Tool 调用入口
# ════════════════════════════════════════════════════════════════════════════

async def handle_sync_rule_tool_call(name: str, arguments: dict) -> dict:
    """处理 sync_rule_execute 工具调用"""
    if name == "sync_rule_execute":
        return await _handle_sync_rule_execute(arguments)
    return {"success": False, "error": f"未知工具: {name}"}


async def _handle_sync_rule_execute(arguments: dict) -> dict:
    """执行数据整理规则，生成输出文件"""
    from proc.mcp_server.tools import _get_proc_rule

    uploaded_files: list[dict] = arguments.get("uploaded_files") or []
    rule_code: str = (arguments.get("rule_code") or "").strip()
    output_dir: str = (arguments.get("output_dir") or "").strip()

    # ── 参数校验 ──────────────────────────────────────────────────────────────
    if not uploaded_files:
        return {"success": False, "error": "uploaded_files 不能为空"}
    if not rule_code:
        return {"success": False, "error": "rule_code 不能为空"}
    if not output_dir:
        return {"success": False, "error": "output_dir 不能为空"}

    # ── 创建输出目录 ──────────────────────────────────────────────────────────
    try:
        Path(output_dir).mkdir(parents=True, exist_ok=True)
    except Exception as e:
        return {"success": False, "error": f"创建输出目录失败: {e}"}

    # ── 获取整理规则 ──────────────────────────────────────────────────────────
    rule_record = _get_proc_rule(rule_code)
    if rule_record is None:
        return {"success": False, "error": f"未找到 rule_code='{rule_code}' 的整理规则"}

    rule_data: dict = rule_record.get("rule") or {}
    rules_list: list[dict] = rule_data.get("rules", [])
    if not rules_list:
        return {"success": False, "error": f"规则 '{rule_code}' 中未定义任何 rules 项"}

    # ── 构建 table_name → file_path 映射 ─────────────────────────────────────
    table_file_map: dict[str, str] = {}
    for item in uploaded_files:
        tname = (item.get("table_name") or "").strip()
        fpath = (item.get("file_path") or item.get("file_name") or "").strip()
        if tname and fpath:
            table_file_map[tname] = fpath

    logger.info(
        f"[sync_rule] 开始执行，rule_code={rule_code!r}，"
        f"文件映射={list(table_file_map.keys())}，rules数量={len(rules_list)}"
    )

    # ── 逐条执行规则 ──────────────────────────────────────────────────────────
    generated_files: list[dict] = []
    errors: list[str] = []

    for rule in rules_list:
        try:
            result = _execute_single_rule(rule, table_file_map, output_dir)
            generated_files.append(result)
            logger.info(
                f"[sync_rule] 规则 {rule.get('rule_id')!r} 执行成功，"
                f"输出文件：{result['output_file']}"
            )
        except Exception as e:
            msg = f"规则 {rule.get('rule_id')!r} 执行失败: {e}"
            logger.error(f"[sync_rule] {msg}", exc_info=True)
            errors.append(msg)

    return {
        "success": len(errors) == 0,
        "rule_code": rule_code,
        "generated_files": generated_files,
        "generated_count": len(generated_files),
        "errors": errors,
        "message": (
            f"成功生成 {len(generated_files)} 个文件"
            + (f"，{len(errors)} 个规则执行失败" if errors else "")
        ),
    }


# ════════════════════════════════════════════════════════════════════════════
# 单条规则执行引擎
# ════════════════════════════════════════════════════════════════════════════

def _execute_single_rule(rule: dict, table_file_map: dict[str, str], output_dir: str) -> dict:
    """
    执行单条规则，返回生成文件的信息。

    Args:
        rule: 单条规则定义（来自 rules 数组的一个元素）
        table_file_map: table_name → 文件绝对路径 的映射
        output_dir: 输出目录

    Returns:
        { rule_id, output_file, row_count }
    """
    rule_id: str = rule.get("rule_id", "UNKNOWN")
    source_tables = rule.get("source_tables") or rule.get("source_table") or ""
    target_table: str = rule.get("target_table", rule_id)
    field_mappings: list[dict] = rule.get("field_mappings", [])
    global_filter: Optional[dict] = rule.get("global_filter")
    lookup_tables_def: list[dict] = rule.get("lookup_tables", [])

    # ── 1. 读取源表数据 ──────────────────────────────────────────────────────
    source_df = _load_source_df(source_tables, table_file_map)

    # ── 2. 应用全局过滤 ──────────────────────────────────────────────────────
    if global_filter:
        source_df = _apply_global_filter(source_df, global_filter)

    logger.info(f"[sync_rule] [{rule_id}] 源数据过滤后行数：{len(source_df)}")

    # ── 3. 加载 lookup 表（如果有）──────────────────────────────────────────
    lookup_data: dict[str, pd.DataFrame] = {}
    for lt in lookup_tables_def:
        lt_name = lt.get("table_name", "")
        if lt_name in table_file_map:
            try:
                lookup_data[lt_name] = _read_file_as_df(table_file_map[lt_name])
                logger.info(f"[sync_rule] [{rule_id}] 加载 lookup 表 '{lt_name}'，{len(lookup_data[lt_name])} 行")
            except Exception as e:
                logger.warning(f"[sync_rule] [{rule_id}] 加载 lookup 表 '{lt_name}' 失败: {e}")

    # ── 4. 按 field_mappings 构建目标 DataFrame ──────────────────────────────
    result_df = _apply_field_mappings(source_df, field_mappings, lookup_data, rule_id)

    # ── 5. 写出 Excel 文件 ───────────────────────────────────────────────────
    safe_target = re.sub(r'[\\/:*?"<>|]', "_", target_table)
    output_filename = f"{rule_id}_{safe_target}.xlsx"
    output_path = str(Path(output_dir) / output_filename)

    _write_excel(result_df, field_mappings, output_path)

    return {
        "rule_id": rule_id,
        "target_table": target_table,
        "output_file": output_path,
        "row_count": len(result_df),
    }


# ════════════════════════════════════════════════════════════════════════════
# 数据加载工具
# ════════════════════════════════════════════════════════════════════════════

def _load_source_df(source_tables: Any, table_file_map: dict[str, str]) -> pd.DataFrame:
    """
    根据 source_tables 字段找到对应文件并加载为 DataFrame。
    source_tables 可以是字符串（单表）或列表（多表合并）。
    """
    if isinstance(source_tables, list):
        table_names = source_tables
    else:
        table_names = [str(source_tables)]

    dfs: list[pd.DataFrame] = []
    for tname in table_names:
        tname = tname.strip()
        if tname not in table_file_map:
            raise ValueError(f"源表 '{tname}' 未在上传文件中找到，可用表：{list(table_file_map.keys())}")
        df = _read_file_as_df(table_file_map[tname])
        dfs.append(df)

    return pd.concat(dfs, ignore_index=True) if len(dfs) > 1 else dfs[0]


def _read_file_as_df(file_path: str) -> pd.DataFrame:
    """读取 CSV 或 Excel 文件为 DataFrame"""
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"文件不存在: {file_path}")
    ext = path.suffix.lower()
    if ext == ".csv":
        try:
            return pd.read_csv(file_path, encoding="utf-8-sig")
        except UnicodeDecodeError:
            import chardet
            with open(file_path, "rb") as f:
                enc = chardet.detect(f.read()).get("encoding", "gbk")
            return pd.read_csv(file_path, encoding=enc)
    elif ext in (".xlsx", ".xls"):
        return pd.read_excel(file_path)
    else:
        raise ValueError(f"不支持的文件格式: {ext}")


# ════════════════════════════════════════════════════════════════════════════
# 全局过滤
# ════════════════════════════════════════════════════════════════════════════

def _apply_global_filter(df: pd.DataFrame, gf: dict) -> pd.DataFrame:
    """
    应用 global_filter 定义的过滤规则。

    支持 operator：
      - "in"          : source_column 的值在 values 列表中
      - "starts_with" : source_column 的值以 values 中任意一项开头
    同时排除 exclude_values（若定义）。
    """
    col: str = gf.get("source_column", "")
    operator: str = gf.get("operator", "in")
    values: list = gf.get("values", [])
    exclude_values: list = gf.get("exclude_values", [])

    if col not in df.columns:
        logger.warning(f"[sync_rule] global_filter 列 '{col}' 不存在，跳过过滤")
        return df

    col_series = df[col].astype(str)

    if operator == "in":
        mask = col_series.isin(values)
    elif operator == "starts_with":
        pattern = "|".join(re.escape(v) for v in values)
        mask = col_series.str.match(f"^({pattern})")
    else:
        logger.warning(f"[sync_rule] 未知 global_filter operator: {operator}，跳过过滤")
        return df

    if exclude_values:
        exclude_mask = col_series.isin(exclude_values)
        mask = mask & ~exclude_mask

    return df[mask].reset_index(drop=True)


# ════════════════════════════════════════════════════════════════════════════
# 字段映射主流程
# ════════════════════════════════════════════════════════════════════════════

def _apply_field_mappings(
    source_df: pd.DataFrame,
    field_mappings: list[dict],
    lookup_data: dict[str, pd.DataFrame],
    rule_id: str,
) -> pd.DataFrame:
    """
    按照 field_mappings 顺序构建目标 DataFrame 的列。

    计算列（is_computed=True 或 rule_type=formula/conditional_formula 且依赖其他目标列）
    需要在普通列全部生成后再计算。
    """
    result: dict[str, Any] = {}
    n_rows = len(source_df)

    # 第一轮：处理非计算列
    deferred: list[dict] = []
    for fm in field_mappings:
        target = fm.get("target_field", "")
        rule_type = fm.get("rule_type", "")

        # 纯计算列（依赖其他目标列）延后处理
        if fm.get("is_computed") or (rule_type in ("formula", "conditional_formula") and fm.get("depends_on")):
            deferred.append(fm)
            continue

        try:
            col_data = _compute_column(fm, source_df, result, lookup_data)
            result[target] = col_data
        except Exception as e:
            logger.warning(f"[sync_rule] [{rule_id}] 字段 '{target}' 计算失败，填充空值: {e}")
            result[target] = [None] * n_rows

    # 第二轮：处理计算列（多轮直到收敛，最多 10 轮防死循环）
    max_rounds = 10
    for _ in range(max_rounds):
        if not deferred:
            break
        still_deferred: list[dict] = []
        for fm in deferred:
            target = fm.get("target_field", "")
            depends_on: list[str] = fm.get("depends_on", [])
            # 检查依赖是否已全部就绪
            if all(d in result for d in depends_on):
                try:
                    col_data = _compute_column(fm, source_df, result, lookup_data)
                    result[target] = col_data
                except Exception as e:
                    logger.warning(f"[sync_rule] [{rule_id}] 计算列 '{target}' 失败，填充空值: {e}")
                    result[target] = [None] * n_rows
            else:
                still_deferred.append(fm)
        if len(still_deferred) == len(deferred):
            # 无进展，剩余列填充空值
            for fm in still_deferred:
                result[fm.get("target_field", "")] = [None] * n_rows
            break
        deferred = still_deferred

    # 按 field_mappings 顺序输出列
    ordered_cols = [fm.get("target_field", "") for fm in field_mappings if fm.get("target_field")]
    out_dict = {col: result.get(col, [None] * n_rows) for col in ordered_cols}
    return pd.DataFrame(out_dict)


# ════════════════════════════════════════════════════════════════════════════
# 各 rule_type 处理器
# ════════════════════════════════════════════════════════════════════════════

def _compute_column(
    fm: dict,
    source_df: pd.DataFrame,
    result: dict[str, Any],
    lookup_data: dict[str, pd.DataFrame],
) -> list:
    """根据单个 field_mapping 计算目标列，返回与源 DataFrame 等长的列表"""
    rule_type = fm.get("rule_type", "")
    target = fm.get("target_field", "")
    n = len(source_df)

    if rule_type == "direct_mapping":
        return _rt_direct_mapping(fm, source_df, n)

    elif rule_type == "constant":
        return _rt_constant(fm, n)

    elif rule_type == "extract":
        return _rt_extract(fm, source_df, n)

    elif rule_type == "formula":
        return _rt_formula(fm, source_df, result, n)

    elif rule_type == "parse_from_field":
        return _rt_parse_from_field(fm, source_df, n)

    elif rule_type == "conditional_value":
        return _rt_conditional_value(fm, source_df, n)

    elif rule_type == "conditional_formula":
        return _rt_conditional_formula(fm, source_df, result, n)

    elif rule_type == "lookup":
        return _rt_lookup(fm, source_df, result, lookup_data, n)

    else:
        logger.warning(f"[sync_rule] 未知 rule_type='{rule_type}'，目标列 '{target}' 填充空值")
        return [None] * n


# ── direct_mapping ───────────────────────────────────────────────────────────

def _rt_direct_mapping(fm: dict, source_df: pd.DataFrame, n: int) -> list:
    src = fm.get("source_field", "")
    if src not in source_df.columns:
        logger.warning(f"[sync_rule] direct_mapping 源字段 '{src}' 不存在")
        return [None] * n
    return source_df[src].tolist()


# ── constant ─────────────────────────────────────────────────────────────────

def _rt_constant(fm: dict, n: int) -> list:
    return [fm.get("value")] * n


# ── extract ──────────────────────────────────────────────────────────────────

def _rt_extract(fm: dict, source_df: pd.DataFrame, n: int) -> list:
    """按分隔符提取第 N 级子串（1-indexed）"""
    src = fm.get("source_field", "")
    delimiter = fm.get("delimiter", "_")
    level = int(fm.get("extract_level", 1)) - 1  # 转为 0-indexed

    if src not in source_df.columns:
        return [None] * n

    def extract(val: Any) -> Any:
        if pd.isna(val):
            return None
        parts = str(val).split(delimiter)
        return parts[level] if level < len(parts) else None

    return source_df[src].apply(extract).tolist()


# ── formula ──────────────────────────────────────────────────────────────────

def _rt_formula(
    fm: dict,
    source_df: pd.DataFrame,
    result: dict[str, Any],
    n: int,
) -> list:
    """
    简单数学公式计算。支持四则运算。
    操作数可以是源 DataFrame 中的列，也可以是已计算的目标列。
    """
    formula: str = fm.get("formula", "")
    source_fields: list[dict] = fm.get("source_fields", [])

    # 收集操作数名称（来自 source_fields 定义 或 depends_on）
    depends_on: list[str] = fm.get("depends_on", [])
    src_field_names: list[str] = [sf.get("field", "") for sf in source_fields]

    results_list: list[Any] = []
    for i in range(n):
        env: dict[str, Any] = {}
        # 填入源 DataFrame 列值
        for col in source_df.columns:
            env[col] = _safe_num(source_df.iloc[i].get(col))
        # 填入已计算的目标列值
        for col_name, col_data in result.items():
            if isinstance(col_data, list) and i < len(col_data):
                env[col_name] = _safe_num(col_data[i])
        try:
            val = _eval_formula(formula, env)
        except Exception as e:
            logger.debug(f"[sync_rule] formula 计算失败 row={i}: {e}")
            val = None
        results_list.append(val)

    return results_list


# ── parse_from_field ─────────────────────────────────────────────────────────

def _rt_parse_from_field(fm: dict, source_df: pd.DataFrame, n: int) -> list:
    """多步骤解析字段值（见 manual_voucher_sync_rule.json 中的税率解析）"""
    src = fm.get("source_field", "")
    parse_rules: list[dict] = fm.get("parse_rules", [])

    if src not in source_df.columns:
        return [None] * n

    def parse_one(raw: Any) -> Any:
        if pd.isna(raw):
            return None
        current: Any = str(raw)
        for pr in parse_rules:
            split_by = pr.get("split_by", "")
            extract_index = int(pr.get("extract_index", 0))
            fallback = pr.get("fallback")
            if split_by:
                parts = current.split(split_by)
                if extract_index < len(parts):
                    current = parts[extract_index].strip()
                else:
                    current = fallback
                    break
            else:
                break
        # 尝试转为数字（税率场景）
        if current is None or current == "":
            return fallback if fallback is not None else None
        try:
            return float(current)
        except (ValueError, TypeError):
            return current

    return source_df[src].apply(parse_one).tolist()


# ── conditional_value ────────────────────────────────────────────────────────

def _rt_conditional_value(fm: dict, source_df: pd.DataFrame, n: int) -> list:
    """
    根据条件映射固定值。
    conditions 顺序匹配，第一个匹配的 condition 取其 value。
    最后一个 condition 为 "其他" 或 "other" 时作为默认值。
    """
    src = fm.get("source_field", "")
    conditions: list[dict] = fm.get("conditions", [])

    if src not in source_df.columns:
        return [None] * n

    def match_one(val: Any) -> Any:
        val_str = str(val) if not pd.isna(val) else ""
        for cond in conditions:
            condition_str = cond.get("condition", "")
            cond_value = cond.get("value")
            # 默认条件
            if condition_str in ("其他", "other", "其它", "default"):
                return cond_value
            # 等于条件
            if "等于" in condition_str:
                m = re.search(r"['\"'\"](.*?)['\"'\"]", condition_str)
                if m and val_str == m.group(1):
                    return cond_value
            # 包含条件
            elif "包含" in condition_str or "为" in condition_str:
                targets = re.findall(r"['\"'\"](.*?)['\"'\"]", condition_str)
                if any(t in val_str for t in targets):
                    return cond_value
            # 科目名称百分比税率 — 直接匹配 condition 文本中的科目
            elif "或" in condition_str:
                targets = re.findall(r"['\"'\"](.*?)['\"'\"]", condition_str)
                if val_str in targets:
                    return cond_value
        return None

    return source_df[src].apply(match_one).tolist()


# ── conditional_formula ──────────────────────────────────────────────────────

def _rt_conditional_formula(
    fm: dict,
    source_df: pd.DataFrame,
    result: dict[str, Any],
    n: int,
) -> list:
    """
    根据条件选择不同公式计算。
    若条件描述无法严格解析，则取第一个条件的公式作为默认。
    """
    src = fm.get("source_field", "")
    conditions: list[dict] = fm.get("conditions", [])

    results_list: list[Any] = []
    for i in range(n):
        row_src_val = source_df.iloc[i].get(src, "") if src in source_df.columns else ""
        row_src_str = str(row_src_val) if not pd.isna(row_src_val) else ""

        chosen_formula: Optional[str] = None
        for cond in conditions:
            condition_str = cond.get("condition", "")
            formula_expr = cond.get("formula", "")
            # 简单匹配：condition 文本中引号内的科目名与当前值对比
            quoted = re.findall(r"['\"'\"](.*?)['\"'\"]", condition_str)
            if not quoted:
                # 无法解析条件，作为默认
                chosen_formula = formula_expr
                break
            if any(q in row_src_str or row_src_str.startswith(q) for q in quoted):
                chosen_formula = formula_expr
                break

        if chosen_formula is None and conditions:
            chosen_formula = conditions[0].get("formula", "")

        if chosen_formula:
            env: dict[str, Any] = {}
            for col in source_df.columns:
                env[col] = _safe_num(source_df.iloc[i].get(col))
            for col_name, col_data in result.items():
                if isinstance(col_data, list) and i < len(col_data):
                    env[col_name] = _safe_num(col_data[i])
            try:
                val = _eval_formula(chosen_formula, env)
            except Exception:
                val = None
        else:
            val = None
        results_list.append(val)

    return results_list


# ── lookup ────────────────────────────────────────────────────────────────────

def _rt_lookup(
    fm: dict,
    source_df: pd.DataFrame,
    result: dict[str, Any],
    lookup_data: dict[str, pd.DataFrame],
    n: int,
) -> list:
    """
    从 lookup_table 中查找对应值。
    若 lookup_table 未加载，返回 no_match_result。
    """
    lookup_table_name: str = fm.get("lookup_table", "")
    lookup_key: str = fm.get("lookup_key") or fm.get("lookup_field", "")
    source_field: str = fm.get("source_field") or ""
    return_field: str = fm.get("return_field", "")
    match_result = fm.get("match_result", True)
    no_match_result = fm.get("no_match_result", None)

    # 源值来自已计算的目标列（lookup_field 指向目标列）或源 DataFrame
    def get_source_val(i: int) -> Any:
        # 优先从已计算列取
        if source_field in result:
            data = result[source_field]
            return data[i] if i < len(data) else None
        # 其次从源 DataFrame 取
        if source_field in source_df.columns:
            return source_df.iloc[i].get(source_field)
        return None

    if lookup_table_name not in lookup_data:
        logger.warning(f"[sync_rule] lookup 表 '{lookup_table_name}' 未加载，填充默认值")
        return [no_match_result] * n

    ldf = lookup_data[lookup_table_name]

    results_list: list[Any] = []
    for i in range(n):
        val = get_source_val(i)
        if val is None or (isinstance(val, float) and pd.isna(val)):
            results_list.append(no_match_result)
            continue

        val_str = str(val)
        # 判断是否命中（match_result 场景：判断是否存在）
        if return_field:
            # 取对应字段值
            if lookup_key and lookup_key in ldf.columns and return_field in ldf.columns:
                match_rows = ldf[ldf[lookup_key].astype(str) == val_str]
                if not match_rows.empty:
                    results_list.append(match_rows.iloc[0][return_field])
                else:
                    results_list.append(no_match_result)
            else:
                results_list.append(no_match_result)
        else:
            # 只判断是否存在（返回 match_result / no_match_result）
            if lookup_key and lookup_key in ldf.columns:
                exists = val_str in ldf[lookup_key].astype(str).values
                results_list.append(match_result if exists else no_match_result)
            else:
                # 任意列中查找
                exists = any(val_str in ldf[c].astype(str).values for c in ldf.columns)
                results_list.append(match_result if exists else no_match_result)

    return results_list


# ════════════════════════════════════════════════════════════════════════════
# 辅助工具
# ════════════════════════════════════════════════════════════════════════════

def _safe_num(val: Any) -> Any:
    """将值转为数字，无法转换则返回 0"""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return 0 if pd.isna(val) else val
    try:
        return float(str(val).strip())
    except (ValueError, TypeError):
        return 0


def _eval_formula(formula: str, env: dict[str, Any]) -> Any:
    """
    安全地计算简单数学公式。
    只允许四则运算和括号，操作数为 env 中定义的变量名。
    """
    # 将变量名替换为数值字符串
    expr = formula
    # 从长到短排序，避免短名称替换了长名称的一部分
    for var_name in sorted(env.keys(), key=len, reverse=True):
        val = env.get(var_name, 0)
        num_str = str(float(val)) if val is not None else "0"
        # 只替换完整的变量名（非字母数字边界）
        expr = re.sub(r'(?<![a-zA-Z0-9_])' + re.escape(var_name) + r'(?![a-zA-Z0-9_])', num_str, expr)

    # 安全检查：只允许数字、运算符、括号和空格
    if not re.match(r'^[\d\s\+\-\*/\.\(\)]+$', expr.strip()):
        raise ValueError(f"公式含有非法字符: {expr!r}")

    return eval(expr)  # nosec: 已限制只含数字和运算符


def _write_excel(df: pd.DataFrame, field_mappings: list[dict], output_path: str) -> None:
    """
    将 DataFrame 写出为 xlsx。
    计算列（is_computed=True）设置背景色以便区分。
    """
    computed_cols = {
        fm.get("target_field", "")
        for fm in field_mappings
        if fm.get("is_computed")
    }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")
        if computed_cols:
            from openpyxl.styles import PatternFill
            ws = writer.sheets["Sheet1"]
            light_yellow = PatternFill(start_color="FFFFD700", end_color="FFFFD700", fill_type="solid")
            header_row = {cell.value: cell.column for cell in ws[1]}
            for col_name, col_idx in header_row.items():
                if col_name in computed_cols:
                    for row_idx in range(2, ws.max_row + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = light_yellow
