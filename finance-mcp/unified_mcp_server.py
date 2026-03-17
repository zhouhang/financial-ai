"""
Financial Agent Unified MCP Server
统一的财务助手 MCP 服务器 - 包含对账和数据整理功能
"""
import os
import sys
import asyncio
from pathlib import Path
from typing import Optional
from dotenv import load_dotenv
from mcp import types

# 加载环境变量
load_dotenv()
from mcp.server import Server
from mcp.server.sse import SseServerTransport
from starlette.applications import Starlette
from starlette.routing import Route, Mount
from starlette.responses import Response, JSONResponse, FileResponse
from starlette.staticfiles import StaticFiles
import uvicorn
import logging

# 导入安全工具
from security_utils import validate_task_id, sanitize_path

# MCP 服务公开访问地址（用于生成下载链接）
MCP_PUBLIC_BASE_URL = os.getenv("MCP_PUBLIC_BASE_URL", "http://localhost:3335")

# 导入对账模块
from reconciliation.mcp_server.config import DEFAULT_HOST, DEFAULT_PORT
from reconciliation.mcp_server.tools import create_tools as create_recon_tools, handle_tool_call as handle_recon_call

# 导入数据整理模块
from data_preparation.mcp_server.tools import create_tools as create_prep_tools, handle_tool_call as handle_prep_call
from data_preparation.mcp_server.config import OUTPUT_DIR, REPORT_DIR as PREP_REPORT_DIR_IMPORT

# 导入认证和规则管理模块
from auth.tools import create_auth_tools, handle_auth_tool_call, _create_guest_tools, _handle_create_guest_token, _handle_verify_guest_token, _handle_list_recommended_rules

# 导入 tools 模块（文件校验和数据同步）
from tools.file_validate_tool import create_file_validate_tools, handle_file_validate_tool_call
from proc.mcp_server.proc_rule import create_proc_rule_tools, handle_proc_rule_tool_call

# 导入 rules 模块（规则查询 + 数字员工管理）
from tools.rules import create_tools as create_rules_tools, handle_tool_call as handle_rules_call, get_rule_from_bus

# 导入对账模块
from recon.mcp_server.recon_tool import create_recon_tools, handle_recon_tool_call

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# 创建统一的 MCP Server
mcp_server = Server("financial-mcp-server")


@mcp_server.list_tools()
async def list_tools() -> list[types.Tool]:
    """列出所有工具（认证 + 对账 + 数据整理）"""
    try:
        auth_tools = create_auth_tools()
        logger.info(f"认证/规则管理工具数量: {len(auth_tools)}")
    except Exception as e:
        logger.error(f"加载认证工具失败: {str(e)}", exc_info=True)
        auth_tools = []

    try:
        guest_tools = _create_guest_tools()
        logger.info(f"游客认证工具数量: {len(guest_tools)}")
    except Exception as e:
        logger.error(f"加载游客工具失败: {str(e)}", exc_info=True)
        guest_tools = []

    try:
        recon_tools = create_recon_tools()
        logger.info(f"对账工具数量: {len(recon_tools)}")
    except Exception as e:
        logger.error(f"加载对账工具失败: {str(e)}", exc_info=True)
        recon_tools = []
    
    try:
        prep_tools = create_prep_tools()
        logger.info(f"数据整理工具数量: {len(prep_tools)}")
    except Exception as e:
        logger.error(f"加载数据整理工具失败: {str(e)}", exc_info=True)
        prep_tools = []
    
    try:
        file_validate_tools = create_file_validate_tools()
        logger.info(f"文件校验工具数量: {len(file_validate_tools)}")
    except Exception as e:
        logger.error(f"加载文件校验工具失败: {str(e)}", exc_info=True)
        file_validate_tools = []

    try:
        sync_rule_tools = create_proc_rule_tools()
        logger.info(f"数据同步规则工具数量: {len(sync_rule_tools)}")
    except Exception as e:
        logger.error(f"加载数据同步规则工具失败: {str(e)}", exc_info=True)
        sync_rule_tools = []

    try:
        rules_tools = create_rules_tools()
        logger.info(f"Rules 工具数量: {len(rules_tools)}")
    except Exception as e:
        logger.error(f"加载 Rules 工具失败: {str(e)}", exc_info=True)
        rules_tools = []

    try:
        recon_tools_2 = create_recon_tools()
        logger.info(f"对账工具数量: {len(recon_tools_2)}")
    except Exception as e:
        logger.error(f"加载对账工具失败: {str(e)}", exc_info=True)
        recon_tools_2 = []
    
    all_tools = auth_tools + guest_tools + recon_tools + prep_tools + file_validate_tools + sync_rule_tools + rules_tools + recon_tools_2
    logger.info(f"总工具数量: {len(all_tools)}")
    return all_tools


# 认证和规则管理工具名集合
_AUTH_TOOL_NAMES = {
    "auth_register", "auth_login", "auth_me",
    "list_reconciliation_rules", "get_reconciliation_rule",
    "save_reconciliation_rule", "update_reconciliation_rule",
    "delete_reconciliation_rule",
    "search_rules_by_mapping", "copy_reconciliation_rule", "batch_get_reconciliation_rules",
    # 管理员功能
    "admin_login", "create_company", "create_department",
    "list_companies", "list_departments", "get_admin_view",
    "list_companies_public", "list_departments_public",
    # 会话管理
    "create_conversation", "list_conversations", "get_conversation",
    "update_conversation", "delete_conversation", "save_message",
}

# 游客工具名集合
_GUEST_TOOL_NAMES = {
    "create_guest_token", "verify_guest_token", "list_recommended_rules"
}

# 文件校验工具名集合
_FILE_VALIDATE_TOOL_NAMES = {
    "validate_uploaded_files",
}

# 数据同步规则工具名集合
_SYNC_RULE_TOOL_NAMES = {
    "proc_rule_execute",
}

# Rules 工具名集合（规则查询 + 数字员工管理）
_RULES_TOOL_NAMES = {
    "get_rule_from_bus",
    "list_digital_employees",
    "list_rules_by_employee",
}

# 对账工具名集合
_RECON_TOOL_NAMES = {
    "recon_execute",
    "recon_list_rules",
}


@mcp_server.call_tool()
async def call_tool(name: str, arguments: dict) -> list[types.TextContent | types.ImageContent | types.EmbeddedResource]:
    """调用工具（自动路由到对应模块）"""
    try:
        import json

        # 1) 认证和规则管理工具
        if name in _AUTH_TOOL_NAMES:
            result = await handle_auth_tool_call(name, arguments)

        # 2) 游客认证工具
        elif name in _GUEST_TOOL_NAMES:
            if name == "create_guest_token":
                result = await _handle_create_guest_token(arguments)
            elif name == "verify_guest_token":
                result = await _handle_verify_guest_token(arguments)
            elif name == "list_recommended_rules":
                result = await _handle_list_recommended_rules(arguments)
            else:
                result = {"error": f"未知的游客工具: {name}"}

        # 3) 对账执行模块工具
        elif name.startswith("reconciliation_") or name in ["file_upload", "get_reconciliation", "analyze_files"]:
            result = await handle_recon_call(name, arguments)

        # 4) 数据整理模块
        elif name.startswith("data_preparation_"):
            result = await handle_prep_call(name, arguments)

        # 5) 文件校验模块
        elif name in _FILE_VALIDATE_TOOL_NAMES:
            result = await handle_file_validate_tool_call(name, arguments)

        # 6) 数据同步规则模块
        elif name in _SYNC_RULE_TOOL_NAMES:
            result = await handle_proc_rule_tool_call(name, arguments)

        # 7) Rules 模块（规则查询 + 数字员工管理）
        elif name in _RULES_TOOL_NAMES:
            result = await handle_rules_call(name, arguments)

        # 8) 对账模块
        elif name in _RECON_TOOL_NAMES:
            result = await handle_recon_tool_call(name, arguments)

        else:
            result = {"error": f"未知的工具: {name}"}
        
        result_str = json.dumps(result, ensure_ascii=False, indent=2)
        return [types.TextContent(type="text", text=result_str)]
    
    except Exception as e:
        error_msg = f"工具调用失败: {str(e)}"
        logger.error(error_msg, exc_info=True)
        return [types.TextContent(type="text", text=error_msg)]


# 创建 SSE Transport
sse_transport = SseServerTransport("/messages/")


# HTTP 端点处理函数
async def handle_sse(request):
    """处理 SSE 连接"""
    async with sse_transport.connect_sse(request.scope, request.receive, request._send) as streams:
        await mcp_server.run(
            streams[0],
            streams[1],
            mcp_server.create_initialization_options()
        )
    return Response()


async def health_check(request):
    """健康检查"""
    return JSONResponse({
        "status": "healthy",
        "service": "financial-mcp-server",
        "version": "1.0.0",
        "modules": ["reconciliation", "data_preparation", "proc"]
    })


async def download_file(request):
    """文件下载端点"""
    task_id = request.path_params.get("task_id")
    logger.info(f"下载文件请求: task_id={task_id}")

    # 验证 task_id 格式，防止路径遍历攻击
    if not validate_task_id(task_id):
        logger.warning(f"非法的 task_id: {task_id}")
        return JSONResponse(
            {"error": "无效的任务ID格式"},
            status_code=400
        )

    # 先查找报告文件获取输出文件路径
    from data_preparation.mcp_server.task_manager import TaskManager
    from data_preparation.mcp_server.config import (
        UPLOAD_DIR as PREP_UPLOAD_DIR,
        OUTPUT_DIR as PREP_OUTPUT_DIR,
        REPORT_DIR as PREP_REPORT_DIR,
        DATA_PREPARATION_SCHEMAS_FILE
    )

    # 尝试从任务管理器获取结果
    task_manager = TaskManager()
    try:
        result = await task_manager.get_task_result(task_id)
    except Exception as e:
        logger.error(f"获取任务结果失败: {str(e)}", exc_info=True)
        result = None

    # 从任务结果中获取输出文件路径或处理ID
    proc_id = None
    output_file = None
    
    if result and not result.get('error'):
        # 如果结果中有 output_file，直接使用
        if result.get('output_file'):
            output_file = sanitize_path(PREP_OUTPUT_DIR, Path(result['output_file']).name)
        # 如果结果中有 task_id（可能是处理ID），用于查找报告文件
        proc_id = result.get('task_id') or task_id
    
    # 如果找不到输出文件，尝试从报告文件中获取
    if (not output_file or not output_file.exists()) and proc_id:
        # 先尝试使用处理ID查找报告文件
        report_file = sanitize_path(PREP_REPORT_DIR, f"{proc_id}_report.json")
        if not report_file or not report_file.exists():
            # 如果处理ID找不到，尝试使用任务ID
            report_file = sanitize_path(PREP_REPORT_DIR, f"{task_id}_report.json")
        
        if report_file and report_file.exists():
            import json
            try:
                with open(report_file, 'r', encoding='utf-8') as f:
                    report_data = json.load(f)
                    # 尝试从 processing_steps 中获取输出文件路径
                    processing_steps = report_data.get('processing_steps', [])
                    if processing_steps:
                        last_step = processing_steps[-1]
                        output_path = last_step.get('details', {}).get('output_file')
                        if output_path:
                            output_file = sanitize_path(PREP_OUTPUT_DIR, Path(output_path).name)
                    # 如果还是找不到，尝试从根级别获取
                    if (not output_file or not output_file.exists()) and report_data.get('output_file'):
                        output_file = sanitize_path(PREP_OUTPUT_DIR, Path(report_data['output_file']).name)
            except (json.JSONDecodeError, IndexError, KeyError) as e:
                logger.error(f"解析报告文件失败: {str(e)}", exc_info=True)
                # 继续尝试其他方法，不直接返回错误

    # 如果仍然找不到，尝试在输出目录中查找包含任务ID或处理ID的文件
    if not output_file or not output_file.exists():
        try:
            output_files = list(PREP_OUTPUT_DIR.glob("*"))
            # 按修改时间排序，取最新的
            if output_files:
                output_files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
                # 如果任务ID或处理ID在文件名中，优先选择
                for f in output_files:
                    if task_id in f.name or (proc_id and proc_id in f.name):
                        output_file = f
                        break
                # 如果没找到匹配的，使用最新的文件
                if (not output_file or not output_file.exists()) and output_files:
                    output_file = output_files[0]
        except Exception as e:
            logger.error(f"查找输出文件失败: {str(e)}", exc_info=True)

    if output_file and output_file.exists():
        logger.info(f"找到输出文件: {output_file}")
        return FileResponse(
            str(output_file),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=output_file.name
        )

    # 记录详细的错误信息
    logger.warning(f"未找到文件: task_id={task_id}, proc_id={proc_id}, output_file={output_file}")
    return JSONResponse({
        "error": f"文件不存在: {task_id}",
        "message": "请确认任务已完成且输出文件已生成",
        "task_id": task_id
    }, status_code=404)


async def preview_file(request):
    """文件预览端点（返回文件基本信息）"""
    task_id = request.path_params.get("task_id")

    # 验证 task_id 格式，防止路径遍历攻击
    if not validate_task_id(task_id):
        logger.warning(f"非法的 task_id: {task_id}")
        return JSONResponse(
            {"error": "无效的任务ID格式"},
            status_code=400
        )

    # 先查找报告文件获取输出文件路径
    from data_preparation.mcp_server.task_manager import TaskManager
    from data_preparation.mcp_server.config import (
        OUTPUT_DIR as PREP_OUTPUT_DIR,
        REPORT_DIR as PREP_REPORT_DIR
    )

    # 尝试从任务管理器获取结果
    task_manager = TaskManager()
    try:
        result = await task_manager.get_task_result(task_id)
    except Exception as e:
        logger.error(f"获取任务结果失败: {str(e)}", exc_info=True)
        result = None

    # 从任务结果中获取输出文件路径或处理ID
    proc_id = None
    output_file = None
    
    if result and not result.get('error'):
        # 如果结果中有 output_file，直接使用
        if result.get('output_file'):
            output_file = sanitize_path(PREP_OUTPUT_DIR, Path(result['output_file']).name)
        # 如果结果中有 task_id（可能是处理ID），用于查找报告文件
        proc_id = result.get('task_id') or task_id

    # 如果找不到，尝试从报告文件中获取
    if (not output_file or not output_file.exists()) and proc_id:
        # 先尝试使用处理ID查找报告文件
        report_file = sanitize_path(PREP_REPORT_DIR, f"{proc_id}_report.json")
        if not report_file or not report_file.exists():
            # 如果处理ID找不到，尝试使用任务ID
            report_file = sanitize_path(PREP_REPORT_DIR, f"{task_id}_report.json")
        
        if report_file and report_file.exists():
            import json
            try:
                with open(report_file, 'r', encoding='utf-8') as f:
                    report_data = json.load(f)
                    # 尝试从 processing_steps 中获取输出文件路径
                    processing_steps = report_data.get('processing_steps', [])
                    if processing_steps:
                        last_step = processing_steps[-1]
                        output_path = last_step.get('details', {}).get('output_file')
                        if output_path:
                            output_file = sanitize_path(PREP_OUTPUT_DIR, Path(output_path).name)
                    # 如果还是找不到，尝试从根级别获取
                    if (not output_file or not output_file.exists()) and report_data.get('output_file'):
                        output_file = sanitize_path(PREP_OUTPUT_DIR, Path(report_data['output_file']).name)
            except (json.JSONDecodeError, IndexError, KeyError) as e:
                logger.error(f"解析报告文件失败: {str(e)}", exc_info=True)
                # 继续尝试其他方法，不直接返回错误

    # 如果仍然找不到，尝试在输出目录中查找包含任务ID或处理ID的文件
    if not output_file or not output_file.exists():
        try:
            output_files = list(PREP_OUTPUT_DIR.glob("*"))
            # 按修改时间排序，取最新的
            if output_files:
                output_files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
                # 如果任务ID或处理ID在文件名中，优先选择
                for f in output_files:
                    if task_id in f.name or (proc_id and proc_id in f.name):
                        output_file = f
                        break
                # 如果没找到匹配的，使用最新的文件
                if (not output_file or not output_file.exists()) and output_files:
                    output_file = output_files[0]
        except Exception as e:
            logger.error(f"查找输出文件失败: {str(e)}", exc_info=True)

    if not output_file or not output_file.exists():
        return JSONResponse({"error": f"文件不存在: {task_id}"}, status_code=404)

    import openpyxl
    try:
        wb = openpyxl.load_workbook(str(output_file), read_only=True, data_only=True)
        sheets = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheets.append({
                "name": sheet_name,
                "rows": ws.max_row,
                "columns": ws.max_column
            })

        return JSONResponse({
            "filename": output_file.name,
            "size": output_file.stat().st_size,
            "sheets": sheets,
            "download_url": f"/download/{task_id}"
        })
    except Exception as e:
        logger.error(f"预览文件失败: {str(e)}", exc_info=True)
        return JSONResponse({"error": f"预览失败: {str(e)}"}, status_code=500)


# ═══════════════════════════════════════════════════════════════════════════════
# 通用文件下载
# ═══════════════════════════════════════════════════════════════════════════════

# 模块输出目录映射
_MODULE_OUTPUT_DIRS = {
    "proc": None,  # 延迟加载
    "recon": None,  # 延迟加载
    "prep": None,  # 延迟加载
}


def _get_module_output_dir(module: str) -> Optional[Path]:
    """获取模块的输出目录（延迟加载）"""
    if module not in _MODULE_OUTPUT_DIRS:
        return None
    
    if _MODULE_OUTPUT_DIRS[module] is not None:
        return _MODULE_OUTPUT_DIRS[module]
    
    try:
        if module == "proc":
            from proc.config.config import OUTPUT_DIR as PROC_OUTPUT_DIR
            _MODULE_OUTPUT_DIRS[module] = Path(PROC_OUTPUT_DIR)
        elif module == "recon":
            from recon.mcp_server.recon_tool import RECON_OUTPUT_DIR
            _MODULE_OUTPUT_DIRS[module] = RECON_OUTPUT_DIR
        elif module == "prep":
            from data_preparation.mcp_server.config import OUTPUT_DIR as PREP_OUTPUT_DIR
            _MODULE_OUTPUT_DIRS[module] = Path(PREP_OUTPUT_DIR)
        return _MODULE_OUTPUT_DIRS[module]
    except ImportError as e:
        logger.error(f"无法导入模块 {module} 的输出目录: {e}")
        return None


async def download_output_file(request):
    """通用文件下载端点。

    路径格式: /output/{module}/{path:path}
    - module: 模块名称（proc/recon/prep）
    - path: 文件相对路径（可包含子目录）

    示例:
    - /output/proc/{rule_code}/{filename} → proc/output/{rule_code}/{filename}
    - /output/recon/{filename} → recon/output/{filename}
    - /output/prep/{filename} → prep/output/{filename}
    """
    module = request.path_params.get("module", "")
    file_path = request.path_params.get("path", "")

    # 参数校验
    if not module or not file_path:
        return JSONResponse({"error": "缺少必要参数"}, status_code=400)

    # 模块白名单校验
    if module not in _MODULE_OUTPUT_DIRS:
        return JSONResponse({"error": f"不支持的模块: {module}"}, status_code=400)

    # 安全校验：禁止路径遍历
    if ".." in file_path:
        return JSONResponse({"error": "无效的文件路径"}, status_code=400)

    # 获取模块输出目录
    output_dir = _get_module_output_dir(module)
    if output_dir is None:
        return JSONResponse({"error": f"模块 {module} 配置错误"}, status_code=500)

    # 构建完整文件路径
    full_path = output_dir / file_path
    logger.info(f"[download] 请求下载: module={module} path={file_path} full_path={full_path}")

    # 安全检查：确保路径在输出目录内
    try:
        full_path.resolve().relative_to(output_dir.resolve())
    except ValueError:
        logger.warning(f"[download] 路径遍历攻击尝试: {file_path}")
        return JSONResponse({"error": "无效的文件路径"}, status_code=400)

    if not full_path.exists() or not full_path.is_file():
        logger.warning(f"[download] 文件不存在: {full_path}")
        return JSONResponse({"error": f"文件不存在: {file_path}"}, status_code=404)

    # 对中文文件名使用 RFC 5987 编码
    from urllib.parse import quote
    filename = full_path.name
    encoded_filename = quote(filename, safe='')

    return FileResponse(
        str(full_path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
        },
    )


async def get_report(request):
    """获取详细报告"""
    task_id = request.path_params.get("task_id")

    # 验证 task_id 格式，防止路径遍历攻击
    if not validate_task_id(task_id):
        logger.warning(f"非法的 task_id: {task_id}")
        return JSONResponse(
            {"error": "无效的任务ID格式"},
            status_code=400
        )

    from data_preparation.mcp_server.task_manager import TaskManager
    from data_preparation.mcp_server.config import (
        REPORT_DIR as PREP_REPORT_DIR,
        OUTPUT_DIR as PREP_OUTPUT_DIR
    )

    logger.info(f"获取报告请求: task_id={task_id}")

    # 尝试从任务管理器获取结果，以获取处理ID
    task_manager = TaskManager()
    proc_id = None
    try:
        result = await task_manager.get_task_result(task_id)
        if result and not result.get('error'):
            # 如果结果中有 task_id（可能是处理ID），用于查找报告文件
            proc_id = result.get('task_id') or task_id
            logger.info(f"从任务管理器获取处理ID: {proc_id}")
    except Exception as e:
        logger.error(f"获取任务结果失败: {str(e)}", exc_info=True)
        result = None

    # 查找报告文件的多种策略
    report_file = None
    
    # 策略1: 使用处理ID查找
    if proc_id and proc_id != task_id:
        report_file = sanitize_path(PREP_REPORT_DIR, f"{proc_id}_report.json")
        if report_file and report_file.exists():
            logger.info(f"使用处理ID找到报告文件: {report_file}")
    
    # 策略2: 使用任务ID查找
    if not report_file or not report_file.exists():
        report_file = sanitize_path(PREP_REPORT_DIR, f"{task_id}_report.json")
        if report_file and report_file.exists():
            logger.info(f"使用任务ID找到报告文件: {report_file}")
    
    # 策略3: 从输出文件推断处理ID（如果输出文件名包含日期时间戳）
    if not report_file or not report_file.exists():
        try:
            # 查找输出目录中可能相关的文件
            output_files = list(PREP_OUTPUT_DIR.glob("*"))
            if output_files:
                # 按修改时间排序，取最新的
                output_files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
                # 尝试从最新的输出文件名中提取日期时间信息
                for f in output_files:
                    import re
                    # 文件名格式可能是: 货币资金审计底稿_20260116_163437.xlsx
                    date_match = re.search(r'(\d{8})_(\d{6})', f.name)
                    if date_match:
                        date_str = date_match.group(1)
                        time_str = date_match.group(2)
                        # 构建可能的处理ID前缀: proc_20260116_163437_
                        proc_prefix = f"proc_{date_str}_{time_str}"
                        logger.info(f"从文件名推断处理ID前缀: {proc_prefix}")
                        # 查找所有匹配的报告文件
                        matching_reports = list(PREP_REPORT_DIR.glob(f"{proc_prefix}*_report.json"))
                        if matching_reports:
                            # 按修改时间排序，取最新的
                            matching_reports.sort(key=lambda p: p.stat().st_mtime, reverse=True)
                            report_file = matching_reports[0]
                            logger.info(f"从输出文件推断找到报告文件: {report_file}")
                            break
        except Exception as e:
            logger.error(f"从输出文件推断处理ID失败: {str(e)}", exc_info=True)
    
    # 策略4: 列出所有报告文件，按修改时间排序，返回最新的
    if not report_file or not report_file.exists():
        try:
            report_files = list(PREP_REPORT_DIR.glob("*_report.json"))
            if report_files:
                # 按修改时间排序，取最新的
                report_files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
                # 如果任务ID在文件名中，优先选择
                for f in report_files:
                    if task_id in f.name:
                        report_file = f
                        logger.info(f"从报告文件列表中找到匹配的报告: {report_file}")
                        break
                # 如果没找到匹配的，使用最新的报告文件
                if (not report_file or not report_file.exists()) and report_files:
                    report_file = report_files[0]
                    logger.info(f"使用最新的报告文件: {report_file}")
        except Exception as e:
            logger.error(f"列出报告文件失败: {str(e)}", exc_info=True)

    if not report_file or not report_file.exists():
        # 列出所有可用的报告文件供参考
        try:
            available_reports = [f.name for f in PREP_REPORT_DIR.glob("*_report.json")]
            available_reports.sort(reverse=True)  # 按名称倒序（最新的在前）
            logger.warning(f"未找到报告文件，可用报告: {available_reports[:5]}")
            return JSONResponse({
                "error": f"报告不存在: {task_id}",
                "message": "请确认任务已完成且报告文件已生成",
                "available_reports": available_reports[:10] if available_reports else [],
                "suggestion": "如果任务已完成，可以尝试使用处理ID（proc_xxx）访问报告"
            }, status_code=404)
        except Exception as e:
            logger.error(f"列出可用报告失败: {str(e)}", exc_info=True)
            return JSONResponse({
                "error": f"报告不存在: {task_id}",
                "message": "请确认任务已完成且报告文件已生成"
            }, status_code=404)

    try:
        import json
        with open(report_file, 'r', encoding='utf-8') as f:
            report_data = json.load(f)
        return JSONResponse(report_data)
    except json.JSONDecodeError as e:
        logger.error(f"报告文件格式错误: {str(e)}", exc_info=True)
        return JSONResponse({"error": "报告文件格式错误"}, status_code=500)
    except Exception as e:
        logger.error(f"读取报告失败: {str(e)}", exc_info=True)
        return JSONResponse({"error": f"读取报告失败: {str(e)}"}, status_code=500)


# 路由配置
routes = [
    Route("/sse", endpoint=handle_sse, methods=["GET", "POST"]),
    Route("/mcp", endpoint=handle_sse, methods=["GET", "POST"]),
    Mount("/messages/", app=sse_transport.handle_post_message),
    Route("/health", endpoint=health_check),
    Route("/download/{task_id}", endpoint=download_file),
    Route("/output/{module}/{path:path}", endpoint=download_output_file),
    Route("/preview/{task_id}", endpoint=preview_file),
    Route("/report/{task_id}", endpoint=get_report),
]

app = Starlette(routes=routes)


async def main():
    """启动服务器"""
    host = DEFAULT_HOST
    port = DEFAULT_PORT
    
    # 动态获取工具列表用于显示
    try:
        tools = await list_tools()
        recon_tools = [t for t in tools if t.name.startswith("reconciliation_") or t.name == "file_upload" or t.name == "get_reconciliation"]
        prep_tools = [t for t in tools if t.name.startswith("data_preparation_")]
        proc_tools = [t for t in tools if t.name in _FILE_VALIDATE_TOOL_NAMES or t.name in _SYNC_RULE_TOOL_NAMES]
    except Exception as e:
        logger.warning(f"获取工具列表失败: {e}")
        recon_tools = []
        prep_tools = []
        proc_tools = []
    
    recon_tools_text = "\n".join([f"  {i+1}. {t.name:<30} - {t.description}" for i, t in enumerate(recon_tools)])
    prep_tools_text = "\n".join([f"  {len(recon_tools)+i+1}. {t.name:<30} - {t.description}" for i, t in enumerate(prep_tools)])
    proc_tools_text = "\n".join([f"  {len(recon_tools)+len(prep_tools)+i+1}. {t.name:<30} - {t.description}" for i, t in enumerate(proc_tools)])
    
    print(f"""
╔══════════════════════════════════════════════════════════════════╗
║          Financial Agent MCP Server 启动中...                    ║
╚══════════════════════════════════════════════════════════════════╝

🌐 服务端点:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  • SSE 端点:        http://{host}:{port}/sse
  • 消息端点:        http://{host}:{port}/messages/
  • 健康检查:        http://{host}:{port}/health
  • 文件下载:        http://{host}:{port}/download/{{task_id}}
  • 输出下载:        http://{host}:{port}/output/{{module}}/{{path}}
  • 文件预览:        http://{host}:{port}/preview/{{task_id}}
  • 详细报告:        http://{host}:{port}/report/{{task_id}}

🛠️  可用工具（对账模块，{len(recon_tools)}个）:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{recon_tools_text}

🛠️  可用工具（数据整理模块，{len(prep_tools)}个）:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{prep_tools_text}

🛠️  可用工具（Proc 模块，{len(proc_tools)}个）:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{proc_tools_text}

📖 使用说明:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  在 Dify 中配置:
    MCP 服务器地址: http://localhost:{port}/sse
    或使用 Docker:   http://host.docker.internal:{port}/sse

服务器正在运行...
""")
    
    config = uvicorn.Config(
        app,
        host=host,
        port=port,
        log_level="info"
    )
    server = uvicorn.Server(config)
    await server.serve()


if __name__ == "__main__":
    asyncio.run(main())
