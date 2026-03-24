"""
核对 MCP 工具模块

根据规则定义，执行源文件与目标文件的数据比对与差异分析。
支持两种规则类型：
1. 对账规则：包含 rules 字段
主要功能：
1. 加载规则（从 rule_detail 表，根据传入的 rule_code）
2. 判断规则类型（对账 vs 普通对账）
3. 执行数据核对：关键列匹配、数值比对、聚合比对
4. 输出差异分析结果：差异记录、源文件独有、目标文件独有
"""
from __future__ import annotations

import json
import logging
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Optional
from urllib.parse import quote

import pandas as pd
from mcp import Tool
from auth.jwt_utils import get_user_from_token
from security_utils import resolve_upload_file_path, write_output_metadata

# 导入数据过滤模块
from tools.data_filter import filter_dataframe_by_rule_config, get_filter_statistics
from tools.rule_schema import load_and_validate_rule

logger = logging.getLogger(__name__)

# ════════════════════════════════════════════════════════════════════════════
# 常量定义
# ════════════════════════════════════════════════════════════════════════════

# 对账报告输出目录（相对于 finance-mcp 目录）
RECON_OUTPUT_DIR = Path(__file__).parent.parent / "output"


# ════════════════════════════════════════════════════════════════════════════
# MCP 工具定义
# ════════════════════════════════════════════════════════════════════════════

def create_recon_tools() -> list[Tool]:
    """创建核对 MCP 工具列表"""
    return [
        Tool(
            name="recon_execute",
            description=(
                "执行对账：根据规则对源文件与目标文件进行数据比对，"
                "支持对账规则（含 rules）和普通对账规则。"
                "输出差异记录、源文件独有记录、目标文件独有记录等。"
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "validated_inputs": {
                        "type": "array",
                        "description": (
                            "统一的输入列表，每个元素通过 input_type 指定来源。"
                            "input_type=file 时需提供 file_path；input_type=dataset 时需提供 dataset_ref。"
                        ),
                        "items": {
                            "type": "object",
                            "properties": {
                                "table_name": {"type": "string"},
                                "input_type": {"type": "string", "enum": ["file", "dataset"]},
                                "file_path": {"type": "string"},
                                "dataset_ref": {"type": "object"},
                            }
                        },
                    },
                    "validated_files": {
                        "type": "array",
                        "description": "兼容旧参数：文件校验结果列表，每个元素包含 file_path 和 table_name",
                        "items": {
                            "type": "object",
                            "properties": {
                                "file_path": {"type": "string"},
                                "table_name": {"type": "string"}
                            }
                        }
                    },
                    "rule_code": {
                        "type": "string",
                        "description": "规则编码（rule_code），用于从 rule_detail 表获取规则定义"
                    },
                    "rule_id": {
                        "type": "string",
                        "description": "要执行的核对规则 ID（如 AUDIT_RECONC_001），仅在对账规则中使用，不指定则执行所有匹配的规则"
                    },
                    "auth_token": {
                        "type": "string",
                        "description": "JWT token，用于校验当前用户是否有权使用该规则"
                    }
                },
                "required": ["rule_code", "auth_token"]
            }
        ),
    ]


async def handle_recon_tool_call(name: str, arguments: dict) -> dict:
    """处理核对工具调用"""
    try:
        if name == "recon_execute":
            return await _handle_recon_execute(arguments)
        return {"success": False, "error": f"未知的工具: {name}"}
    except Exception as e:
        logger.error(f"核对工具调用失败 [{name}]: {e}", exc_info=True)
        return {"success": False, "error": f"工具调用失败: {str(e)}"}


# ════════════════════════════════════════════════════════════════════════════
# 规则加载（复用 tools.rules 中的公共方法）
# ════════════════════════════════════════════════════════════════════════════

def _get_rule(rule_code: str, user_id: str) -> Optional[dict]:
    """
    从 rule_detail 表加载指定 rule_code 的规则配置
    
    复用 tools.rules 中的 get_rule 函数
    
    Args:
        rule_code: 规则编码
        
    Returns:
        规则字典，包含 id, user_id, rule_code, rule, rule_type, remark 等字段；未找到返回 None
    """
    try:
        from tools.rules import get_rule
        return get_rule(rule_code, user_id=user_id)
    except ImportError:
        logger.error(f"[recon] 无法导入 tools.rules.get_rule")
        return None

def find_recon_rule_by_id(rules_config: dict, rule_id: str) -> Optional[dict]:
    """根据 rule_id 查找规则"""
    if rules_config.get("rule_id") == rule_id:
        rules = rules_config.get("rules", [])
        return rules[0] if rules else None
    for rule in rules_config.get("rules", []):
        if rule.get("rule_id") == rule_id:
            return rule
    return None


def _normalize_validated_inputs(
    *,
    validated_inputs: list[dict[str, Any]] | None,
    validated_files: list[dict[str, Any]] | None,
) -> tuple[list[dict[str, Any]], str | None]:
    """统一 recon 输入，兼容 validated_inputs / validated_files。"""
    normalized: list[dict[str, Any]] = []

    for item in validated_inputs or []:
        if not isinstance(item, dict):
            continue
        table_name = str(item.get("table_name") or "").strip()
        input_type = str(item.get("input_type") or "").strip().lower()
        if not table_name or input_type not in {"file", "dataset"}:
            continue
        if input_type == "file":
            file_path = str(item.get("file_path") or "").strip()
            if not file_path:
                continue
            normalized.append({
                "table_name": table_name,
                "input_type": "file",
                "file_path": file_path,
            })
            continue
        dataset_ref = item.get("dataset_ref")
        if not isinstance(dataset_ref, dict):
            dataset_ref = {}
        normalized.append({
            "table_name": table_name,
            "input_type": "dataset",
            "dataset_ref": dataset_ref,
        })

    if not normalized:
        for item in validated_files or []:
            if not isinstance(item, dict):
                continue
            table_name = str(item.get("table_name") or "").strip()
            file_path = str(item.get("file_path") or "").strip()
            if not table_name or not file_path:
                continue
            normalized.append({
                "table_name": table_name,
                "input_type": "file",
                "file_path": file_path,
            })

    if not normalized:
        return [], "validated_inputs / validated_files 不能为空"
    return normalized, None


def _find_input_by_identification(
    identification: dict[str, Any],
    table_input_map: dict[str, dict[str, Any]],
    rule_id: str,
    file_type: str,
) -> Optional[dict[str, Any]]:
    """根据表名识别输入对象。"""
    match_by = identification.get("match_by", "table_name")
    match_value = identification.get("match_value", "")
    match_strategy = identification.get("match_strategy", "exact")

    if not match_value:
        logger.warning(f"[recon] [{rule_id}] {file_type} 输入 match_value 未配置")
        return None

    if match_by == "table_name":
        if match_strategy == "exact":
            return table_input_map.get(match_value)
        if match_strategy == "contains":
            for table_name, input_item in table_input_map.items():
                if match_value in table_name:
                    return input_item
        if match_strategy == "startswith":
            for table_name, input_item in table_input_map.items():
                if table_name.startswith(match_value):
                    return input_item

    logger.warning(f"[recon] [{rule_id}] 未找到匹配的 {file_type} 输入: {match_value}")
    return None


def _dataset_display_name(dataset_ref: dict[str, Any], table_name: str) -> str:
    source_key = str(dataset_ref.get("source_key") or "").strip()
    if source_key:
        return source_key
    return table_name


def _read_dataset_as_df(dataset_ref: dict[str, Any], table_name: str) -> pd.DataFrame:
    """读取 dataset 输入。

    当前仅支持测试/占位数据：
    - dataset_ref.rows: list[dict]
    - dataset_ref.data: list[dict]
    真实 db/api 抓取逻辑后续再接入。
    """
    rows = dataset_ref.get("rows")
    if isinstance(rows, list):
        return pd.DataFrame(rows)
    rows = dataset_ref.get("data")
    if isinstance(rows, list):
        return pd.DataFrame(rows)
    source_type = str(dataset_ref.get("source_type") or "").strip() or "dataset"
    source_key = str(dataset_ref.get("source_key") or "").strip() or table_name
    raise NotImplementedError(
        f"dataset 输入暂未接入真实 {source_type} 数据源读取，请为 source_key={source_key} 提供 rows/data 占位数据。"
    )


def _resolve_input_to_df(input_item: dict[str, Any], rule_id: str, table_name: str) -> tuple[pd.DataFrame, str]:
    """将 file / dataset 输入统一解析为 DataFrame。"""
    input_type = str(input_item.get("input_type") or "").strip().lower()
    if input_type == "file":
        file_path = str(input_item.get("file_path") or "").strip()
        return _read_file_as_df(file_path), file_path
    if input_type == "dataset":
        dataset_ref = input_item.get("dataset_ref")
        if not isinstance(dataset_ref, dict):
            dataset_ref = {}
        return _read_dataset_as_df(dataset_ref, table_name), _dataset_display_name(dataset_ref, table_name)
    raise ValueError(f"[{rule_id}] 不支持的输入类型: {input_type}")


# ════════════════════════════════════════════════════════════════════════════
# 工具处理函数
# ════════════════════════════════════════════════════════════════════════════

async def _handle_recon_execute(arguments: dict) -> dict:
    """执行对账（支持对账和普通对账）"""
    validated_inputs_raw = arguments.get("validated_inputs", [])
    validated_files = arguments.get("validated_files", [])
    rule_code = arguments.get("rule_code", "")
    rule_id = arguments.get("rule_id")
    auth_token = arguments.get("auth_token", "").strip()

    validated_inputs, input_error = _normalize_validated_inputs(
        validated_inputs=validated_inputs_raw,
        validated_files=validated_files,
    )
    if input_error:
        return {"success": False, "error": input_error}
    if not rule_code:
        return {"success": False, "error": "rule_code 不能为空"}
    if not auth_token:
        return {"success": False, "error": "未提供认证 token，请先登录"}

    user = get_user_from_token(auth_token)
    if not user:
        return {"success": False, "error": "token 无效或已过期，请重新登录"}

    user_id = str(user.get("user_id") or user.get("id") or "")
    if not user_id:
        return {"success": False, "error": "token 中缺少用户标识"}
    
    # 使用常量定义的输出目录
    output_dir = str(RECON_OUTPUT_DIR)
    
    # 确保输出目录存在
    RECON_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    # 加载规则并做结构校验
    validation_result = load_and_validate_rule(rule_code, expected_kind="recon", user_id=user_id)
    if not validation_result.get("success"):
        validation_result["status"] = "invalid_request"
        validation_result["success"] = False
        return validation_result

    rule_content = validation_result.get("rule", {})
    
    # 判断规则类型
    is_recon = isinstance(rule_content, dict) and rule_content.get("rules") is not None
    
    logger.info(f"[recon] 规则类型: {'对账' if is_recon else '普通对账'}, rule_code={rule_code}")
    
    # 构建 table_name -> input 映射
    table_input_map: dict[str, dict[str, Any]] = {}
    for item in validated_inputs:
        table_name = item["table_name"]
        if table_name in table_input_map:
            logger.warning(f"[recon] table_name={table_name} 重复输入，后者覆盖前者")
        table_input_map[table_name] = item

    input_labels = [
        f"{name}({table_input_map[name].get('input_type')})"
        for name in table_input_map.keys()
    ]
    logger.info(f"[recon] 输入映射: {input_labels}")
    
    if is_recon:
        # 对账逻辑
        rules_config = rule_content
        
        # 确定要执行的规则
        rules = rules_config.get("rules", [])
        if rule_id:
            target_rule = find_recon_rule_by_id(rules_config, rule_id)
            if target_rule is None:
                return {"success": False, "error": f"未找到 rule_id='{rule_id}' 的规则"}
            rules = [target_rule]
        
        # 执行核对
        results = []
        for rule in rules:
            if not rule.get("enabled", True):
                results.append(
                    {
                        "success": False,
                        "status": "skipped",
                        "rule_id": rules_config.get("rule_id", "UNKNOWN"),
                        "rule_name": rules_config.get("rule_name", "未命名规则"),
                        "skip_reason": "disabled",
                        "message": "规则已禁用",
                    }
                )
                continue

            results.append(
                execute_single_recon(
                    rule,
                    table_input_map,
                    output_dir,
                    auth_token,
                    user_id,
                    rule_code,
                    rules_config,
                )
            )

        summary = _build_recon_summary(results)
        status = _derive_recon_status(summary)

        return {
            "success": status in {"success", "partial_success"},
            "status": status,
            "rule_code": rule_code,
            "rule_type": "recon",
            "total_rules": len(results),
            "success_count": summary["succeeded_rules"],
            "summary": summary,
            "results": results
        }
    else:
        # 普通对账暂未支持 dataset 输入，保持旧行为并返回明确错误
        has_dataset_input = any(item.get("input_type") == "dataset" for item in validated_inputs)
        if has_dataset_input:
            return {
                "success": False,
                "status": "invalid_request",
                "error": "普通对账暂不支持 dataset 输入，请先使用 file 输入。",
            }
        legacy_validated_files = [
            {"table_name": item["table_name"], "file_path": item.get("file_path", "")}
            for item in validated_inputs
            if item.get("input_type") == "file" and item.get("file_path")
        ]
        # 普通对账逻辑
        return await _execute_normal_reconciliation(
            rule_content=rule_content,
            rule_code=rule_code,
            table_file_map={item["table_name"]: item.get("file_path", "") for item in legacy_validated_files},
            output_dir=output_dir,
            validated_files=legacy_validated_files
        )


# ════════════════════════════════════════════════════════════════════════════
# 核对执行逻辑
# ════════════════════════════════════════════════════════════════════════════

def execute_single_recon(
    rule: dict,
    table_input_map: dict[str, dict[str, Any]],
    output_dir: str,
    auth_token: str,
    user_id: str,
    rule_code: str,
    rule_meta: dict[str, Any] | None = None,
) -> dict:
    """
    执行单个对账规则
    
    Args:
        rule: 规则配置
        table_input_map: table_name -> input 映射
        output_dir: 输出目录
    
    Returns:
        核对结果
    """
    meta = rule_meta or {}
    rule_id = meta.get("rule_id") or rule.get("rule_id", "UNKNOWN")
    rule_name = meta.get("rule_name") or rule.get("rule_name", "未命名规则")
    
    logger.info(f"[recon] [{rule_id}] 开始执行: {rule_name}")
    
    # 1. 识别源文件和目标文件
    source_file_config = rule.get("source_file", {})
    target_file_config = rule.get("target_file", {})
    
    source_input = _find_input_by_identification(
        source_file_config.get("identification", {}),
        table_input_map,
        rule_id,
        "source"
    )
    target_input = _find_input_by_identification(
        target_file_config.get("identification", {}),
        table_input_map,
        rule_id,
        "target"
    )
    
    if source_input is None:
        return {
            "success": False,
            "status": "skipped",
            "rule_id": rule_id,
            "rule_name": rule_name,
            "skip_reason": "missing_source_file",
            "message": "未找到源文件",
        }
    
    if target_input is None:
        return {
            "success": False,
            "status": "skipped",
            "rule_id": rule_id,
            "rule_name": rule_name,
            "skip_reason": "missing_target_file",
            "message": "未找到目标文件",
        }
    
    logger.info(f"[recon] [{rule_id}] 源输入: {source_input}")
    logger.info(f"[recon] [{rule_id}] 目标输入: {target_input}")
    
    # 2. 读取输入
    try:
        df_source, source_path = _resolve_input_to_df(
            source_input,
            rule_id,
            str(source_file_config.get("table_name") or "源文件"),
        )
        df_target, target_path = _resolve_input_to_df(
            target_input,
            rule_id,
            str(target_file_config.get("table_name") or "目标文件"),
        )
    except Exception as e:
        return {
            "success": False,
            "status": "failed",
            "rule_id": rule_id,
            "rule_name": rule_name,
            "error_code": "read_file_failed",
            "error": f"读取文件失败: {e}"
        }
    
    logger.info(f"[recon] [{rule_id}] 源文件 {len(df_source)} 行，目标文件 {len(df_target)} 行")
    
    # 3. 应用数据过滤
    source_filter_stats = None
    target_filter_stats = None
    
    # 过滤源文件
    df_source_original = df_source.copy()
    logger.info(f"[recon] [{rule_id}] 源文件过滤前: {len(df_source)} 行, filter配置: {source_file_config.get('filter')}")
    df_source = filter_dataframe_by_rule_config(df_source, source_file_config)
    logger.info(f"[recon] [{rule_id}] 源文件过滤后: {len(df_source)} 行")
    if len(df_source) != len(df_source_original):
        source_filter_stats = get_filter_statistics(
            df_source_original, 
            df_source, 
            source_file_config.get("table_name", "源文件")
        )
    else:
        logger.info(f"[recon] [{rule_id}] 源文件无过滤或过滤前后行数相同")
    
    # 过滤目标文件
    df_target_original = df_target.copy()
    logger.info(f"[recon] [{rule_id}] 目标文件过滤前: {len(df_target)} 行, filter配置: {target_file_config.get('filter')}")
    df_target = filter_dataframe_by_rule_config(df_target, target_file_config)
    logger.info(f"[recon] [{rule_id}] 目标文件过滤后: {len(df_target)} 行")
    if len(df_target) != len(df_target_original):
        target_filter_stats = get_filter_statistics(
            df_target_original, 
            df_target, 
            target_file_config.get("table_name", "目标文件")
        )
    else:
        logger.info(f"[recon] [{rule_id}] 目标文件无过滤或过滤前后行数相同")
    
    logger.info(f"[recon] [{rule_id}] 过滤后：源文件 {len(df_source)} 行，目标文件 {len(df_target)} 行")
    
    # 4. 应用列映射
    df_source = _apply_column_mapping(df_source, source_file_config.get("column_mapping", {}))
    df_target = _apply_column_mapping(df_target, target_file_config.get("column_mapping", {}))
    
    # 4. 获取核对配置
    recon_config = rule.get("recon") or rule.get("reconciliation_config", {})
    key_columns_config = recon_config.get("key_columns", {})
    key_mappings = _get_key_mappings(key_columns_config)
    compare_columns_config = recon_config.get("compare_columns", {}).get("columns", [])
    aggregation_config = recon_config.get("aggregation", {})
    
    # 5. 执行聚合（如果启用）
    if aggregation_config.get("enabled", False):
        df_source = _apply_aggregation(df_source, aggregation_config, rule_id, "source")
        df_target = _apply_aggregation(df_target, aggregation_config, rule_id, "target")
    
    # 6. 执行核对比较
    diff_result = _execute_comparison(
        df_source=df_source,
        df_target=df_target,
        key_mappings=key_mappings,
        compare_columns_config=compare_columns_config,
        rule_id=rule_id,
        key_columns_config=key_columns_config
    )
    
    # 7. 输出结果
    output_config = rule.get("output", {})
    output_path = _write_recon_result(
        diff_result=diff_result,
        output_dir=output_dir,
        output_config=output_config,
        rule_id=rule_id,
        rule_name=rule_name,
        rule=rule,
        key_columns_config=key_columns_config,
        compare_columns_config=compare_columns_config
    )
    
    # 8. 生成下载链接
    download_url = None
    if output_path:
        file_name = Path(output_path).name
        # MCP_PUBLIC_BASE_URL 在 unified_mcp_server.py 中定义
        try:
            import unified_mcp_server
            base_url = unified_mcp_server.MCP_PUBLIC_BASE_URL.rstrip("/")
        except (ImportError, AttributeError):
            base_url = os.getenv("MCP_PUBLIC_BASE_URL", "http://localhost:3335").rstrip("/")
        write_output_metadata(
            output_path,
            {
                "owner_user_id": user_id,
                "module": "recon",
                "rule_code": rule_code,
                "rule_id": rule_id,
            },
        )
        download_url = f"{base_url}/output/recon/{file_name}?auth_token={quote(auth_token, safe='')}"
    
    # 9. 构建过滤提示信息
    filter_messages = []
    if source_filter_stats:
        filter_messages.append(
            f"源文件【{source_filter_stats['file_name']}】"
            f"原记录 {source_filter_stats['original_count']} 条，"
            f"过滤后参与对账 {source_filter_stats['filtered_count']} 条"
        )
    if target_filter_stats:
        filter_messages.append(
            f"目标文件【{target_filter_stats['file_name']}】"
            f"原记录 {target_filter_stats['original_count']} 条，"
            f"过滤后参与对账 {target_filter_stats['filtered_count']} 条"
        )
    
    # 构建完整消息
    message_parts = []
    if filter_messages:
        message_parts.append("数据过滤：" + "；".join(filter_messages))
    message_parts.append(
        f"核对完成：差异 {len(diff_result.get('matched_with_diff', []))} 条，"
        f"源独有 {len(diff_result.get('source_only', []))} 条，"
        f"目标独有 {len(diff_result.get('target_only', []))} 条"
    )
    
    result = {
        "success": True,
        "status": "succeeded",
        "rule_id": rule_id,
        "rule_name": rule_name,
        "source_file": source_path,
        "target_file": target_path,
        "source_rows": len(df_source),
        "target_rows": len(df_target),
        "matched_with_diff": len(diff_result.get("matched_with_diff", [])),
        "source_only": len(diff_result.get("source_only", [])),
        "target_only": len(diff_result.get("target_only", [])),
        "matched_exact": len(diff_result.get("matched_exact", [])),
        "output_file": output_path,
        "download_url": download_url,
        "message": "；".join(message_parts)
    }
    
    # 添加过滤统计详情（可选）
    if source_filter_stats:
        result["source_filter_stats"] = {
            "original_count": source_filter_stats["original_count"],
            "filtered_count": source_filter_stats["filtered_count"],
            "removed_count": source_filter_stats["removed_count"],
            "filter_rate": source_filter_stats["filter_rate"]
        }
    if target_filter_stats:
        result["target_filter_stats"] = {
            "original_count": target_filter_stats["original_count"],
            "filtered_count": target_filter_stats["filtered_count"],
            "removed_count": target_filter_stats["removed_count"],
            "filter_rate": target_filter_stats["filter_rate"]
        }
    
    return result


def _build_recon_summary(results: list[dict[str, Any]]) -> dict[str, int]:
    """统计对账执行结果。"""
    summary = {
        "total_rules": len(results),
        "enabled_rules": 0,
        "succeeded_rules": 0,
        "skipped_rules": 0,
        "failed_rules": 0,
    }
    for item in results:
        status = item.get("status")
        if status != "skipped" or item.get("skip_reason") != "disabled":
            summary["enabled_rules"] += 1
        if status == "succeeded":
            summary["succeeded_rules"] += 1
        elif status == "skipped":
            summary["skipped_rules"] += 1
        elif status == "failed":
            summary["failed_rules"] += 1
    return summary


def _derive_recon_status(summary: dict[str, int]) -> str:
    """根据统计结果推导顶层执行状态。"""
    succeeded = summary.get("succeeded_rules", 0)
    skipped = summary.get("skipped_rules", 0)
    failed = summary.get("failed_rules", 0)

    if succeeded > 0 and skipped == 0 and failed == 0:
        return "success"
    if succeeded > 0:
        return "partial_success"
    if failed > 0:
        return "failed"
    return "skipped"


def _find_file_by_identification(
    identification: dict,
    table_file_map: dict[str, str],
    rule_id: str,
    file_type: str
) -> Optional[str]:
    """根据识别规则查找文件"""
    match_by = identification.get("match_by", "table_name")
    match_value = identification.get("match_value", "")
    match_strategy = identification.get("match_strategy", "exact")
    
    if not match_value:
        logger.warning(f"[recon] [{rule_id}] {file_type} 文件 match_value 未配置")
        return None
    
    if match_by == "table_name":
        if match_strategy == "exact":
            return table_file_map.get(match_value)
        elif match_strategy == "contains":
            for table_name, file_path in table_file_map.items():
                if match_value in table_name:
                    return file_path
        elif match_strategy == "startswith":
            for table_name, file_path in table_file_map.items():
                if table_name.startswith(match_value):
                    return file_path
    
    logger.warning(f"[recon] [{rule_id}] 未找到匹配的 {file_type} 文件: {match_value}")
    return None


def _apply_column_mapping(df: pd.DataFrame, column_mapping: dict) -> pd.DataFrame:
    """应用列名映射"""
    mappings = column_mapping.get("mappings", {})
    if not mappings:
        return df
    
    # 构建重命名字典（source_col -> target_col）
    rename_dict = {}
    for source_col, target_col in mappings.items():
        if source_col in df.columns and source_col != target_col:
            rename_dict[source_col] = target_col
    
    if rename_dict:
        df = df.rename(columns=rename_dict)
    
    return df


def _get_key_mappings(key_columns_config: dict | None) -> list[dict[str, str]]:
    """解析关键列映射，兼容单字段与多字段配置。"""
    config = key_columns_config or {}
    mappings = config.get("mappings") or []
    normalized: list[dict[str, str]] = []

    for item in mappings:
        source_field = str(item.get("source_field") or "").strip()
        target_field = str(item.get("target_field") or "").strip()
        if source_field and target_field:
            normalized.append({"source_field": source_field, "target_field": target_field})

    if normalized:
        return normalized

    source_field = str(config.get("source_field") or "").strip()
    target_field = str(config.get("target_field") or "").strip()
    if source_field and target_field:
        return [{"source_field": source_field, "target_field": target_field}]
    return []


def _resolve_group_by_columns(aggregation_config: dict, file_type: str) -> list[str]:
    """解析聚合 group_by 配置，主形态为数组，兼容旧对象结构。"""
    group_by = aggregation_config.get("group_by", [])
    if isinstance(group_by, dict):
        group_by = [group_by]
    if isinstance(group_by, list):
        fields: list[str] = []
        for item in group_by:
            if isinstance(item, str):
                fields.append(str(item).strip())
                continue
            if not isinstance(item, dict):
                continue
            field = item.get(f"{file_type}_field") or item.get(file_type)
            if isinstance(field, list):
                fields.extend(str(sub).strip() for sub in field if str(sub).strip())
            elif field:
                fields.append(str(field).strip())
        return fields
    return []


def _build_aggregation_specs(aggregation_config: dict, df: pd.DataFrame, file_type: str) -> dict[str, tuple[str, str]]:
    """构建 pandas NamedAgg 配置。"""
    specs: dict[str, tuple[str, str]] = {}
    aggregations = aggregation_config.get("aggregations", []) or aggregation_config.get("aggre_fields", []) or []
    if isinstance(aggregations, dict):
        aggregations = [aggregations]
    for agg in aggregations:
        input_col = (
            agg.get(f"{file_type}_field")
            or agg.get("source_field" if file_type == "source" else "target_field")
            or agg.get(f"{file_type}_column")
            or agg.get("source_column" if file_type == "source" else "target_column")
            or agg.get("column")
        )
        output_col = agg.get("alias") or input_col
        func = agg.get("function", "sum")
        if input_col and output_col and input_col in df.columns:
            specs[output_col] = (input_col, func)
    return specs


def _get_compare_name(compare_cfg: dict[str, Any]) -> str:
    """获取比较项展示名，兼容旧 column 字段。"""
    return (
        str(compare_cfg.get("name") or "").strip()
        or str(compare_cfg.get("column") or "").strip()
        or str(compare_cfg.get("source_column") or "").strip()
        or str(compare_cfg.get("target_column") or "").strip()
        or "数值"
    )


def _get_diff_display_name(compare_cfg: dict[str, Any]) -> str:
    """获取差异列的业务展示名。"""
    compare_name = _get_compare_name(compare_cfg)
    if compare_name.endswith("差异"):
        return compare_name
    return f"{compare_name}差异"


def _format_export_dataframe(
    df: pd.DataFrame,
    sheet_type: str,
    rule: dict[str, Any],
    compare_columns_config: list[dict[str, Any]] | None = None,
) -> pd.DataFrame:
    """按 sheet 类型格式化导出列名。"""
    if df is None or df.empty:
        return df

    formatted = df.copy()
    source_name = ((rule.get("source_file") or {}).get("table_name") or "源").strip()
    target_name = ((rule.get("target_file") or {}).get("table_name") or "目标").strip()
    rename_map: dict[str, str] = {}

    if sheet_type == "source_only":
        formatted = formatted[[col for col in formatted.columns if not col.startswith("target_")]]
    elif sheet_type == "target_only":
        formatted = formatted[[col for col in formatted.columns if not col.startswith("source_")]]

    compare_name_map = {
        f"diff_{_get_compare_name(cfg)}": _get_diff_display_name(cfg)
        for cfg in (compare_columns_config or [])
    }

    for col in formatted.columns:
        if col in compare_name_map:
            rename_map[col] = compare_name_map[col]
            continue
        if sheet_type == "matched_with_diff":
            if col.startswith("source_"):
                rename_map[col] = f"{source_name}.{col[len('source_'):]}"
                continue
            if col.startswith("target_"):
                rename_map[col] = f"{target_name}.{col[len('target_'):]}"
                continue
        elif sheet_type == "source_only" and col.startswith("source_"):
            rename_map[col] = col[len("source_"):]
            continue
        elif sheet_type == "target_only" and col.startswith("target_"):
            rename_map[col] = col[len("target_"):]
            continue

    if rename_map:
        formatted = formatted.rename(columns=rename_map)
    return formatted


def _apply_aggregation(
    df: pd.DataFrame,
    aggregation_config: dict,
    rule_id: str,
    file_type: str
) -> pd.DataFrame:
    """应用分组聚合"""
    group_by = _resolve_group_by_columns(aggregation_config, file_type)
    aggregations = aggregation_config.get("aggregations", []) or aggregation_config.get("aggre_fields", [])
    if isinstance(aggregations, dict):
        aggregations = [aggregations]
    
    if not group_by or not aggregations:
        return df
    
    # 检查 group_by 列是否存在
    missing_cols = [col for col in group_by if col not in df.columns]
    if missing_cols:
        logger.warning(f"[recon] [{rule_id}] {file_type} 缺少分组列: {missing_cols}")
        return df
    
    agg_specs = _build_aggregation_specs(aggregation_config, df, file_type)

    if not agg_specs:
        return df
    
    try:
        grouped = df.groupby(group_by, as_index=False).agg(
            **{
                output_col: pd.NamedAgg(column=input_col, aggfunc=func)
                for output_col, (input_col, func) in agg_specs.items()
            }
        )
        logger.info(f"[recon] [{rule_id}] {file_type} 聚合后 {len(grouped)} 行")
        return grouped
    except Exception as e:
        logger.error(f"[recon] [{rule_id}] {file_type} 聚合失败: {e}")
        return df


def _apply_key_transformations(
    df: pd.DataFrame,
    key_column: str,
    transformations: list[dict[str, Any]],
    file_type: str,
    rule_id: str
) -> pd.DataFrame:
    """
    应用关键列的数据清洗转换
    
    支持的转换操作（按执行顺序）：
    1. regex_extract: 使用正则表达式提取匹配内容
    2. regex_replace: 使用正则表达式替换匹配内容
    3. strip_prefix: 去除前缀字符串
    4. strip_suffix: 去除后缀字符串
    5. strip_whitespace: 去除首尾空白字符
    6. lowercase: 转换为小写
    
    Args:
        df: DataFrame
        key_column: 关键列名
        transformations: 转换配置
        file_type: "source" 或 "target"
        rule_id: 规则ID
    
    Returns:
        转换后的 DataFrame
    """
    if key_column not in df.columns:
        return df

    original_values = df[key_column].astype(str)
    transformed_values = original_values

    # 允许对同一字段按顺序应用多条转换规则
    for trans_config in [t for t in transformations if t]:
        op_type = trans_config.get("type")
        # 1. 正则表达式提取
        regex_extract = trans_config.get("pattern") if op_type == "regex_extract" else None
        if regex_extract:
            try:
                extracted = transformed_values.str.extract(regex_extract, expand=False)
                if isinstance(extracted, pd.DataFrame):
                    for col in extracted.columns:
                        if extracted[col].notna().any():
                            extracted = extracted[col]
                            break
                transformed_values = extracted.fillna(transformed_values)
                matched_count = extracted.notna().sum()
                logger.info(f"[recon] [{rule_id}] {file_type} 列 '{key_column}' 正则提取 '{regex_extract}' 匹配 {matched_count} 个值")
            except Exception as e:
                logger.warning(f"[recon] [{rule_id}] {file_type} 列 '{key_column}' 正则提取失败: {e}")

        # 2. 正则表达式替换
        if op_type == "regex_replace":
            pattern = trans_config.get("pattern")
            replacement = trans_config.get("replacement", "")
            if pattern:
                try:
                    transformed_values = transformed_values.str.replace(pattern, replacement, regex=True)
                    logger.info(f"[recon] [{rule_id}] {file_type} 列 '{key_column}' 正则替换 '{pattern}' -> '{replacement}'")
                except Exception as e:
                    logger.warning(f"[recon] [{rule_id}] {file_type} 列 '{key_column}' 正则替换失败: {e}")

        strip_prefix = trans_config.get("value") if op_type == "strip_prefix" else None
        if strip_prefix:
            transformed_values = transformed_values.str.lstrip(strip_prefix)
            logger.info(f"[recon] [{rule_id}] {file_type} 列 '{key_column}' 去除前缀 '{strip_prefix}'")

        strip_suffix = trans_config.get("value") if op_type == "strip_suffix" else None
        if strip_suffix:
            transformed_values = transformed_values.str.removesuffix(strip_suffix)
            logger.info(f"[recon] [{rule_id}] {file_type} 列 '{key_column}' 去除后缀 '{strip_suffix}'")

        if op_type == "strip_whitespace":
            transformed_values = transformed_values.str.strip()
            logger.info(f"[recon] [{rule_id}] {file_type} 列 '{key_column}' 去除首尾空白")

        if op_type == "lowercase":
            transformed_values = transformed_values.str.lower()
            logger.info(f"[recon] [{rule_id}] {file_type} 列 '{key_column}' 转换为小写")

    df[key_column] = transformed_values

    changed_count = (original_values != transformed_values).sum()
    if changed_count > 0:
        logger.info(f"[recon] [{rule_id}] {file_type} 列 '{key_column}' 共转换了 {changed_count} 个值")
    
    return df


def _get_transformation_chain(
    key_columns_config: dict | None,
    file_type: str,
    key_column: str,
) -> list[dict[str, Any]]:
    """按字段解析转换链，仅支持 source/target -> field -> operations 结构。"""
    transformations = (key_columns_config or {}).get("transformations", {})
    if not isinstance(transformations, dict):
        return []
    side_config = transformations.get(file_type, {})
    if not isinstance(side_config, dict):
        return []
    operations = side_config.get(key_column, [])
    if not isinstance(operations, list):
        return []
    return [dict(item) for item in operations if isinstance(item, dict)]


def _execute_comparison(
    df_source: pd.DataFrame,
    df_target: pd.DataFrame,
    key_mappings: list[dict[str, str]],
    compare_columns_config: list[dict],
    rule_id: str,
    key_columns_config: dict = None
) -> dict:
    """
    执行数据比较
    
    Returns:
        {
            "matched_with_diff": DataFrame,  # 匹配但有差异
            "source_only": DataFrame,        # 源文件独有
            "target_only": DataFrame,        # 目标文件独有
            "matched_exact": DataFrame       # 完全匹配
        }
    """
    result = {
        "matched_with_diff": pd.DataFrame(),
        "source_only": pd.DataFrame(),
        "target_only": pd.DataFrame(),
        "matched_exact": pd.DataFrame()
    }
    
    if not key_mappings:
        logger.warning(f"[recon] [{rule_id}] 未配置关键列，无法执行比较")
        return result

    source_key_cols_raw = [item["source_field"] for item in key_mappings]
    target_key_cols_raw = [item["target_field"] for item in key_mappings]
    source_missing = [col for col in source_key_cols_raw if col not in df_source.columns]
    target_missing = [col for col in target_key_cols_raw if col not in df_target.columns]
    
    if source_missing:
        logger.warning(f"[recon] [{rule_id}] 源文件缺少关键列: {source_missing}")
        return result
    if target_missing:
        logger.warning(f"[recon] [{rule_id}] 目标文件缺少关键列: {target_missing}")
        return result
    
    # 应用数据清洗转换（在添加前缀之前）
    if key_columns_config:
        if key_columns_config.get("transformations"):
            df_source = df_source.copy()
            df_target = df_target.copy()
            for source_key_col in source_key_cols_raw:
                trans_chain = _get_transformation_chain(key_columns_config, "source", source_key_col)
                if source_key_col in df_source.columns and trans_chain:
                    df_source = _apply_key_transformations(
                        df_source, source_key_col, trans_chain, "source", rule_id
                    )
            for target_key_col in target_key_cols_raw:
                trans_chain = _get_transformation_chain(key_columns_config, "target", target_key_col)
                if target_key_col in df_target.columns and trans_chain:
                    df_target = _apply_key_transformations(
                        df_target, target_key_col, trans_chain, "target", rule_id
                    )
    
    # 添加前缀以区分来源
    df_source_prefixed = df_source.add_prefix("source_")
    df_target_prefixed = df_target.add_prefix("target_")
    
    source_key_cols = [f"source_{col}" for col in source_key_cols_raw]
    target_key_cols = [f"target_{col}" for col in target_key_cols_raw]
    
    # 创建合并键（处理 NaN 值）
    def _create_merge_key(df, cols):
        """创建合并键，处理 NaN 值"""
        # 将每列转换为字符串，并用空字符串填充 NaN
        str_cols = df[cols].astype(str).fillna('')
        # 使用 join 连接各列
        return str_cols.agg("||".join, axis=1)
    
    df_source_prefixed["_merge_key"] = _create_merge_key(df_source_prefixed, source_key_cols)
    df_target_prefixed["_merge_key"] = _create_merge_key(df_target_prefixed, target_key_cols)
    
    # 执行外连接
    merged = pd.merge(
        df_source_prefixed,
        df_target_prefixed,
        on="_merge_key",
        how="outer",
        indicator=True
    )
    
    # 分类记录
    source_only = merged[merged["_merge"] == "left_only"].drop(columns=["_merge_key", "_merge"])
    target_only = merged[merged["_merge"] == "right_only"].drop(columns=["_merge_key", "_merge"])
    both = merged[merged["_merge"] == "both"].drop(columns=["_merge_key", "_merge"])
    
    result["source_only"] = source_only
    result["target_only"] = target_only
    
    logger.info(f"[recon] [{rule_id}] 源独有: {len(source_only)}, 目标独有: {len(target_only)}, 匹配: {len(both)}")
    
    if len(both) == 0:
        return result
    
    # 比较匹配记录中的数值差异
    if not compare_columns_config:
        result["matched_exact"] = both
        return result
    
    # 计算差异
    has_diff_mask = pd.Series([False] * len(both), index=both.index)
    
    for cfg in compare_columns_config:
        compare_name = _get_compare_name(cfg)
        if not compare_name:
            continue
        
        source_col_name = cfg.get("source_column")
        target_col_name = cfg.get("target_column")
        if not source_col_name or not target_col_name:
            logger.warning(f"[recon] [{rule_id}] 比较项缺少 source_column 或 target_column: {compare_name}")
            continue
        
        source_col = f"source_{source_col_name}"
        target_col = f"target_{target_col_name}"
        
        if source_col not in both.columns or target_col not in both.columns:
            logger.warning(f"[recon] [{rule_id}] 比较列不存在: {source_col} 或 {target_col}")
            continue
        
        tolerance = cfg.get("tolerance", 0)
        
        # 转换为数值
        source_vals = pd.to_numeric(both[source_col], errors="coerce").fillna(0)
        target_vals = pd.to_numeric(both[target_col], errors="coerce").fillna(0)
        
        diff = (source_vals - target_vals).abs()
        
        col_has_diff = diff > tolerance
        
        has_diff_mask = has_diff_mask | col_has_diff
        
        both[f"diff_{compare_name}"] = source_vals - target_vals
        
        logger.info(f"[recon] [{rule_id}] 比较列 {compare_name}: {source_col_name} vs {target_col_name}, 差异 {col_has_diff.sum()} 条")
    
    result["matched_with_diff"] = both[has_diff_mask]
    result["matched_exact"] = both[~has_diff_mask]
    
    logger.info(f"[recon] [{rule_id}] 有差异: {len(result['matched_with_diff'])}, 完全匹配: {len(result['matched_exact'])}")
    
    return result


def _write_recon_result(
    diff_result: dict,
    output_dir: str,
    output_config: dict,
    rule_id: str,
    rule_name: str,
    rule: dict | None = None,
    key_columns_config: dict = None,
    compare_columns_config: list = None
) -> str:
    """写出核对结果，并对关键列添加颜色标记"""
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter
    
    # 生成文件名
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_rule_name = re.sub(r'[\\/:*?"<>|]', "_", rule_name)
    filename = f"{safe_rule_name}_核对结果_{timestamp}.xlsx"
    output_path = str(Path(output_dir) / filename)
    
    sheets_config = output_config.get("sheets", {})
    
    # 收集需要标记的列
    _get_marked_columns._rule_context = rule or {}
    marked_columns = _get_marked_columns(key_columns_config, compare_columns_config)
    
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # 汇总表
        if sheets_config.get("summary", {}).get("enabled", True):
            summary_data = {
                "项目": ["规则ID", "规则名称", "差异记录数", "源文件独有数", "目标文件独有数", "完全匹配数", "生成时间"],
                "值": [
                    rule_id,
                    rule_name,
                    len(diff_result.get("matched_with_diff", [])),
                    len(diff_result.get("source_only", [])),
                    len(diff_result.get("target_only", [])),
                    len(diff_result.get("matched_exact", [])),
                    timestamp
                ]
            }
            pd.DataFrame(summary_data).to_excel(
                writer,
                sheet_name=sheets_config.get("summary", {}).get("name", "核对汇总"),
                index=False
            )
        
        # 差异记录
        sheet_name_matched_with_diff = sheets_config.get("matched_with_diff", {}).get("name", "差异记录")
        if sheets_config.get("matched_with_diff", {}).get("enabled", True):
            df = diff_result.get("matched_with_diff", pd.DataFrame())
            if len(df) > 0:
                df = _format_export_dataframe(
                    df,
                    "matched_with_diff",
                    rule or {},
                    compare_columns_config or [],
                )
                df.to_excel(
                    writer,
                    sheet_name=sheet_name_matched_with_diff,
                    index=False
                )
        
        # 源文件独有
        sheet_name_source_only = sheets_config.get("source_only", {}).get("name", "源文件独有")
        if sheets_config.get("source_only", {}).get("enabled", True):
            df = diff_result.get("source_only", pd.DataFrame())
            if len(df) > 0:
                df = _format_export_dataframe(
                    df,
                    "source_only",
                    rule or {},
                    compare_columns_config or [],
                )
                df.to_excel(
                    writer,
                    sheet_name=sheet_name_source_only,
                    index=False
                )
        
        # 目标文件独有
        sheet_name_target_only = sheets_config.get("target_only", {}).get("name", "目标文件独有")
        if sheets_config.get("target_only", {}).get("enabled", True):
            df = diff_result.get("target_only", pd.DataFrame())
            if len(df) > 0:
                df = _format_export_dataframe(
                    df,
                    "target_only",
                    rule or {},
                    compare_columns_config or [],
                )
                df.to_excel(
                    writer,
                    sheet_name=sheet_name_target_only,
                    index=False
                )
        
        # 应用颜色标记
        workbook = writer.book
        
        # 定义颜色
        mapping_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")  # 黄色 - mapping列
        compare_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")  # 蓝色 - 比对列
        diff_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")    # 橙色 - 差异列
        header_font = Font(bold=True)
        
        # 标记差异记录sheet
        if sheet_name_matched_with_diff in workbook.sheetnames:
            _apply_column_highlighting(
                workbook[sheet_name_matched_with_diff],
                marked_columns,
                mapping_fill,
                compare_fill,
                diff_fill,
                header_font
            )
        
        # 标记源文件独有sheet
        if sheet_name_source_only in workbook.sheetnames:
            _apply_column_highlighting(
                workbook[sheet_name_source_only],
                marked_columns,
                mapping_fill,
                compare_fill,
                diff_fill,
                header_font
            )
        
        # 标记目标文件独有sheet
        if sheet_name_target_only in workbook.sheetnames:
            _apply_column_highlighting(
                workbook[sheet_name_target_only],
                marked_columns,
                mapping_fill,
                compare_fill,
                diff_fill,
                header_font
            )
    
    logger.info(f"[recon] [{rule_id}] 结果已输出: {output_path}")
    return output_path


def _get_marked_columns(key_columns_config: dict, compare_columns_config: list) -> dict:
    """
    获取需要标记的列信息
    
    Returns:
        {
            "mapping_source": ["source_sup订单号", ...],  # 黄色 - source mapping列
            "mapping_target": ["target_第三方订单号", ...],  # 黄色 - target mapping列
            "compare_source": ["source_发生-", ...],  # 蓝色 - source 比对列
            "compare_target": ["target_合作方分销收入", ...],  # 蓝色 - target 比对列
            "diff": ["diff_发生减", ...]  # 橙色 - 差异列
        }
    """
    marked = {
        "mapping_source": [],
        "mapping_target": [],
        "compare_source": [],
        "compare_target": [],
        "diff": []
    }
    
    source_label = "源"
    target_label = "目标"
    if rule_context := getattr(_get_marked_columns, "_rule_context", None):
        source_label = ((rule_context.get("source_file") or {}).get("table_name") or source_label).strip()
        target_label = ((rule_context.get("target_file") or {}).get("table_name") or target_label).strip()

    if key_columns_config:
        for mapping in _get_key_mappings(key_columns_config):
            source_col = mapping.get("source_field")
            target_col = mapping.get("target_field")
            if source_col:
                marked["mapping_source"].extend([f"source_{source_col}", f"{source_label}.{source_col}", source_col])
            if target_col:
                marked["mapping_target"].extend([f"target_{target_col}", f"{target_label}.{target_col}", target_col])
    
    if compare_columns_config:
        for cfg in compare_columns_config:
            compare_name = _get_compare_name(cfg)
            source_col = cfg.get("source_column")
            target_col = cfg.get("target_column")
            
            if source_col:
                marked["compare_source"].extend([f"source_{source_col}", f"{source_label}.{source_col}", source_col])
            if target_col:
                marked["compare_target"].extend([f"target_{target_col}", f"{target_label}.{target_col}", target_col])
            if compare_name:
                marked["diff"].extend([f"diff_{compare_name}", _get_diff_display_name(cfg)])
    
    return marked


def _apply_column_highlighting(
    worksheet,
    marked_columns: dict,
    mapping_fill: PatternFill,
    compare_fill: PatternFill,
    diff_fill: PatternFill,
    header_font: Font
):
    """对 worksheet 的指定列应用颜色标记"""
    if not worksheet or worksheet.max_row == 0:
        return
    
    # 获取表头行
    header_row = 1
    header_map = {}
    for col_idx in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=header_row, column=col_idx)
        header_map[cell.value] = col_idx
        # 表头加粗
        cell.font = header_font
    
    # 标记 mapping 列（黄色）
    for col_name in marked_columns.get("mapping_source", []) + marked_columns.get("mapping_target", []):
        if col_name in header_map:
            col_idx = header_map[col_name]
            for row_idx in range(header_row, worksheet.max_row + 1):
                worksheet.cell(row=row_idx, column=col_idx).fill = mapping_fill
    
    # 标记比对列（蓝色）
    for col_name in marked_columns.get("compare_source", []) + marked_columns.get("compare_target", []):
        if col_name in header_map:
            col_idx = header_map[col_name]
            for row_idx in range(header_row, worksheet.max_row + 1):
                worksheet.cell(row=row_idx, column=col_idx).fill = compare_fill
    
    # 标记差异列（橙色）
    for col_name in marked_columns.get("diff", []):
        if col_name in header_map:
            col_idx = header_map[col_name]
            for row_idx in range(header_row, worksheet.max_row + 1):
                worksheet.cell(row=row_idx, column=col_idx).fill = diff_fill


# ════════════════════════════════════════════════════════════════════════════
# 文件读取
# ════════════════════════════════════════════════════════════════════════════

def _read_file_as_df(file_path: str) -> pd.DataFrame:
    """读取 CSV 或 Excel 文件为 DataFrame"""
    path = resolve_upload_file_path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"文件不存在: {file_path}")
    
    ext = path.suffix.lower()
    if ext == ".csv":
        try:
            return pd.read_csv(path, encoding="utf-8-sig")
        except UnicodeDecodeError:
            import chardet
            with open(path, "rb") as f:
                enc = chardet.detect(f.read()).get("encoding", "gbk")
            return pd.read_csv(path, encoding=enc)
    elif ext in (".xlsx", ".xls"):
        return pd.read_excel(path)
    else:
        raise ValueError(f"不支持的文件格式: {ext}")
