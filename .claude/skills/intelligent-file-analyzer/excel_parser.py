"""Excel 文件解析工具

用于智能文件分析技能，提供 Excel 文件解析、表头提取、数据类型检测等功能。
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any
from dataclasses import dataclass

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    raise ImportError("需要安装 openpyxl: pip install openpyxl")

logger = logging.getLogger(__name__)


@dataclass
class SheetInfo:
    """Sheet 信息数据类"""
    name: str
    headers: list[str]
    column_types: dict[str, str]
    row_count: int
    has_data: bool
    sample_data: list[dict[str, Any]]


@dataclass
class ExcelFileInfo:
    """Excel 文件信息数据类"""
    filename: str
    filepath: str
    sheet_count: int
    sheets: list[SheetInfo]


def detect_column_type(values: list[Any]) -> str:
    """检测列的数据类型

    Args:
        values: 列的值列表（不包括表头）

    Returns:
        数据类型字符串: "number", "date", "text", "boolean", "empty"
    """
    if not values or all(v is None or v == "" for v in values):
        return "empty"

    # 过滤掉空值
    non_empty_values = [v for v in values if v is not None and v != ""]

    if not non_empty_values:
        return "empty"

    # 检查是否为数字类型
    number_count = sum(1 for v in non_empty_values if isinstance(v, (int, float)))
    if number_count / len(non_empty_values) > 0.8:  # 80%以上是数字
        return "number"

    # 检查是否为日期类型
    date_count = sum(1 for v in non_empty_values
                     if hasattr(v, '__class__') and 'datetime' in v.__class__.__name__.lower())
    if date_count / len(non_empty_values) > 0.8:
        return "date"

    # 检查是否为布尔类型
    bool_count = sum(1 for v in non_empty_values if isinstance(v, bool))
    if bool_count / len(non_empty_values) > 0.8:
        return "boolean"

    # 默认为文本类型
    return "text"


def parse_sheet(sheet, max_sample_rows: int = 5) -> SheetInfo | None:
    """解析单个 sheet

    Args:
        sheet: openpyxl worksheet 对象
        max_sample_rows: 最大采样行数（用于数据预览和类型检测）

    Returns:
        SheetInfo 对象，如果 sheet 为空则返回 None
    """
    # 获取 sheet 名称
    sheet_name = sheet.title

    # 获取所有行
    rows = list(sheet.iter_rows(values_only=True))

    if not rows or len(rows) < 1:
        logger.warning(f"Sheet '{sheet_name}' 为空")
        return None

    # 第一行为表头
    header_row = rows[0]
    headers = [str(cell) if cell is not None else f"Column_{i}"
               for i, cell in enumerate(header_row)]

    # 检查是否有数据行
    data_rows = rows[1:]
    has_data = len(data_rows) > 0
    row_count = len(data_rows)

    if not has_data:
        logger.info(f"Sheet '{sheet_name}' 只有表头，没有数据行")
        return SheetInfo(
            name=sheet_name,
            headers=headers,
            column_types={},
            row_count=0,
            has_data=False,
            sample_data=[]
        )

    # 采样数据用于类型检测（最多取 max_sample_rows 或所有行）
    sample_size = min(max_sample_rows, len(data_rows))
    sample_rows = data_rows[:sample_size]

    # 检测每列的数据类型
    column_types = {}
    for i, header in enumerate(headers):
        column_values = [row[i] if i < len(row) else None for row in data_rows]
        column_types[header] = detect_column_type(column_values)

    # 构建采样数据
    sample_data = []
    for row in sample_rows:
        row_dict = {}
        for i, header in enumerate(headers):
            value = row[i] if i < len(row) else None
            row_dict[header] = value
        sample_data.append(row_dict)

    return SheetInfo(
        name=sheet_name,
        headers=headers,
        column_types=column_types,
        row_count=row_count,
        has_data=has_data,
        sample_data=sample_data
    )


def parse_excel_file(filepath: str | Path, skip_empty_sheets: bool = True) -> ExcelFileInfo:
    """解析 Excel 文件，提取所有 sheets 的信息

    Args:
        filepath: Excel 文件路径
        skip_empty_sheets: 是否跳过空 sheet（只有表头或完全为空）

    Returns:
        ExcelFileInfo 对象

    Raises:
        FileNotFoundError: 文件不存在
        Exception: 文件解析失败
    """
    filepath = Path(filepath)

    if not filepath.exists():
        raise FileNotFoundError(f"文件不存在: {filepath}")

    try:
        # 加载工作簿（read_only=True 提高性能）
        wb = load_workbook(filepath, read_only=True, data_only=True)

        sheets_info = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_info = parse_sheet(sheet)

            # 如果 skip_empty_sheets=True，跳过空 sheet
            if sheet_info and (not skip_empty_sheets or sheet_info.has_data):
                sheets_info.append(sheet_info)

        wb.close()

        return ExcelFileInfo(
            filename=filepath.name,
            filepath=str(filepath),
            sheet_count=len(sheets_info),
            sheets=sheets_info
        )

    except Exception as e:
        logger.error(f"解析 Excel 文件失败: {filepath}, 错误: {e}")
        raise Exception(f"Excel 文件解析失败: {str(e)}")


def parse_multiple_files(filepaths: list[str | Path], skip_empty_sheets: bool = True) -> list[ExcelFileInfo]:
    """解析多个 Excel 文件

    Args:
        filepaths: Excel 文件路径列表
        skip_empty_sheets: 是否跳过空 sheet

    Returns:
        ExcelFileInfo 对象列表
    """
    results = []
    for filepath in filepaths:
        try:
            file_info = parse_excel_file(filepath, skip_empty_sheets)
            results.append(file_info)
        except Exception as e:
            logger.error(f"跳过文件 {filepath}: {e}")
            continue

    return results


def get_all_sheets(files: list[ExcelFileInfo]) -> list[tuple[str, str, SheetInfo]]:
    """获取所有文件的所有 sheets（扁平化）

    Args:
        files: ExcelFileInfo 列表

    Returns:
        (filename, sheet_name, sheet_info) 元组列表
    """
    all_sheets = []
    for file_info in files:
        for sheet_info in file_info.sheets:
            all_sheets.append((file_info.filename, sheet_info.name, sheet_info))

    return all_sheets


# 测试代码
if __name__ == "__main__":
    import sys

    logging.basicConfig(level=logging.INFO)

    if len(sys.argv) < 2:
        print("用法: python excel_parser.py <excel_file_path>")
        sys.exit(1)

    filepath = sys.argv[1]

    try:
        file_info = parse_excel_file(filepath)
        print(f"\n文件: {file_info.filename}")
        print(f"Sheet 数量: {file_info.sheet_count}")
        print("-" * 60)

        for sheet in file_info.sheets:
            print(f"\nSheet: {sheet.name}")
            print(f"  行数: {sheet.row_count}")
            print(f"  列数: {len(sheet.headers)}")
            print(f"  表头: {sheet.headers}")
            print(f"  列类型: {sheet.column_types}")
            if sheet.sample_data:
                print(f"  样例数据（前{len(sheet.sample_data)}行）:")
                for i, row in enumerate(sheet.sample_data, 1):
                    print(f"    行{i}: {row}")

    except Exception as e:
        print(f"错误: {e}")
        sys.exit(1)
