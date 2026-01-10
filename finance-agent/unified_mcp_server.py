"""
Financial Agent Unified MCP Server
统一的财务助手 MCP 服务器 - 包含对账和数据整理功能
"""
import sys
import asyncio
from pathlib import Path
from mcp import types
from mcp.server import Server
from mcp.server.sse import SseServerTransport
from starlette.applications import Starlette
from starlette.routing import Route, Mount
from starlette.responses import Response, JSONResponse, FileResponse
from starlette.staticfiles import StaticFiles
import uvicorn
import logging

# 导入对账模块
from reconciliation.mcp_server.config import DEFAULT_HOST, DEFAULT_PORT
from reconciliation.mcp_server.tools import create_tools as create_recon_tools, handle_tool_call as handle_recon_call

# 导入数据整理模块  
from data_preparation.mcp_server.tools import create_tools as create_prep_tools, handle_tool_call as handle_prep_call
from data_preparation.mcp_server.config import OUTPUT_DIR, REPORT_DIR

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# 创建统一的 MCP Server
mcp_server = Server("financial-agent-mcp-server")


@mcp_server.list_tools()
async def list_tools() -> list[types.Tool]:
    """列出所有工具（对账 + 数据整理）"""
    try:
        recon_tools = create_recon_tools()
        logger.info(f"对账工具数量: {len(recon_tools)}")
    except Exception as e:
        logger.error(f"加载对账工具失败: {str(e)}", exc_info=True)
        recon_tools = []
    
    try:
        prep_tools = create_prep_tools()
        logger.info(f"数据整理工具数量: {len(prep_tools)}")
    except Exception as e:
        logger.error(f"加载数据整理工具失败: {str(e)}", exc_info=True)
        prep_tools = []
    
    all_tools = recon_tools + prep_tools
    logger.info(f"总工具数量: {len(all_tools)}")
    return all_tools


@mcp_server.call_tool()
async def call_tool(name: str, arguments: dict) -> list[types.TextContent | types.ImageContent | types.EmbeddedResource]:
    """调用工具（自动路由到对应模块）"""
    try:
        # 根据工具名前缀路由到对应模块
        if name.startswith("reconciliation_"):
            result = await handle_recon_call(name, arguments)
        elif name.startswith("data_preparation_"):
            result = await handle_prep_call(name, arguments)
        else:
            result = {"error": f"未知的工具: {name}"}
        
        # 转换为字符串
        import json
        result_str = json.dumps(result, ensure_ascii=False, indent=2)
        
        return [types.TextContent(type="text", text=result_str)]
    
    except Exception as e:
        error_msg = f"工具调用失败: {str(e)}"
        logger.error(error_msg, exc_info=True)
        return [types.TextContent(type="text", text=error_msg)]


# 创建 SSE Transport
sse_transport = SseServerTransport("/messages/")


# HTTP 端点处理函数
async def handle_sse(request):
    """处理 SSE 连接"""
    async with sse_transport.connect_sse(request.scope, request.receive, request._send) as streams:
        await mcp_server.run(
            streams[0],
            streams[1],
            mcp_server.create_initialization_options()
        )


async def health_check(request):
    """健康检查"""
    return JSONResponse({
        "status": "healthy",
        "service": "financial-agent-mcp-server",
        "version": "1.0.0",
        "modules": ["reconciliation", "data_preparation"]
    })


async def download_file(request):
    """文件下载端点"""
    task_id = request.path_params.get("task_id")
    
    # 先查找报告文件获取输出文件路径
    from data_preparation.mcp_server.task_manager import TaskManager
    from data_preparation.mcp_server.config import (
        UPLOAD_DIR as PREP_UPLOAD_DIR,
        OUTPUT_DIR as PREP_OUTPUT_DIR,
        REPORT_DIR as PREP_REPORT_DIR,
        DATA_PREPARATION_SCHEMAS_FILE
    )
    
    # 尝试从任务管理器获取结果
    task_manager = TaskManager()
    result = task_manager.get_task_result(task_id)
    
    if result and result.get('output_file'):
        output_file = Path(result['output_file'])
        if output_file.exists():
            return FileResponse(
                str(output_file),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=output_file.name
            )
    
    # 如果找不到，尝试从报告文件中获取
    report_file = REPORT_DIR / f"{task_id}_report.json"
    if report_file.exists():
        import json
        with open(report_file, 'r') as f:
            report_data = json.load(f)
            output_path = report_data.get('processing_steps', [])[-1].get('details', {}).get('output_file')
            if output_path:
                output_file = Path(output_path)
                if output_file.exists():
                    return FileResponse(
                        str(output_file),
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        filename=output_file.name
                    )
    
    return JSONResponse({"error": f"文件不存在: {task_id}"}, status_code=404)


async def preview_file(request):
    """文件预览端点（返回文件基本信息）"""
    task_id = request.path_params.get("task_id")
    
    # 先查找报告文件获取输出文件路径
    from data_preparation.mcp_server.task_manager import TaskManager
    
    # 尝试从任务管理器获取结果
    task_manager = TaskManager()
    result = task_manager.get_task_result(task_id)
    
    output_file = None
    if result and result.get('output_file'):
        output_file = Path(result['output_file'])
    
    # 如果找不到，尝试从报告文件中获取
    if not output_file or not output_file.exists():
        report_file = REPORT_DIR / f"{task_id}_report.json"
        if report_file.exists():
            import json
            with open(report_file, 'r') as f:
                report_data = json.load(f)
                output_path = report_data.get('processing_steps', [])[-1].get('details', {}).get('output_file')
                if output_path:
                    output_file = Path(output_path)
    
    if not output_file or not output_file.exists():
        return JSONResponse({"error": f"文件不存在: {task_id}"}, status_code=404)
    
    import openpyxl
    try:
        wb = openpyxl.load_workbook(str(output_file), read_only=True, data_only=True)
        sheets = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheets.append({
                "name": sheet_name,
                "rows": ws.max_row,
                "columns": ws.max_column
            })
        
        return JSONResponse({
            "filename": output_file.name,
            "size": output_file.stat().st_size,
            "sheets": sheets,
            "download_url": f"/download/{task_id}"
        })
    except Exception as e:
        return JSONResponse({"error": f"预览失败: {str(e)}"}, status_code=500)


async def get_report(request):
    """获取详细报告"""
    task_id = request.path_params.get("task_id")
    
    # 查找报告文件
    report_file = REPORT_DIR / f"{task_id}_report.json"
    
    if not report_file.exists():
        return JSONResponse({"error": f"报告不存在: {task_id}"}, status_code=404)
    
    try:
        import json
        with open(report_file, 'r', encoding='utf-8') as f:
            report_data = json.load(f)
        return JSONResponse(report_data)
    except Exception as e:
        return JSONResponse({"error": f"读取报告失败: {str(e)}"}, status_code=500)


# 路由配置
routes = [
    Route("/sse", endpoint=handle_sse, methods=["GET", "POST"]),
    Route("/mcp", endpoint=handle_sse, methods=["GET", "POST"]),
    Mount("/messages/", app=sse_transport.handle_post_message),
    Route("/health", endpoint=health_check),
    Route("/download/{task_id}", endpoint=download_file),
    Route("/preview/{task_id}", endpoint=preview_file),
    Route("/report/{task_id}", endpoint=get_report),
]

app = Starlette(routes=routes)


async def main():
    """启动服务器"""
    host = DEFAULT_HOST
    port = DEFAULT_PORT
    
    # 动态获取工具列表用于显示
    try:
        tools = await list_tools()
        recon_tools = [t for t in tools if t.name.startswith("reconciliation_") or t.name == "file_upload" or t.name == "get_reconciliation"]
        prep_tools = [t for t in tools if t.name.startswith("data_preparation_")]
    except Exception as e:
        logger.warning(f"获取工具列表失败: {e}")
        recon_tools = []
        prep_tools = []
    
    recon_tools_text = "\n".join([f"  {i+1}. {t.name:<30} - {t.description}" for i, t in enumerate(recon_tools)])
    prep_tools_text = "\n".join([f"  {len(recon_tools)+i+1}. {t.name:<30} - {t.description}" for i, t in enumerate(prep_tools)])
    
    print(f"""
╔══════════════════════════════════════════════════════════════════╗
║          Financial Agent MCP Server 启动中...                    ║
╚══════════════════════════════════════════════════════════════════╝

🌐 服务端点:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  • SSE 端点:        http://{host}:{port}/sse
  • 消息端点:        http://{host}:{port}/messages/
  • 健康检查:        http://{host}:{port}/health
  • 文件下载:        http://{host}:{port}/download/{{task_id}}
  • 文件预览:        http://{host}:{port}/preview/{{task_id}}
  • 详细报告:        http://{host}:{port}/report/{{task_id}}

🛠️  可用工具（对账模块，{len(recon_tools)}个）:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{recon_tools_text}

🛠️  可用工具（数据整理模块，{len(prep_tools)}个）:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{prep_tools_text}

📖 使用说明:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  在 Dify 中配置:
    MCP 服务器地址: http://localhost:{port}/sse
    或使用 Docker:   http://host.docker.internal:{port}/sse

服务器正在运行...
""")
    
    config = uvicorn.Config(
        app,
        host=host,
        port=port,
        log_level="info"
    )
    server = uvicorn.Server(config)
    await server.serve()


if __name__ == "__main__":
    asyncio.run(main())
