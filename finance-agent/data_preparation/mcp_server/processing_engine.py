"""
数据整理处理引擎 - 协调整个数据处理流程
"""
import logging
import time
import pandas as pd
from pathlib import Path
from typing import Dict, List, Any
from datetime import datetime

from .models import ProcessingResult, ProcessingStep, ProcessingMetadata
from .extractor import DataExtractor
from .transformer import DataTransformer
from .template_writer import TemplateWriter
from .file_matcher import FileMatcher
from .template_reader import TemplateReader

logger = logging.getLogger(__name__)


class ProcessingEngine:
    """数据整理处理引擎"""
    
    def __init__(self, schema: Dict):
        self.schema = schema
        self.extractor = DataExtractor(schema)
        self.transformer = DataTransformer(schema)
        self.template_writer = TemplateWriter(schema)
        self.file_matcher = FileMatcher(schema)

        # 检测schema类型
        self.schema_type = schema.get("schema_type", "traditional")
        self.is_step_based = self.schema_type == "step_based"

        # 工作流控制
        self.workflow_controls = schema.get("workflow_controls", {})
        self.error_handling = self.workflow_controls.get("error_handling", {})
        self.max_errors = self.error_handling.get("max_errors", 999)
        self.error_count = 0

        # 处理步骤记录
        self.steps: List[ProcessingStep] = []

        # 步骤化处理的上下文（存储步骤间共享的数据）
        self.step_context: Dict[str, Any] = {}

        logger.info(f"初始化处理引擎: schema_type={self.schema_type}")
    
    def process(self, file_paths: List[str], output_dir: str, report_dir: str = None) -> ProcessingResult:
        """
        执行完整的数据整理流程

        Args:
            file_paths: 上传的文件路径列表
            output_dir: 输出目录
            report_dir: 报告目录

        Returns:
            处理结果
        """
        # 根据schema类型选择处理方式
        if self.is_step_based:
            return self.process_step_based(file_paths, output_dir, report_dir)
        else:
            return self.process_traditional(file_paths, output_dir, report_dir)

    def process_traditional(self, file_paths: List[str], output_dir: str, report_dir: str = None) -> ProcessingResult:
        """
        传统方式处理（向后兼容）

        Args:
            file_paths: 上传的文件路径列表
            output_dir: 输出目录
            report_dir: 报告目录

        Returns:
            处理结果
        """
        start_time = time.time()
        task_id = f"proc_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{id(self) & 0xFFFFFF:06x}"

        try:
            # 1. 文件匹配
            matched_files = self._match_files(file_paths)
            
            # 2. 数据提取
            extracted_data = self._extract_data(matched_files)
            
            # 3. 数据转换
            calculation_results = self._transform_data(extracted_data)
            
            # 4. 写入模板
            output_file = self._write_output(extracted_data, calculation_results, output_dir)
            
            # 5. 生成详细报告
            if report_dir:
                self._generate_report(task_id, matched_files, extracted_data, calculation_results, report_dir)
            
            # 6. 生成结果
            execution_time = time.time() - start_time
            metadata = ProcessingMetadata(
                project_name=self.schema.get("metadata", {}).get("project_name", "数据整理"),
                rule_version=self.schema.get("version", "1.0"),
                execution_time_seconds=round(execution_time, 2)
            )
            
            result = ProcessingResult(
                task_id=task_id,
                status="success",
                output_file=output_file,
                metadata=metadata,
                steps=self.steps
            )
            
            logger.info(f"数据整理完成: task_id={task_id}, 耗时={execution_time:.2f}秒")
            return result
        
        except Exception as e:
            logger.error(f"数据整理失败: {str(e)}", exc_info=True)
            return ProcessingResult(
                task_id=task_id,
                status="failed",
                error=str(e),
                steps=self.steps
            )
    
    def _generate_report(
        self,
        task_id: str,
        matched_files: Dict[str, str],
        extracted_data: Dict[str, Any],
        calculation_results: Dict[str, Any],
        report_dir: str
    ):
        """生成详细报告"""
        import json
        
        report = {
            "task_id": task_id,
            "timestamp": datetime.now().isoformat(),
            "schema_version": self.schema.get("version", "1.0"),
            "project_name": self.schema.get("metadata", {}).get("project_name", ""),
            
            "files_processed": {
                source_id: Path(file_path).name
                for source_id, file_path in matched_files.items()
            },
            
            "data_extraction": {
                source_id: {
                    "records": len(df) if hasattr(df, '__len__') else 0,
                    "columns": list(df.columns) if hasattr(df, 'columns') else []
                }
                for source_id, df in extracted_data.items()
            },
            
            "calculations": calculation_results,
            
            "processing_steps": [step.to_dict() for step in self.steps],
            
            "summary": {
                "total_steps": len(self.steps),
                "completed_steps": len([s for s in self.steps if s.status == "completed"]),
                "failed_steps": len([s for s in self.steps if s.status == "failed"])
            }
        }
        
        report_file = Path(report_dir) / f"{task_id}_report.json"
        with open(report_file, 'w', encoding='utf-8') as f:
            json.dump(report, f, ensure_ascii=False, indent=2)
        
        logger.info(f"详细报告已生成: {report_file}")
    
    def _match_files(self, file_paths: List[str]) -> Dict[str, str]:
        """匹配文件到数据源"""
        step = ProcessingStep(
            step_name="文件匹配",
            step_type="file_matching",
            status="processing",
            start_time=datetime.now().isoformat()
        )
        self.steps.append(step)
        
        try:
            matched = self.file_matcher.match_files(file_paths)
            step.status = "completed"
            step.end_time = datetime.now().isoformat()
            step.details = {"matched_files": matched}
            logger.info(f"文件匹配完成: {len(matched)} 个数据源")
            return matched
        except Exception as e:
            step.status = "failed"
            step.error_message = str(e)
            self._handle_error("file_matching", e)
            raise
    
    def _extract_data(self, matched_files: Dict[str, str]) -> Dict[str, Any]:
        """提取数据"""
        step = ProcessingStep(
            step_name="数据提取",
            step_type="extraction",
            status="processing",
            start_time=datetime.now().isoformat()
        )
        self.steps.append(step)
        
        extracted_data = {}
        total_records = 0
        
        try:
            for source_id, file_path in matched_files.items():
                try:
                    logger.info(f"提取数据: {source_id} <- {file_path}")
                    df = self.extractor.extract(source_id, file_path)
                    extracted_data[source_id] = df
                    total_records += len(df)
                except Exception as e:
                    error_action = self._handle_error("extraction", e)
                    if error_action == "skip_and_log":
                        logger.warning(f"跳过数据源 {source_id}: {str(e)}")
                        continue
                    else:
                        raise
            
            step.status = "completed"
            step.end_time = datetime.now().isoformat()
            step.records_processed = total_records
            step.details = {
                "sources_processed": len(extracted_data),
                "total_records": total_records
            }
            logger.info(f"数据提取完成: {len(extracted_data)} 个数据源, {total_records} 条记录")
            return extracted_data
        
        except Exception as e:
            step.status = "failed"
            step.error_message = str(e)
            raise
    
    def _transform_data(self, extracted_data: Dict[str, Any]) -> Dict[str, Any]:
        """转换数据"""
        step = ProcessingStep(
            step_name="数据转换",
            step_type="transformation",
            status="processing",
            start_time=datetime.now().isoformat()
        )
        self.steps.append(step)
        
        try:
            results = self.transformer.transform(extracted_data)
            step.status = "completed"
            step.end_time = datetime.now().isoformat()
            step.details = {"calculated_fields": list(results.keys())}
            logger.info(f"数据转换完成: {len(results)} 个计算字段")
            return results
        except Exception as e:
            step.status = "failed"
            step.error_message = str(e)
            error_action = self._handle_error("calculation", e)
            if error_action.startswith("use_default:"):
                default_value = float(error_action.split(":")[1])
                logger.warning(f"计算失败，使用默认值: {default_value}")
                return {"error": default_value}
            raise
    
    def _write_output(
        self,
        extracted_data: Dict[str, Any],
        calculation_results: Dict[str, Any],
        output_dir: str
    ) -> str:
        """写入输出文件"""
        step = ProcessingStep(
            step_name="写入输出",
            step_type="output",
            status="processing",
            start_time=datetime.now().isoformat()
        )
        self.steps.append(step)
        
        try:
            # 获取模板文件路径
            from .config import TEMPLATES_DIR, FINANCE_AGENT_DIR
            template_file = self.schema.get("template_mapping", {}).get("template_file", "template.xlsx")
            
            # 尝试多个可能的模板路径
            possible_paths = [
                TEMPLATES_DIR / template_file,  # data_preparation/templates/
                FINANCE_AGENT_DIR / "templates" / template_file,  # 绝对路径
                Path(template_file) if Path(template_file).is_absolute() and Path(template_file).exists() else None  # 绝对路径
            ]
            
            template_path = None
            for path in possible_paths:
                if path and path.exists():
                    template_path = path
                    break
            
            if not template_path:
                # 最后尝试相对路径（向后兼容）
                template_path = Path(template_file)
            
            # 如果没有模板，创建一个简单的输出文件
            if not template_path.exists():
                logger.warning(f"模板文件不存在: {template_path}, 创建简单输出")
                output_file = self._create_simple_output(extracted_data, calculation_results, output_dir)
            else:
                output_file = self.template_writer.write_to_template(
                    str(template_path),
                    output_dir,
                    calculation_results,
                    extracted_data
                )
            
            step.status = "completed"
            step.end_time = datetime.now().isoformat()
            step.details = {"output_file": output_file}
            logger.info(f"输出文件创建完成: {output_file}")
            return output_file
        
        except Exception as e:
            step.status = "failed"
            step.error_message = str(e)
            raise
    
    def _create_simple_output(
        self,
        extracted_data: Dict[str, Any],
        calculation_results: Dict[str, Any],
        output_dir: str
    ) -> str:
        """创建简单的输出文件（当没有模板时）"""
        import pandas as pd
        from datetime import datetime
        
        output_filename = f"data_preparation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = Path(output_dir) / output_filename
        
        with pd.ExcelWriter(str(output_path), engine='openpyxl') as writer:
            # 写入计算结果
            if calculation_results:
                calc_df = pd.DataFrame([calculation_results])
                calc_df.to_excel(writer, sheet_name='计算结果', index=False)
            
            # 写入提取的数据
            for source_id, df in extracted_data.items():
                if isinstance(df, pd.DataFrame):
                    sheet_name = source_id[:31]  # Excel sheet name 最多31字符
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        return str(output_path)
    
    def _handle_error(self, error_type: str, error: Exception) -> str:
        """
        处理错误
        
        Args:
            error_type: 错误类型 (extraction, calculation, etc.)
            error: 异常对象
        
        Returns:
            错误处理动作 (skip_and_log, use_default:0, 等)
        """
        self.error_count += 1
        
        # 检查是否超过最大错误数
        if self.error_count > self.max_errors:
            logger.error(f"错误次数超过最大限制: {self.max_errors}")
            raise Exception(f"错误次数超过最大限制: {self.max_errors}")
        
        # 根据错误类型返回处理动作
        if error_type == "extraction":
            return self.error_handling.get("on_extraction_error", "raise")
        elif error_type == "calculation":
            return self.error_handling.get("on_calculation_error", "raise")
        else:
            return "raise"

    def process_step_based(self, file_paths: List[str], output_dir: str, report_dir: str = None) -> ProcessingResult:
        """
        步骤化处理方式（新架构）

        Args:
            file_paths: 上传的文件路径列表
            output_dir: 输出目录
            report_dir: 报告目录

        Returns:
            处理结果
        """
        start_time = time.time()
        task_id = f"proc_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{id(self) & 0xFFFFFF:06x}"

        logger.info(f"开始步骤化处理: task_id={task_id}")

        try:
            # 获取模板配置
            template_config = self.schema.get("template_config", {})
            template_file = template_config.get("template_file", "template.xlsx")

            # 准备模板文件路径
            from .config import TEMPLATES_DIR
            template_path = TEMPLATES_DIR / template_file
            if not template_path.exists():
                raise FileNotFoundError(f"模板文件不存在: {template_path}")

            # 复制模板到输出目录作为工作文件
            import shutil
            output_filename_pattern = template_config.get("output_filename_pattern", "output_{timestamp}.xlsx")
            output_filename = output_filename_pattern.replace("{timestamp}", datetime.now().strftime('%Y%m%d_%H%M%S'))
            working_file = Path(output_dir) / output_filename
            shutil.copy(template_path, working_file)
            logger.info(f"工作文件已创建: {working_file}")

            # 获取处理步骤
            processing_steps = self.schema.get("processing_steps", [])
            if not processing_steps:
                raise ValueError("步骤化schema必须包含processing_steps")

            # 执行每个步骤
            for step_config in processing_steps:
                if not step_config.get("enabled", True):
                    logger.info(f"跳过已禁用的步骤: {step_config.get('step_id')}")
                    continue

                self._execute_step(step_config, file_paths, str(working_file))

            # 生成结果
            execution_time = time.time() - start_time
            metadata = ProcessingMetadata(
                project_name=self.schema.get("metadata", {}).get("project_name", "数据整理"),
                rule_version=self.schema.get("version", "3.0"),
                execution_time_seconds=round(execution_time, 2)
            )

            result = ProcessingResult(
                task_id=task_id,
                status="success",
                output_file=str(working_file),
                metadata=metadata,
                steps=self.steps
            )

            logger.info(f"步骤化处理完成: task_id={task_id}, 耗时={execution_time:.2f}秒")
            return result

        except Exception as e:
            logger.error(f"步骤化处理失败: {str(e)}", exc_info=True)
            return ProcessingResult(
                task_id=task_id,
                status="failed",
                error=str(e),
                steps=self.steps
            )

    def _execute_step(self, step_config: Dict[str, Any], file_paths: List[str], working_file: str):
        """
        执行单个处理步骤

        Args:
            step_config: 步骤配置
            file_paths: 上传的文件路径列表
            working_file: 工作文件路径（模板副本）
        """
        step_id = step_config.get("step_id", "unknown")
        step_name = step_config.get("step_name", step_id)
        step_type = step_config.get("step_type", "unknown")

        logger.info(f"执行步骤: {step_id} - {step_name}")

        # 创建步骤记录
        step = ProcessingStep(
            step_name=step_name,
            step_type=step_type,
            status="processing",
            start_time=datetime.now().isoformat()
        )
        # 记录 step_id 到 details
        step.details = {"step_id": step_id}
        self.steps.append(step)

        try:
            # 1. 检查依赖
            depends_on = step_config.get("depends_on", [])
            if depends_on:
                self._check_dependencies(depends_on)

            # 2. 读取数据
            data = self._read_step_data(step_config, file_paths, working_file)

            # 3. 执行模板操作
            if data is not None:
                self._execute_template_action(step_config, data, working_file)

            # 4. 存储输出变量
            output_variables = step_config.get("output_variables", {})
            if output_variables and data is not None:
                for var_name, var_config in output_variables.items():
                    self.step_context[var_name] = data

            step.status = "completed"
            step.end_time = datetime.now().isoformat()
            logger.info(f"步骤完成: {step_id}")

        except Exception as e:
            step.status = "failed"
            step.error_message = str(e)
            step.end_time = datetime.now().isoformat()
            logger.error(f"步骤失败: {step_id}, 错误: {str(e)}", exc_info=True)
            raise

    def _check_dependencies(self, depends_on: List[str]):
        """
        检查步骤依赖是否满足

        Args:
            depends_on: 依赖的步骤ID列表
        """
        completed_step_ids = set()
        for step in self.steps:
            if step.status == "completed":
                # 从step的details中提取step_id，如果没有则使用step_name
                step_id = step.details.get("step_id") if hasattr(step, 'details') and step.details else None
                if step_id:
                    completed_step_ids.add(step_id)

        for dep_step_id in depends_on:
            if dep_step_id not in completed_step_ids and dep_step_id not in self.step_context:
                raise ValueError(f"依赖的步骤未完成: {dep_step_id}")

        logger.debug(f"依赖检查通过: {depends_on}")

    def _read_step_data(self, step_config: Dict[str, Any], file_paths: List[str], working_file: str):
        """
        读取步骤数据

        Args:
            step_config: 步骤配置
            file_paths: 上传的文件路径列表
            working_file: 工作文件路径

        Returns:
            读取的数据（DataFrame或其他格式）
        """
        import pandas as pd

        # 检查是否有template_reference（用于读取模板数据）
        template_reference = step_config.get("template_reference")

        # 获取data_source配置
        data_source = step_config.get("data_source", {})
        source_type = data_source.get("source_type", "uploaded_file")

        logger.info(f"读取数据: source_type={source_type}")

        # 根据source_type读取数据
        if source_type == "template_range":
            # 从模板读取数据
            reader = TemplateReader(working_file)
            df = reader.read_by_config(data_source.get("template_reference", data_source))

            # 应用转换（如果有）
            transformations = step_config.get("transformations", [])
            if transformations:
                df = self._apply_transformations(df, transformations)

            return df

        elif source_type == "uploaded_file":
            # 从上传的文件读取数据
            file_pattern = data_source.get("file_pattern", "*")

            # 匹配文件
            matched_file = None
            for file_path in file_paths:
                file_name = Path(file_path).name
                if self._match_pattern(file_name, file_pattern):
                    matched_file = file_path
                    break

            if not matched_file:
                raise FileNotFoundError(f"未找到匹配的文件: {file_pattern}")

            # 使用extractor提取数据（需要临时创建data_source配置）
            # 这里简化处理，直接读取Excel
            extraction_rules = data_source.get("extraction_rules", {})
            conditional_extractions = data_source.get("conditional_extractions")

            df = self._extract_from_file(matched_file, extraction_rules, conditional_extractions)

            # 应用条件提取（如果有 match_with_template）
            if conditional_extractions and template_reference:
                # 读取模板数据用于匹配
                reader = TemplateReader(working_file)
                template_df = reader.read_by_config(template_reference)

                # 执行匹配过滤
                df = self._apply_template_matching(df, template_df, conditional_extractions)

            return df

        elif source_type == "step_output":
            # 从之前步骤的输出读取
            step_id = data_source.get("step_id")
            variable_name = data_source.get("variable_name")

            if variable_name in self.step_context:
                return self.step_context[variable_name]
            else:
                raise ValueError(f"步骤输出变量不存在: {variable_name}")

        else:
            raise ValueError(f"不支持的数据源类型: {source_type}")

    def _apply_transformations(self, df: pd.DataFrame, transformations: List[Dict[str, Any]]) -> pd.DataFrame:
        """
        应用数据转换

        Args:
            df: 输入DataFrame
            transformations: 转换配置列表

        Returns:
            转换后的DataFrame
        """
        import pandas as pd
        import re

        for trans in transformations:
            operation = trans.get("operation")
            output_field = trans.get("output_field")
            default_value = trans.get("default_value", 0)

            if operation == "calculate":
                # 计算公式
                formula = trans.get("formula", "")

                # 解析公式中的变量 {{field_name}}
                variables = re.findall(r'\{\{(\w+)\}\}', formula)

                # 先将NaN值填充为默认值
                for var in variables:
                    if var in df.columns:
                        df[var] = df[var].fillna(default_value)

                # 构建计算表达式
                calc_formula = formula
                for var in variables:
                    if var in df.columns:
                        calc_formula = calc_formula.replace(f"{{{{{var}}}}}", f"df['{var}']")
                    else:
                        logger.warning(f"字段不存在: {var}，使用默认值 {default_value}")
                        calc_formula = calc_formula.replace(f"{{{{{var}}}}}", str(default_value))

                try:
                    # 执行计算
                    df[output_field] = eval(calc_formula)
                    # 处理NaN值
                    df[output_field] = df[output_field].fillna(default_value)
                    logger.info(f"计算完成: {output_field} = {formula}")
                except Exception as e:
                    logger.error(f"计算失败: {formula}, 错误: {str(e)}")
                    df[output_field] = default_value

            elif operation == "aggregate":
                # 聚合操作
                group_by = trans.get("group_by", [])
                agg_func = trans.get("agg_func", "sum")
                input_field = trans.get("input_field")

                if group_by and input_field:
                    df[output_field] = df.groupby(group_by)[input_field].transform(agg_func)
                    logger.info(f"聚合完成: {output_field}")

            elif operation == "copy":
                # 复制字段
                source_field = trans.get("source_field")
                if source_field and source_field in df.columns:
                    df[output_field] = df[source_field]
                    logger.info(f"复制字段: {source_field} -> {output_field}")

        return df

    def _match_pattern(self, filename: str, pattern: str) -> bool:
        """
        匹配文件名与模式

        Args:
            filename: 文件名
            pattern: 模式（支持通配符 * 和 ?）

        Returns:
            是否匹配
        """
        import fnmatch
        return fnmatch.fnmatch(filename.lower(), pattern.lower())

    def _extract_from_file(self, file_path: str, extraction_rules: Dict[str, Any], conditional_extractions: Dict[str, Any] = None) -> pd.DataFrame:
        """
        从文件中提取数据

        Args:
            file_path: 文件路径
            extraction_rules: 提取规则
            conditional_extractions: 条件提取配置（可选）

        Returns:
            提取的DataFrame
        """
        import pandas as pd
        import openpyxl

        # 检查是否有多级表头
        multi_index_header = extraction_rules.get("multi_index_header")
        sheet_name = extraction_rules.get("sheet_name", 0)
        skip_rows = extraction_rules.get("skip_rows", 0)
        columns_mapping = extraction_rules.get("columns_mapping", {})

        # 读取Excel
        try:
            if multi_index_header:
                # 使用多级表头
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=multi_index_header)
                logger.info(f"Excel 读取成功（多级表头）: {file_path}, 行数={len(df)}")

                # 将多级列名扁平化
                if isinstance(df.columns, pd.MultiIndex):
                    new_columns = []
                    seen_cols = {}
                    for i, col in enumerate(df.columns.values):
                        if isinstance(col, tuple):
                            # 过滤掉 "Unnamed" 和空值
                            parts = [str(c) for c in col if c and 'Unnamed' not in str(c)]
                            new_col = '_'.join(parts) if parts else f'Unnamed_{i}'
                        else:
                            new_col = str(col)

                        # 处理重复列名
                        if new_col in seen_cols:
                            seen_cols[new_col] += 1
                            new_col = f"{new_col}_{seen_cols[new_col]}"
                        else:
                            seen_cols[new_col] = 0

                        new_columns.append(new_col)

                    df.columns = new_columns
                    logger.info(f"多级表头扁平化完成，列名: {list(df.columns[:10])}")
            else:
                df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=skip_rows)
                logger.info(f"Excel 读取成功: {file_path}, 行数={len(df)}")
        except Exception as e:
            logger.error(f"Excel 读取失败: {file_path}, 错误: {str(e)}")
            raise

        # 应用列映射
        if columns_mapping:
            df = self._apply_column_mapping(df, columns_mapping)

        # 应用条件提取
        if conditional_extractions:
            df = self._apply_conditional_extraction_in_step(df, conditional_extractions)

        return df

    def _apply_column_mapping(self, df: pd.DataFrame, mapping: Dict) -> pd.DataFrame:
        """应用列映射"""
        rename_dict = {}

        for key, value in mapping.items():
            # 直接列名映射或部分匹配
            if key in df.columns:
                rename_dict[key] = value
            else:
                # 尝试部分匹配（扁平化后的列名可能包含原列名）
                matching_cols = [col for col in df.columns if str(col).startswith(key) or key in str(col)]
                if matching_cols:
                    rename_dict[matching_cols[0]] = value
                    logger.info(f"列映射部分匹配: '{key}' -> '{matching_cols[0]}' -> '{value}'")

        if rename_dict:
            df = df.rename(columns=rename_dict)
            logger.info(f"应用列映射: {len(rename_dict)} 列")

        return df

    def _apply_conditional_extraction_in_step(self, df: pd.DataFrame, config: Dict[str, Any]) -> pd.DataFrame:
        """应用条件提取（在步骤中）"""
        import re

        condition = config.get("condition", {})
        extraction = config.get("extraction", {})
        output_type = extraction.get("output_type", "value")

        # 评估条件
        mask = self._evaluate_condition(df, condition)

        # 筛选满足条件的行
        filtered_df = df[mask].copy()
        logger.info(f"条件筛选: {len(df)} 行 -> {len(filtered_df)} 行")

        if output_type == "table":
            # 表格输出：返回筛选后的DataFrame，选择指定字段
            target_fields = extraction.get("target_fields", [])
            if target_fields:
                result_df = pd.DataFrame()

                for field in target_fields:
                    if isinstance(field, str):
                        # 简单字段名
                        if field in filtered_df.columns:
                            result_df[field] = filtered_df[field]
                    elif isinstance(field, dict):
                        # 多级字段：{"field": "期初余额", "sub_field": "借方", "output_field": "opening_balance_debit"}
                        main_field = field.get("field")
                        sub_field = field.get("sub_field")
                        output_field = field.get("output_field")

                        # 查找匹配的列名（扁平化后的格式：期初余额_借方）
                        col_pattern = f"{main_field}_{sub_field}"
                        matching_col = None
                        for col in filtered_df.columns:
                            if col_pattern in str(col):
                                matching_col = col
                                break

                        if matching_col:
                            result_df[output_field] = filtered_df[matching_col]
                            logger.info(f"提取多级字段: {main_field}.{sub_field} -> {output_field}")
                        else:
                            logger.warning(f"未找到多级字段: {main_field}.{sub_field}")

                return result_df

        return filtered_df

    def _evaluate_condition(self, df: pd.DataFrame, condition: Dict[str, Any]) -> pd.Series:
        """评估条件，返回布尔掩码"""
        import re

        cond_type = condition.get("type")

        if cond_type == "and":
            # AND 条件
            conditions = condition.get("conditions", [])
            mask = pd.Series([True] * len(df), index=df.index)
            for sub_cond in conditions:
                mask = mask & self._evaluate_condition(df, sub_cond)
            return mask

        elif cond_type == "or":
            # OR 条件
            conditions = condition.get("conditions", [])
            mask = pd.Series([False] * len(df), index=df.index)
            for sub_cond in conditions:
                mask = mask | self._evaluate_condition(df, sub_cond)
            return mask

        elif cond_type == "column_matches":
            # 列匹配条件（支持正则）
            column_header = condition.get("column_header")
            regex_pattern = condition.get("regex_pattern")

            if column_header not in df.columns:
                logger.warning(f"列不存在: {column_header}")
                return pd.Series([False] * len(df), index=df.index)

            # 应用正则匹配
            mask = df[column_header].astype(str).str.match(regex_pattern, na=False)
            return mask

        elif cond_type == "column_empty":
            # 列为空/非空条件
            column_header = condition.get("column_header")
            empty_check = condition.get("empty_check", True)

            if column_header not in df.columns:
                logger.warning(f"列不存在: {column_header}")
                return pd.Series([False] * len(df), index=df.index)

            if empty_check:
                # 检查为空
                mask = df[column_header].isna() | (df[column_header].astype(str).str.strip() == "")
            else:
                # 检查非空
                mask = df[column_header].notna() & (df[column_header].astype(str).str.strip() != "")

            return mask

        elif cond_type == "column_equals":
            # 列等于某值
            column_header = condition.get("column_header")
            value = condition.get("value")

            if column_header not in df.columns:
                logger.warning(f"列不存在: {column_header}")
                return pd.Series([False] * len(df), index=df.index)

            mask = df[column_header] == value
            return mask

        else:
            logger.warning(f"不支持的条件类型: {cond_type}")
            return pd.Series([True] * len(df), index=df.index)

    def _apply_template_matching(
        self,
        data_df: pd.DataFrame,
        template_df: pd.DataFrame,
        conditional_extractions: Dict[str, Any]
    ) -> pd.DataFrame:
        """
        应用模板匹配过滤

        Args:
            data_df: 从文件提取的数据
            template_df: 从模板读取的数据
            conditional_extractions: 条件提取配置

        Returns:
            过滤后的DataFrame
        """
        import pandas as pd

        match_config = conditional_extractions.get("match_with_template", {})
        if not match_config:
            return data_df

        template_fields = match_config.get("template_fields", [])
        data_fields = match_config.get("data_fields", [])
        match_type = match_config.get("match_type", "inner_join")

        if not template_fields or not data_fields:
            logger.warning("匹配字段未配置，返回原始数据")
            return data_df

        logger.info(f"应用模板匹配: template_fields={template_fields}, data_fields={data_fields}, match_type={match_type}")

        try:
            # 执行匹配（类似SQL JOIN）
            if match_type == "inner_join":
                # 内连接：只保留在模板中存在的数据
                merged_df = pd.merge(
                    data_df,
                    template_df[template_fields],
                    left_on=data_fields,
                    right_on=template_fields,
                    how="inner"
                )
                # 移除重复的匹配字段列
                for tf in template_fields:
                    if tf in merged_df.columns and tf not in data_fields:
                        merged_df.drop(columns=[tf], inplace=True)

                logger.info(f"匹配完成: 原始{len(data_df)}行 -> 匹配后{len(merged_df)}行")
                return merged_df

            elif match_type == "left_join":
                # 左连接：保留所有数据，标记是否在模板中
                merged_df = pd.merge(
                    data_df,
                    template_df[template_fields],
                    left_on=data_fields,
                    right_on=template_fields,
                    how="left"
                )
                return merged_df

            else:
                logger.warning(f"不支持的匹配类型: {match_type}")
                return data_df

        except Exception as e:
            logger.error(f"模板匹配失败: {str(e)}", exc_info=True)
            return data_df

    def _execute_template_action(
        self,
        step_config: Dict[str, Any],
        data: pd.DataFrame,
        working_file: str
    ):
        """
        执行模板操作（写入数据到模板）

        Args:
            step_config: 步骤配置
            data: 要写入的数据
            working_file: 工作文件路径
        """
        import pandas as pd
        import openpyxl
        from openpyxl.utils import get_column_letter, column_index_from_string

        template_action = step_config.get("template_action", {})
        if not template_action:
            logger.warning("未配置template_action，跳过写入")
            return

        action_type = template_action.get("action_type", "write_table")
        target = template_action.get("target", {})

        logger.info(f"执行模板操作: action_type={action_type}")

        # 加载工作簿
        wb = openpyxl.load_workbook(working_file)
        sheet_name = target.get("sheet", "Sheet1")

        if sheet_name not in wb.sheetnames:
            logger.error(f"工作表不存在: {sheet_name}")
            wb.close()
            return

        ws = wb[sheet_name]

        try:
            if action_type == "write_table":
                # 写入表格数据
                self._write_table_action(ws, target, data)

            elif action_type == "write_matched":
                # 匹配写入（根据key字段匹配已有数据并写入）
                self._write_matched_action(ws, target, data, working_file)

            elif action_type == "write_column":
                # 写入单列数据
                self._write_column_action(ws, target, data)

            elif action_type == "write_value":
                # 写入单个值
                self._write_value_action(ws, target, data)

            else:
                logger.warning(f"不支持的操作类型: {action_type}")

            # 保存工作簿
            wb.save(working_file)
            logger.info(f"模板操作完成，已保存: {working_file}")

        except Exception as e:
            logger.error(f"执行模板操作失败: {str(e)}", exc_info=True)
            raise
        finally:
            wb.close()

    def _write_table_action(self, ws, target: Dict[str, Any], data: pd.DataFrame):
        """写入表格数据"""
        from openpyxl.utils import column_index_from_string
        import re

        start_cell = target.get("start_cell", "A1")
        header_mapping = target.get("header_mapping", {})
        write_mode = target.get("write_mode", "overwrite")

        # 解析起始单元格
        match = re.match(r"([A-Z]+)(\d+)", start_cell)
        if not match:
            raise ValueError(f"无效的起始单元格: {start_cell}")

        start_col_letter = match.group(1)
        start_row = int(match.group(2))
        start_col = column_index_from_string(start_col_letter)

        logger.info(f"写入表格: start_cell={start_cell}, rows={len(data)}")

        # 如果有header_mapping，按映射写入
        if header_mapping:
            # header_mapping 格式: {"field_name": "A"}
            for row_idx, (_, df_row) in enumerate(data.iterrows(), start=start_row):
                for field_name, col_letter in header_mapping.items():
                    if field_name in df_row:
                        col_idx = column_index_from_string(col_letter)
                        value = df_row[field_name]
                        # 处理NaN值
                        if pd.isna(value):
                            value = None
                        ws.cell(row=row_idx, column=col_idx, value=value)
        else:
            # 按顺序写入所有列
            for row_idx, row_data in enumerate(data.itertuples(index=False), start=start_row):
                for col_offset, value in enumerate(row_data):
                    # 处理NaN值
                    if pd.isna(value):
                        value = None
                    ws.cell(row=row_idx, column=start_col + col_offset, value=value)

        logger.info(f"表格写入完成: {len(data)} 行")

    def _write_matched_action(self, ws, target: Dict[str, Any], data: pd.DataFrame, working_file: str):
        """匹配写入（根据key字段匹配已有数据并写入）"""
        from openpyxl.utils import column_index_from_string

        match_by = target.get("match_by", {})
        write_columns = target.get("write_columns", {})
        aggregation = target.get("aggregation", {})

        template_columns = match_by.get("template_columns", [])
        data_fields = match_by.get("data_fields", [])

        if not template_columns or not data_fields:
            logger.error("match_by配置不完整")
            return

        logger.info(f"匹配写入: template_columns={template_columns}, data_fields={data_fields}")

        # 读取模板中已有的数据（用于匹配）
        reader = TemplateReader(working_file)
        sheet_name = target.get("sheet", "Sheet1")

        # 确定读取范围（从第2行开始，读取到空行）
        # 假设模板列从A开始
        max_col_letter = max(template_columns + list(write_columns.values()))
        max_col_idx = column_index_from_string(max_col_letter)
        from openpyxl.utils import get_column_letter
        range_str = f"A2:{get_column_letter(max_col_idx)}1000"

        # 构建columns_mapping
        columns_mapping = {}
        for col_letter in template_columns:
            columns_mapping[col_letter] = f"key_{col_letter}"

        try:
            template_df = reader.read_range(
                sheet_name=sheet_name,
                range_str=range_str,
                columns_mapping=columns_mapping,
                read_until_empty=True
            )
        except Exception as e:
            logger.error(f"读取模板数据失败: {str(e)}")
            return

        # 构建匹配字典：{(key1, key2, ...): row_number}
        match_dict = {}
        for idx, row in template_df.iterrows():
            key = tuple(row.get(f"key_{col}", "") for col in template_columns)
            excel_row = idx + 2  # Excel行号（从2开始）
            match_dict[key] = excel_row

        logger.info(f"模板中有 {len(match_dict)} 行数据可供匹配")

        # 如果需要聚合，先进行聚合
        if aggregation:
            # 按data_fields分组聚合
            agg_dict = {}
            for field, agg_func in aggregation.items():
                if agg_func == "sum":
                    agg_dict[field] = "sum"
                elif agg_func == "count":
                    agg_dict[field] = "count"
                elif agg_func == "mean":
                    agg_dict[field] = "mean"
                else:
                    agg_dict[field] = "first"

            data = data.groupby(data_fields, as_index=False).agg(agg_dict)
            logger.info(f"聚合后数据: {len(data)} 行")

        # 匹配并写入
        matched_count = 0
        for _, row in data.iterrows():
            # 构建匹配key
            key = tuple(row.get(field, "") for field in data_fields)

            if key in match_dict:
                excel_row = match_dict[key]

                # 写入数据
                for field_name, col_letter in write_columns.items():
                    if field_name in row:
                        col_idx = column_index_from_string(col_letter)
                        value = row[field_name]
                        # 处理NaN值
                        if pd.isna(value):
                            value = None
                        ws.cell(row=excel_row, column=col_idx, value=value)

                matched_count += 1

        logger.info(f"匹配写入完成: {matched_count}/{len(data)} 行成功匹配")

    def _write_column_action(self, ws, target: Dict[str, Any], data: pd.DataFrame):
        """写入单列数据"""
        from openpyxl.utils import column_index_from_string
        import re

        start_cell = target.get("start_cell", "A1")
        field = target.get("field")

        if not field or field not in data.columns:
            logger.error(f"字段不存在: {field}")
            return

        # 解析起始单元格
        match = re.match(r"([A-Z]+)(\d+)", start_cell)
        if not match:
            raise ValueError(f"无效的起始单元格: {start_cell}")

        col_letter = match.group(1)
        start_row = int(match.group(2))
        col_idx = column_index_from_string(col_letter)

        logger.info(f"写入列: field={field}, start_cell={start_cell}, rows={len(data)}")

        # 写入数据
        for row_offset, value in enumerate(data[field]):
            # 处理NaN值
            if pd.isna(value):
                value = None
            ws.cell(row=start_row + row_offset, column=col_idx, value=value)

        logger.info(f"列写入完成: {len(data)} 行")

    def _write_value_action(self, ws, target: Dict[str, Any], data):
        """写入单个值"""
        cell = target.get("cell", "A1")
        value_source = target.get("value_source")

        if isinstance(data, pd.DataFrame) and len(data) > 0:
            # 如果data是DataFrame，取第一行第一列
            value = data.iloc[0, 0]
        else:
            value = data

        # 处理NaN值
        if pd.isna(value):
            value = None

        ws[cell] = value
        logger.info(f"写入单元格: {cell} = {value}")
