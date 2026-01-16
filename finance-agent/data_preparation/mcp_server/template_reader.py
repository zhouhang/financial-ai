"""
模板读取器 - 从Excel模板中读取已写入的数据
用于步骤化数据整理流程中，后续步骤读取前面步骤写入模板的数据
"""
import logging
import pandas as pd
import openpyxl
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from openpyxl.utils import get_column_letter, column_index_from_string

logger = logging.getLogger(__name__)


class TemplateReader:
    """从Excel模板中读取数据的组件"""

    def __init__(self, template_path: str):
        """
        初始化模板读取器

        Args:
            template_path: 模板文件路径
        """
        self.template_path = Path(template_path)
        if not self.template_path.exists():
            raise FileNotFoundError(f"模板文件不存在: {template_path}")

        logger.info(f"初始化模板读取器: {template_path}")

    def read_range(
        self,
        sheet_name: str,
        range_str: str,
        columns_mapping: Optional[Dict[str, str]] = None,
        read_until_empty: bool = False
    ) -> pd.DataFrame:
        """
        从模板中读取指定范围的数据

        Args:
            sheet_name: 工作表名称
            range_str: 单元格范围，如 "A2:E100"
            columns_mapping: 列映射，如 {"A": "customer_id", "B": "customer_name"}
            read_until_empty: 是否读取到空行为止（忽略range_str的结束行）

        Returns:
            包含读取数据的DataFrame
        """
        logger.info(f"读取模板范围: sheet={sheet_name}, range={range_str}")

        try:
            # 加载工作簿
            wb = openpyxl.load_workbook(self.template_path, data_only=True)

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表不存在: {sheet_name}")

            ws = wb[sheet_name]

            # 解析范围
            start_cell, end_cell = self._parse_range(range_str)
            start_col, start_row = start_cell
            end_col, end_row = end_cell

            # 读取数据
            if read_until_empty:
                data = self._read_until_empty(ws, start_col, start_row, end_col)
            else:
                data = self._read_fixed_range(ws, start_col, start_row, end_col, end_row)

            # 创建DataFrame
            df = pd.DataFrame(data)

            # 应用列映射
            if columns_mapping:
                df = self._apply_column_mapping(df, columns_mapping, start_col)

            logger.info(f"成功读取 {len(df)} 行数据")
            wb.close()
            return df

        except Exception as e:
            logger.error(f"读取模板失败: {str(e)}", exc_info=True)
            raise

    def _parse_range(self, range_str: str) -> Tuple[Tuple[int, int], Tuple[int, int]]:
        """
        解析Excel范围字符串

        Args:
            range_str: 范围字符串，如 "A2:E100"

        Returns:
            ((start_col, start_row), (end_col, end_row))
        """
        try:
            start, end = range_str.split(":")

            # 解析起始单元格
            start_col_letter = ''.join(filter(str.isalpha, start))
            start_row = int(''.join(filter(str.isdigit, start)))
            start_col = column_index_from_string(start_col_letter)

            # 解析结束单元格
            end_col_letter = ''.join(filter(str.isalpha, end))
            end_row = int(''.join(filter(str.isdigit, end)))
            end_col = column_index_from_string(end_col_letter)

            return (start_col, start_row), (end_col, end_row)

        except Exception as e:
            raise ValueError(f"无效的范围格式: {range_str}, 错误: {str(e)}")

    def _read_fixed_range(
        self,
        ws,
        start_col: int,
        start_row: int,
        end_col: int,
        end_row: int
    ) -> List[List[Any]]:
        """
        读取固定范围的单元格数据

        Args:
            ws: 工作表对象
            start_col: 起始列索引
            start_row: 起始行索引
            end_col: 结束列索引
            end_row: 结束行索引

        Returns:
            二维列表，包含单元格数据
        """
        data = []
        for row_idx in range(start_row, end_row + 1):
            row_data = []
            for col_idx in range(start_col, end_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                row_data.append(cell.value)
            data.append(row_data)

        return data

    def _read_until_empty(
        self,
        ws,
        start_col: int,
        start_row: int,
        end_col: int
    ) -> List[List[Any]]:
        """
        从起始位置读取数据，直到遇到空行为止

        Args:
            ws: 工作表对象
            start_col: 起始列索引
            start_row: 起始行索引
            end_col: 结束列索引

        Returns:
            二维列表，包含单元格数据
        """
        data = []
        row_idx = start_row
        max_empty_rows = 3  # 连续3行为空则停止

        empty_row_count = 0
        while empty_row_count < max_empty_rows:
            row_data = []
            is_empty_row = True

            for col_idx in range(start_col, end_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = cell.value
                row_data.append(value)

                if value is not None and str(value).strip() != "":
                    is_empty_row = False

            if is_empty_row:
                empty_row_count += 1
            else:
                empty_row_count = 0
                data.append(row_data)

            row_idx += 1

            # 安全限制：最多读取10000行
            if row_idx > start_row + 10000:
                logger.warning(f"读取行数超过10000，停止读取")
                break

        return data

    def _apply_column_mapping(
        self,
        df: pd.DataFrame,
        columns_mapping: Dict[str, str],
        start_col: int
    ) -> pd.DataFrame:
        """
        应用列名映射

        Args:
            df: 原始DataFrame
            columns_mapping: 列映射，如 {"A": "customer_id", "B": "customer_name"}
            start_col: 起始列索引

        Returns:
            应用映射后的DataFrame
        """
        # 构建新的列名列表
        new_columns = []
        for idx, old_col in enumerate(df.columns):
            # 计算当前列的Excel字母
            col_letter = get_column_letter(start_col + idx)

            # 如果有映射，使用映射的名称，否则保持原列名
            if col_letter in columns_mapping:
                new_columns.append(columns_mapping[col_letter])
            else:
                new_columns.append(f"col_{col_letter}")

        df.columns = new_columns
        return df

    def read_by_config(self, config: Dict[str, Any]) -> pd.DataFrame:
        """
        根据配置读取模板数据（用于步骤化schema）

        Args:
            config: 配置字典，包含 sheet, range, columns_mapping, read_until_empty 等

        Returns:
            读取的DataFrame
        """
        sheet_name = config.get("sheet")
        range_str = config.get("range")
        columns_mapping = config.get("columns_mapping")
        read_until_empty = config.get("read_until_empty", False)

        if not sheet_name or not range_str:
            raise ValueError("配置必须包含 sheet 和 range")

        return self.read_range(
            sheet_name=sheet_name,
            range_str=range_str,
            columns_mapping=columns_mapping,
            read_until_empty=read_until_empty
        )
