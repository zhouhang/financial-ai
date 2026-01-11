"""
模板写入器 - 将数据写入 Excel 模板
"""
import pandas as pd
import logging
import re
from pathlib import Path
from typing import Dict, Any
import openpyxl
from openpyxl.styles import Font
from datetime import datetime

logger = logging.getLogger(__name__)


class TemplateWriter:
    """模板写入器"""
    
    def __init__(self, schema: Dict):
        self.schema = schema
        self.template_mapping = schema.get("template_mapping", {})
    
    def write_to_template(
        self,
        template_path: str,
        output_path: str,
        calculation_results: Dict[str, Any],
        extracted_data: Dict[str, pd.DataFrame]
    ) -> str:
        """
        将数据写入 Excel 模板
        
        Args:
            template_path: 模板文件路径
            output_path: 输出文件路径
            calculation_results: 计算结果
            extracted_data: 提取的原始数据
        
        Returns:
            输出文件路径
        """
        # 加载模板
        try:
            wb = openpyxl.load_workbook(template_path)
            logger.info(f"加载模板成功: {template_path}")
        except Exception as e:
            logger.error(f"加载模板失败: {template_path}, 错误: {str(e)}")
            raise
        
        # 获取映射配置
        cell_mappings = self.template_mapping.get("cell_mappings", [])
        
        # 写入数据
        for mapping in cell_mappings:
            try:
                self._write_mapping(wb, mapping, calculation_results, extracted_data)
            except Exception as e:
                logger.error(f"写入映射失败: {mapping.get('target')}, 错误: {str(e)}")
        
        # 应用条件格式
        conditional_formats = self.template_mapping.get("conditional_formats", [])
        for cf in conditional_formats:
            try:
                self._apply_conditional_format(wb, cf)
            except Exception as e:
                logger.error(f"应用条件格式失败: {cf.get('range')}, 错误: {str(e)}")
        
        # 生成输出文件名
        # 如果配置了 output_file_pattern，使用它；否则使用模板文件名+日期后缀
        output_file_pattern = self.template_mapping.get("output_file_pattern")
        if output_file_pattern:
            output_filename = self._format_filename(output_file_pattern)
        else:
            # 使用模板文件名+日期后缀
            template_name = Path(template_path).stem  # 不带扩展名的文件名
            template_ext = Path(template_path).suffix  # 扩展名
            output_filename = f"{template_name}_{datetime.now().strftime('%Y%m%d')}{template_ext}"
        final_output_path = Path(output_path) / output_filename
        
        # 保存文件
        try:
            wb.save(str(final_output_path))
            logger.info(f"输出文件保存成功: {final_output_path}")
            return str(final_output_path)
        except Exception as e:
            logger.error(f"保存文件失败: {final_output_path}, 错误: {str(e)}")
            raise
    
    def _write_mapping(
        self,
        wb: openpyxl.Workbook,
        mapping: Dict,
        calculation_results: Dict[str, Any],
        extracted_data: Dict[str, pd.DataFrame]
    ):
        """写入单个映射"""
        data_source = mapping.get("data_source")
        field = mapping.get("field")
        target = mapping.get("target", {})
        
        sheet_name = target.get("sheet")
        cell_address = target.get("cell")
        range_address = target.get("range")
        mapping_type = target.get("type", "value")
        match_by = target.get("match_by")
        
        # 获取工作表
        if sheet_name not in wb.sheetnames:
            logger.warning(f"工作表 {sheet_name} 不存在")
            return
        
        ws = wb[sheet_name]
        
        # 获取数据
        if data_source == "calculation_result":
            value = calculation_results.get(field)
        elif data_source in extracted_data:
            # 如果是表格类型，直接使用 DataFrame
            if mapping_type == "table" and field == "extracted_data":
                value = extracted_data[data_source]
            else:
                value = extracted_data[data_source]
        else:
            logger.warning(f"数据源 {data_source} 不存在")
            return
        
        # 写入数据
        if mapping_type == "value" and cell_address:
            # 单元格值
            ws[cell_address] = value
            self._apply_format(ws[cell_address], target.get("format", {}))
            logger.info(f"写入单元格: {sheet_name}!{cell_address} = {value}")
        
        elif mapping_type == "table" and range_address:
            # 表格数据
            if isinstance(value, pd.DataFrame):
                if match_by:
                    # 需要匹配写入：通过指定字段匹配已写入的数据
                    self._write_table_with_match(
                        ws, range_address, value, target.get("header_mapping", {}),
                        match_by, extracted_data
                    )
                else:
                    # 直接写入
                    self._write_table(ws, range_address, value, target.get("header_mapping", {}))
            else:
                logger.warning(f"表格数据类型错误: {type(value)}")
    
    def _write_table(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        range_address: str,
        df: pd.DataFrame,
        header_mapping: Dict
    ):
        """写入表格数据"""
        # 解析范围 "B9:F100"
        start_cell, end_cell = range_address.split(":")
        start_col_letter = re.match(r"([A-Z]+)", start_cell).group(1)
        start_col = openpyxl.utils.column_index_from_string(start_col_letter)
        start_row = int(re.match(r"[A-Z]+(\d+)", start_cell).group(1))
        
        # 如果 header_mapping 的值是列字母（如 "B", "C"），需要转换为列索引并重新排列
        if header_mapping and any(isinstance(v, str) and len(v) == 1 and v.isalpha() for v in header_mapping.values()):
            # header_mapping 格式: {"field_name": "B"}，表示字段写入到B列
            # 按列字母顺序排列字段
            col_order = sorted(header_mapping.items(), key=lambda x: openpyxl.utils.column_index_from_string(x[1]))
            
            # 写入数据，按列映射写入
            for row_idx, (_, df_row) in enumerate(df.iterrows(), start=start_row):
                for field_name, col_letter in col_order:
                    if field_name in df_row:
                        col_idx = openpyxl.utils.column_index_from_string(col_letter)
                        value = df_row[field_name]
                        ws.cell(row=row_idx, column=col_idx, value=value)
        else:
            # 原有的连续写入方式
            # 写入数据
            for row_idx, row in enumerate(df.itertuples(index=False), start=start_row):
                for col_idx, value in enumerate(row, start=start_col):
                    ws.cell(row=row_idx, column=col_idx, value=value)
        
        logger.info(f"写入表格: {range_address}, {len(df)} 行")
    
    def _write_table_with_match(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        range_address: str,
        df: pd.DataFrame,
        header_mapping: Dict,
        match_by: Dict,
        extracted_data: Dict[str, pd.DataFrame]
    ):
        """
        通过匹配字段写入表格数据
        用于将 source_2 的数据匹配到 source_1 已写入的行
        """
        # 解析范围 "H9:H100"
        start_cell, end_cell = range_address.split(":")
        start_col_letter = re.match(r"([A-Z]+)", start_cell).group(1)
        start_col = openpyxl.utils.column_index_from_string(start_col_letter)
        start_row = int(re.match(r"[A-Z]+(\d+)", start_cell).group(1))
        
        # 获取匹配字段
        match_fields = match_by.get("fields", [])
        source_1_fields = match_by.get("source_1_fields", match_fields)
        
        # 获取 source_1 的数据（应该已经写入）
        source_1_id = "source_1"
        if source_1_id not in extracted_data:
            logger.warning(f"无法找到 source_1 数据进行匹配")
            return
        
        source_1_df = extracted_data[source_1_id]
        
        # 构建匹配键的字典：{("科目名称", "核算项目"): row_index_in_excel}
        match_dict = {}
        for idx, row in source_1_df.iterrows():
            key = tuple(row.get(field, "") for field in source_1_fields if field in row)
            match_dict[key] = start_row + idx  # Excel 行号（从start_row开始）
        
        # 获取要写入的字段和列
        if header_mapping:
            # header_mapping 格式: {"prior_period_credit": "H"}
            write_field = list(header_mapping.keys())[0]
            write_col_letter = list(header_mapping.values())[0]
            write_col = openpyxl.utils.column_index_from_string(write_col_letter)
        else:
            write_field = df.columns[0] if len(df.columns) > 0 else None
            write_col = start_col
        
        # 匹配并写入数据
        matched_count = 0
        for idx, row in df.iterrows():
            # 构建当前行的匹配键
            key = tuple(row.get(field, "") for field in match_fields if field in row)
            
            # 查找匹配的行
            if key in match_dict:
                excel_row = match_dict[key]
                # 写入数据到对应列的匹配行
                value = row.get(write_field)
                if value is not None:
                    ws.cell(row=excel_row, column=write_col, value=value)
                    matched_count += 1
        
        logger.info(f"匹配写入表格: {range_address}, 匹配 {matched_count}/{len(df)} 行到 source_1 的数据")
    
    def _apply_format(self, cell, format_config: Dict):
        """应用单元格格式"""
        if not format_config:
            return
        
        # 数字格式
        number_format = format_config.get("number_format")
        if number_format:
            cell.number_format = number_format
        
        # 字体
        font_config = format_config.get("font", {})
        if font_config:
            cell.font = Font(
                bold=font_config.get("bold", False),
                color=font_config.get("color")
            )
    
    def _apply_conditional_format(self, wb: openpyxl.Workbook, cf_config: Dict):
        """应用条件格式"""
        from openpyxl.formatting.rule import DataBarRule
        from openpyxl.styles import Color
        
        range_address = cf_config.get("range", "")  # 如 "汇总表!C5:C10"
        cf_type = cf_config.get("type", "")
        color = cf_config.get("color", "63BE7B")
        
        # 解析 sheet!range 格式
        if "!" in range_address:
            sheet_name, cell_range = range_address.split("!", 1)
        else:
            logger.warning(f"条件格式范围格式错误: {range_address}")
            return
        
        # 获取工作表
        if sheet_name not in wb.sheetnames:
            logger.warning(f"工作表 {sheet_name} 不存在")
            return
        
        ws = wb[sheet_name]
        
        # 应用条件格式
        if cf_type == "data_bar":
            rule = DataBarRule(
                start_type='min',
                end_type='max',
                color=color
            )
            ws.conditional_formatting.add(cell_range, rule)
            logger.info(f"应用条件格式 data_bar: {range_address}")
        else:
            logger.warning(f"不支持的条件格式类型: {cf_type}")
    
    @staticmethod
    def _format_filename(pattern: str) -> str:
        """格式化文件名"""
        now = datetime.now()
        replacements = {
            "{YYYY}": now.strftime("%Y"),
            "{MM}": now.strftime("%m"),
            "{DD}": now.strftime("%d"),
            "{YYYYMMDD}": now.strftime("%Y%m%d"),
            "{HH}": now.strftime("%H"),
            "{mm}": now.strftime("%M"),
            "{SS}": now.strftime("%S"),
            "{timestamp}": now.strftime("%Y%m%d_%H%M%S")
        }
        
        result = pattern
        for key, value in replacements.items():
            result = result.replace(key, value)
        
        return result
